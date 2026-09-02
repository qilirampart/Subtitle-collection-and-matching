from __future__ import annotations

from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait
import os
import shutil
import subprocess
import sys
import tempfile
from threading import Lock, local
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qs, urlsplit
from pathlib import Path

from PySide6.QtCore import QLockFile, QThread, QTimer, Qt, QUrl, Signal
from PySide6.QtGui import QDesktopServices
from PySide6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QProgressDialog,
    QProgressBar,
    QScrollArea,
    QSpinBox,
    QStackedWidget,
    QSplitter,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from app.models import YouTubeVideo
from app.config.settings import YOUTUBE_PROXY_CONFIG_PATH
from app.services.matching_api import (
    DramaSubtitleMatchingClient,
    MatchingServiceConfig,
    build_matching_result_rows,
)
from app.services.youtube_cookie_service import YouTubeCookieSyncError, sync_browser_cookie_file
from app.services.youtube_cover_service import CoverDownloadResult, YouTubeCoverService
from app.services.youtube_cover_review_service import CoverReviewResult, YouTubeCoverReviewService
from app.services.subtitle_excel_exporter import export_subtitles_to_xlsx
from app.services.review_excel_exporter import (
    export_cover_review_results_to_xlsx,
    export_matching_results_to_xlsx,
)
from app.services.proxy_discovery_service import ProxyDiscoveryService
from app.services.api_config_service import ApiConfigService
from app.services.update_service import ApplicationUpdateService, AvailableUpdate
from app.services.youtube_audio_service import YouTubeAudioService
from app.services.youtube_service import YouTubeDownloadCancelled, YouTubeService
from app.services.youtube_asr import YouTubeAsrService
from app.services.task_state import TaskStateStore, has_recoverable_work
from app.settings import COVER_DIR, PROJECT_ROOT, RUNTIME_DIR
from app.workflow import VerificationWorkflow
from app.ui.asr_config_dialog import AsrConfigDialog
from app.ui.llm_config_dialog import LlmConfigDialog
from app.ui.matching_config_dialog import MatchingConfigDialog
from app.ui.youtube_login_dialog import YouTubeLoginDialog
from app.ui.browser_audio_capture_dialog import BrowserAudioCaptureDialog
from app.ui.theme import apply_app_theme, load_app_icon
from app.ui.window_geometry import apply_responsive_window_geometry
from app.ui.workspace_pages import (
    DashboardPage,
    CoverPage,
    DownloadPage,
    MatchingPage,
    SettingsHubPage,
    SubtitlePage,
    TaskCenterPage,
)
from app.task_control import TaskControl
from app.utils.logger import get_logger
from app.version import APP_VERSION


LOGGER = get_logger(__name__)


class ExportDestinationDialog(QDialog):
    """Make the default export location and filename visible before writing."""

    def __init__(self, title: str, default_path: Path, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(620)
        self._output_path: Path | None = None

        layout = QVBoxLayout(self)
        hint = QLabel("默认会保存到软件的 output 目录。可在导出前修改文件夹或文件名。")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        form = QFormLayout()
        self.directory_input = QLineEdit(str(default_path.parent))
        self.file_name_input = QLineEdit(default_path.name)
        choose_folder_button = QPushButton("选择文件夹")
        choose_folder_button.clicked.connect(self._choose_directory)
        directory_layout = QHBoxLayout()
        directory_layout.addWidget(self.directory_input, 1)
        directory_layout.addWidget(choose_folder_button)
        form.addRow("保存位置", directory_layout)
        form.addRow("文件名称", self.file_name_input)
        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Save
        )
        buttons.rejected.connect(self.reject)
        buttons.accepted.connect(self._confirm)
        layout.addWidget(buttons)

    @property
    def output_path(self) -> Path | None:
        return self._output_path

    def _choose_directory(self) -> None:
        directory = QFileDialog.getExistingDirectory(
            self,
            "选择导出文件夹",
            self.directory_input.text().strip(),
        )
        if directory:
            self.directory_input.setText(directory)

    def _confirm(self) -> None:
        directory_text = self.directory_input.text().strip()
        file_name = self.file_name_input.text().strip()
        if not directory_text or not file_name:
            QMessageBox.warning(self, "缺少导出信息", "请填写保存位置和文件名称。")
            return
        output_path = Path(directory_text).expanduser() / file_name
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            QMessageBox.warning(self, "无法创建文件夹", str(exc))
            return
        if output_path.exists():
            answer = QMessageBox.question(
                self,
                "确认覆盖",
                f"文件已存在，是否覆盖？\n{output_path}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if answer != QMessageBox.StandardButton.Yes:
                return
        self._output_path = output_path
        self.accept()


class _CollectThread(QThread):
    succeeded = Signal(object)
    failed = Signal(str)

    def __init__(self, url: str, limit: int) -> None:
        super().__init__()
        self.url = url
        self.limit = limit

    def run(self) -> None:
        try:
            self.succeeded.emit(VerificationWorkflow().collector.collect_channel(self.url, max_items=self.limit))
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _CoverCollectThread(QThread):
    """Collect multiple channel video lists for the independent cover page."""

    progress = Signal(int, int, str, int, object)
    batch = Signal(object)
    succeeded = Signal(object)
    failed = Signal(str)

    def __init__(self, urls: list[str], limit: int, control: TaskControl | None = None) -> None:
        super().__init__()
        self.urls = urls
        self.limit = limit
        self.control = control

    def run(self) -> None:
        collected: list[dict[str, object]] = []
        seen_ids: set[str] = set()
        collector = VerificationWorkflow().collector
        total = len(self.urls)
        for index, url in enumerate(self.urls, start=1):
            if self.control is not None and not self.control.checkpoint():
                break
            try:
                videos = collector.collect_channel(url, max_items=self.limit)
                added = 0
                for video in videos:
                    if video.video_id in seen_ids:
                        continue
                    seen_ids.add(video.video_id)
                    collected.append({
                        "video": video.to_dict(),
                        "collection_status": "success",
                        "cover_status": "pending",
                        "review_status": "pending",
                    })
                    added += 1
                if added:
                    self.batch.emit(collected[-added:])
                self.progress.emit(index, total, url, added, "")
            except Exception as exc:  # noqa: BLE001
                self.progress.emit(index, total, url, 0, str(exc))
        self.succeeded.emit(collected)


class _UpdateCheckThread(QThread):
    completed = Signal(object)
    failed = Signal(str)

    def run(self) -> None:
        try:
            self.completed.emit(ApplicationUpdateService().check_for_update(APP_VERSION))
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _UpdateDownloadThread(QThread):
    completed = Signal(object)
    failed = Signal(str)
    progress = Signal(int, int)

    def __init__(self, update: AvailableUpdate) -> None:
        super().__init__()
        self.update = update

    def run(self) -> None:
        try:
            archive = ApplicationUpdateService().download_update(
                self.update,
                progress_callback=lambda received, total: self.progress.emit(received, total),
            )
            self.completed.emit(archive)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _CoverThread(QThread):
    succeeded = Signal(object)
    cancelled = Signal(object)
    failed = Signal(str)
    started = Signal(int, int, object)
    progress = Signal(int, int, object, object)

    def __init__(self, videos: list[YouTubeVideo], control: TaskControl) -> None:
        super().__init__()
        self.videos = videos
        self.control = control

    def run(self) -> None:
        try:
            results, was_cancelled = YouTubeCoverService().download_batch(
                self.videos,
                task_control=self.control,
                started_callback=lambda index, total, video: self.started.emit(index, total, video),
                progress_callback=lambda index, total, video, result: self.progress.emit(
                    index, total, video, result
                ),
            )
            (self.cancelled if was_cancelled else self.succeeded).emit(results)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _CoverReviewThread(QThread):
    succeeded = Signal(object)
    cancelled = Signal(object)
    failed = Signal(str)
    progress = Signal(int, int, object, object)

    def __init__(self, videos: list[YouTubeVideo], cover_paths: dict[str, str], control: TaskControl, *, prefer_url: bool = False, concurrency: int | None = None) -> None:
        super().__init__()
        self.videos = videos
        self.cover_paths = cover_paths
        self.control = control
        self.prefer_url = prefer_url
        self.concurrency = concurrency

    def run(self) -> None:
        try:
            results, was_cancelled = YouTubeCoverReviewService().review_batch(
                self.videos,
                self.cover_paths,
                task_control=self.control,
                progress_callback=lambda index, total, video, result: self.progress.emit(
                    index, total, video, result
                ),
                prefer_url=self.prefer_url,
                concurrency=self.concurrency,
            )
            (self.cancelled if was_cancelled else self.succeeded).emit(results)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _PrepareThread(QThread):
    succeeded = Signal(object, object)
    failed = Signal(str)
    progress = Signal(int, int, object, object)
    stage = Signal(str)
    cancelled = Signal(object, object)

    def __init__(
        self,
        videos: list[YouTubeVideo],
        seconds: int,
        control: TaskControl,
        *,
        allow_asr_fallback: bool = False,
        skip_caption_probe: bool = False,
        caption_concurrency: int = 1,
        download_concurrency: int = 1,
        asr_concurrency: int = 1,
    ) -> None:
        super().__init__()
        self.videos = videos
        self.seconds = seconds
        self.control = control
        self.allow_asr_fallback = allow_asr_fallback
        self.skip_caption_probe = skip_caption_probe
        self.caption_concurrency = caption_concurrency
        self.download_concurrency = download_concurrency
        self.asr_concurrency = asr_concurrency

    def run(self) -> None:
        try:
            ready, pending_asr = VerificationWorkflow().prepare_batch_items(
                self.videos,
                leading_seconds=self.seconds,
                allow_asr_fallback=self.allow_asr_fallback,
                skip_caption_probe=self.skip_caption_probe,
                caption_concurrency=self.caption_concurrency,
                download_concurrency=self.download_concurrency,
                asr_concurrency=self.asr_concurrency,
                progress_callback=lambda index, total, video, inspection: self.progress.emit(
                    index, total, video, inspection
                ),
                stage_callback=self.stage.emit,
                task_control=self.control,
            )
            (self.cancelled if self.control.cancelled else self.succeeded).emit(ready, pending_asr)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _AsrFallbackThread(QThread):
    succeeded = Signal(object, object)
    failed = Signal(str)
    progress = Signal(int, int, object, object)
    stage = Signal(str)
    cancelled = Signal(object, object)

    def __init__(self, pending_items: list[dict[str, object]], seconds: int, control: TaskControl, audio_sources: dict[str, str] | None = None, download_concurrency: int = 1, asr_concurrency: int = 1) -> None:
        super().__init__()
        self.pending_items = pending_items
        self.seconds = seconds
        self.audio_sources = audio_sources or {}
        self.control = control
        self.download_concurrency = download_concurrency
        self.asr_concurrency = asr_concurrency

    def run(self) -> None:
        try:
            ready, still_pending = VerificationWorkflow().prepare_asr_fallback_items(
                self.pending_items,
                leading_seconds=self.seconds,
                audio_sources=self.audio_sources,
                download_concurrency=self.download_concurrency,
                asr_concurrency=self.asr_concurrency,
                task_control=self.control,
                progress_callback=lambda index, total, video, inspection: self.progress.emit(
                    index, total, video, inspection
                ),
                stage_callback=self.stage.emit,
            )
            (self.cancelled if self.control.cancelled else self.succeeded).emit(ready, still_pending)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class _MatchThread(QThread):
    succeeded = Signal(object)
    failed = Signal(str)
    task_created = Signal(str)
    progress = Signal(object)

    def __init__(
        self,
        items: list[dict[str, object]],
        server: str,
        username: str,
        password: str,
        top_k: int = 10,
    ) -> None:
        super().__init__()
        self.items = items
        self.server = server
        self.username = username
        self.password = password
        self.top_k = max(1, min(int(top_k or 10), 20))
        self.task_id = "video-compare"
        self._control_lock = Lock()
        self._control_request = ""
        self._paused = False
        self._cancelled = False

    def run(self) -> None:
        try:
            self._run_batch_task()
            return
            # Verify credentials once before scheduling the independent video requests.
            probe = DramaSubtitleMatchingClient(MatchingServiceConfig(self.server))
            probe.login(self.username, self.password)
            self.task_created.emit(self.task_id)

            total = len(self.items)
            completed = 0
            failed = 0
            next_index = 0
            cancelled = False
            results: list[dict[str, object] | None] = [None] * total
            pending: dict[object, int] = {}
            clients = local()

            def compare_one(index: int) -> dict[str, object]:
                client = getattr(clients, "client", None)
                if client is None:
                    client = DramaSubtitleMatchingClient(MatchingServiceConfig(self.server))
                    client.login(self.username, self.password)
                    clients.client = client
                item = self.items[index]
                response = client.video_compare(
                    str(item.get("query_text") or ""),
                    cues=list(item.get("cues") or []),
                    language_code=str(item.get("matching_language_code") or item.get("source_caption_language") or ""),
                    top_k=self.top_k,
                    translation_fallback=True,
                )
                return self._build_video_result_item(item, response)

            with ThreadPoolExecutor(max_workers=min(3, max(1, total)), thread_name_prefix="video-compare") as executor:
                while next_index < total or pending:
                    while next_index < total and len(pending) < 3:
                        if not self._checkpoint():
                            cancelled = True
                            break
                        pending[executor.submit(compare_one, next_index)] = next_index
                        next_index += 1
                    if cancelled:
                        for future in pending:
                            future.cancel()
                    if not pending:
                        break
                    finished, _ = wait(tuple(pending), return_when=FIRST_COMPLETED)
                    for future in finished:
                        index = pending.pop(future)
                        try:
                            results[index] = future.result()
                            completed += 1
                        except Exception as exc:  # noqa: BLE001
                            failed += 1
                            results[index] = self._build_video_error_item(self.items[index], exc)
                        self._emit_video_progress(completed, failed, total, cancelled)

            status = "cancelled" if cancelled else "partial_failed" if failed else "completed"
            self.succeeded.emit(
                {
                    "task": {
                        "status": status,
                        "accepted_input_count": total,
                        "completed_input_count": completed,
                        "failed_input_count": failed,
                        "progress_scope": "video",
                    },
                    "items": [item for item in results if item is not None],
                }
            )
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))

    def _run_batch_task(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig(self.server))
        client.login(self.username, self.password)
        created = client.submit_batch(self.items, top_k=self.top_k)
        task_id = str(created.get("task_id") or "").strip()
        if not task_id:
            raise RuntimeError("匹配服务未返回任务 ID。")
        self.task_id = task_id
        self.task_created.emit(task_id)

        while True:
            action = self._take_control_request()
            if action == "pause":
                client.pause_task(task_id)
            elif action == "resume":
                client.resume_task(task_id)
            elif action == "cancel":
                client.cancel_task(task_id)

            detail = client.task_detail(task_id)
            self.progress.emit(detail)
            task = detail.get("task") if isinstance(detail.get("task"), dict) else {}
            status = str(task.get("status") or "").lower()
            if status in {"completed", "partial_failed", "failed", "cancelled", "canceled"}:
                self.succeeded.emit(detail)
                return
            # Keep polling a paused task so the existing Continue button can
            # resume the same service-side task without resubmitting it.
            self.msleep(2000)

    def _take_control_request(self) -> str:
        with self._control_lock:
            action = self._control_request
            self._control_request = ""
        return action

    def request_control(self, action: str) -> None:
        if action not in {"pause", "resume", "cancel"}:
            return
        with self._control_lock:
            self._control_request = action

    def _checkpoint(self) -> bool:
        """Apply user controls between video requests; an in-flight request finishes safely."""
        while True:
            self._apply_control_request()
            if self._cancelled:
                return False
            if not self._paused:
                return True
            self.msleep(100)

    def _apply_control_request(self) -> None:
        with self._control_lock:
            action = self._control_request
            self._control_request = ""
        if action == "pause":
            self._paused = True
        elif action == "resume":
            self._paused = False
        elif action == "cancel":
            self._cancelled = True
            self._paused = False

    def _emit_video_progress(self, completed: int, failed: int, total: int, cancelled: bool) -> None:
        self.progress.emit(
            {
                "task": {
                    "status": "cancel_requested" if cancelled else "running",
                    "accepted_input_count": total,
                    "completed_input_count": completed,
                    "failed_input_count": failed,
                    "progress_scope": "video",
                }
            }
        )

    @staticmethod
    def _build_video_result_item(item: dict[str, object], response: dict[str, object]) -> dict[str, object]:
        payload = response.get("payload") if isinstance(response.get("payload"), dict) else response
        payload = payload if isinstance(payload, dict) else {}
        decision = payload.get("video_decision") if isinstance(payload.get("video_decision"), dict) else {}
        execution = payload.get("execution") if isinstance(payload.get("execution"), dict) else {}
        segments = payload.get("segments") if isinstance(payload.get("segments"), list) else []
        segment_order = decision.get("matched_segment_order") if isinstance(decision, dict) else None
        matched_segment = next(
            (segment for segment in segments if isinstance(segment, dict) and segment.get("segment_order") == segment_order),
            {},
        )
        fast_screen = payload.get("fast_screen") if isinstance(payload.get("fast_screen"), dict) else {}
        candidate_source = matched_segment if matched_segment else fast_screen
        translation_candidates = payload.get("translation_candidates") if isinstance(payload.get("translation_candidates"), list) else []
        candidates = translation_candidates or (candidate_source.get("candidates") if isinstance(candidate_source.get("candidates"), list) else [])
        result_payload = {
            "decision": decision,
            "candidates": candidates,
            "execution": execution,
            "translation_fallback": payload.get("translation_fallback") if isinstance(payload.get("translation_fallback"), dict) else {},
            "translation_candidates": translation_candidates,
            "translated_query_text": str(payload.get("translated_query_text") or ""),
            "review_feedback": payload.get("review_feedback") if isinstance(payload.get("review_feedback"), dict) else {},
        }
        result = dict(item)
        result.update(
            {
                "source_segment_order": segment_order or 1,
                "source_time_start": str(matched_segment.get("source_time_start") or item.get("source_time_start") or "0"),
                "source_time_end": str(matched_segment.get("source_time_end") or item.get("source_time_end") or "0"),
                "query_text": str(matched_segment.get("query_text") or item.get("query_text") or ""),
                "source_text_original": str(matched_segment.get("query_text") or item.get("source_text_original") or item.get("query_text") or ""),
                "result_payload_json": result_payload,
            }
        )
        return result

    @staticmethod
    def _build_video_error_item(item: dict[str, object], error: Exception) -> dict[str, object]:
        result = dict(item)
        result["result_payload_json"] = {
            "decision": {
                "matched": False,
                "status": "review_required",
                "outcome": "potential_match",
                "content_match_status": "uncertain",
                "title_resolution": "unresolved",
                "user_message": f"视频级匹配请求失败：{type(error).__name__}",
                "reason": "video_compare_request_failed",
            },
            "candidates": [],
        }
        return result


class _EvidenceContextThread(QThread):
    succeeded = Signal(str, str)
    failed = Signal(str)

    def __init__(self, server: str, username: str, password: str, window_uid: str) -> None:
        super().__init__()
        self.server = server
        self.username = username
        self.password = password
        self.window_uid = window_uid

    def run(self) -> None:
        try:
            client = DramaSubtitleMatchingClient(MatchingServiceConfig(self.server))
            client.login(self.username, self.password)
            payload = client.evidence_context(self.window_uid)
            context: object = payload.get("context")
            for _ in range(2):
                if isinstance(context, dict) and isinstance(context.get("context"), dict):
                    context = context["context"]
                else:
                    break
            if not isinstance(context, dict):
                raise RuntimeError("Evidence context response is invalid.")
            text = str(context.get("text") or "").strip()
            if not text:
                raise RuntimeError("Evidence context contains no subtitle text.")
            LOGGER.info("Matching evidence context loaded. window_uid=%s chars=%s", self.window_uid, len(text))
            self.succeeded.emit(self.window_uid, text)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Matching evidence context failed. error_type=%s", type(exc).__name__)
            self.failed.emit(str(exc))


class _StandaloneDownloadThread(QThread):
    """Run the existing video/audio download services without entering the full workflow."""

    item_updated = Signal(int, str, str)
    completed = Signal(object, bool)
    failed = Signal(str)

    def __init__(
        self,
        urls: list[str],
        *,
        content_kind: str,
        duration_seconds: int,
        audio_concurrency: int,
        control: TaskControl,
    ) -> None:
        super().__init__()
        self.urls = urls
        self.content_kind = content_kind
        self.duration_seconds = duration_seconds
        self.audio_concurrency = audio_concurrency
        self.control = control

    def run(self) -> None:
        results: list[dict[str, object]] = []
        try:
            for index, url in enumerate(self.urls):
                if not self.control.checkpoint():
                    break
                self.item_updated.emit(index, "正在准备下载", "")
                output_paths: list[str] = []
                try:
                    if self.content_kind in {"video", "both"}:
                        video_result = YouTubeService().download_video(
                            url,
                            max_duration_seconds=self.duration_seconds or None,
                            progress_callback=lambda current, total, row=index: self._emit_video_progress(
                                row, current, total
                            ),
                            should_cancel=lambda: not self.control.checkpoint(),
                        )
                        output_paths.append(video_result.local_path)
                    if self.content_kind in {"audio", "both"}:
                        audio_result = YouTubeAudioService().download_audio(
                            url,
                            max_duration_seconds=self.duration_seconds,
                            concurrency=self.audio_concurrency,
                            progress_callback=lambda current, total, row=index: self.item_updated.emit(
                                row, f"正在下载音频（分段 {current}/{total}）", ""
                            ),
                            should_cancel=lambda: not self.control.checkpoint(),
                        )
                        output_paths.append(audio_result.local_path)
                    self.item_updated.emit(index, "下载完成", "\n".join(output_paths))
                    results.append({"url": url, "ok": True, "paths": output_paths})
                except YouTubeDownloadCancelled:
                    break
                except Exception as exc:  # noqa: BLE001
                    message = str(exc)
                    self.item_updated.emit(index, "下载失败", message)
                    results.append({"url": url, "ok": False, "error": message})
            self.completed.emit(results, self.control.cancelled)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))

    def _emit_video_progress(self, row: int, current: int, total: int) -> None:
        if total > 0:
            percent = min(100, int(current * 100 / total))
            self.item_updated.emit(row, f"正在下载视频（{percent}%）", "")
        else:
            self.item_updated.emit(row, "正在下载视频", "")


class _StandaloneSubtitleThread(QThread):
    """Acquire direct captions or transcribe local/remote media independently."""

    item_updated = Signal(int, str, str, str)
    completed = Signal(object, object, bool)
    failed = Signal(str)

    def __init__(
        self,
        sources: list[str],
        seconds: int,
        allow_asr_fallback: bool,
        control: TaskControl,
        skip_caption_probe: bool = False,
    ) -> None:
        super().__init__()
        self.sources = sources
        self.seconds = seconds
        self.allow_asr_fallback = allow_asr_fallback
        self.skip_caption_probe = skip_caption_probe
        self.control = control

    def run(self) -> None:
        ready: list[dict[str, object]] = []
        pending: list[dict[str, object]] = []
        workflow = VerificationWorkflow()
        asr_service = YouTubeAsrService()
        try:
            for index, source in enumerate(self.sources):
                if not self.control.checkpoint():
                    break
                self.item_updated.emit(index, "处理中", "", "")
                try:
                    if YouTubeService.is_youtube_url(source):
                        video = self._video_from_url(source)
                        inspection = (
                            workflow.asr_required_inspection(video, self.seconds)
                            if self.skip_caption_probe
                            else workflow.inspect_video(video, leading_seconds=self.seconds)
                        )
                        if inspection.get("status") == "asr_required" and self.allow_asr_fallback:
                            inspection = workflow._try_asr_fallback(  # noqa: SLF001
                                video,
                                dict(inspection),
                                leading_seconds=self.seconds,
                            )
                        if inspection.get("status") == "asr_required":
                            pending.append({"video": video.to_dict(), "inspection": inspection})
                            self.item_updated.emit(index, "待 ASR 兜底", str(inspection.get("source_kind") or "无字幕"), "")
                            continue
                        items = workflow._build_batch_items(video, inspection)  # noqa: SLF001
                        ready.extend(items)
                        text = str(inspection.get("text") or "")
                        self.item_updated.emit(index, "识别完成", str(inspection.get("source_kind") or "直出字幕"), text)
                        continue

                    path = Path(source)
                    if not path.is_file():
                        raise RuntimeError("本地素材不存在。")
                    transcript = asr_service.transcribe_audio_source(str(path))
                    video = YouTubeVideo(
                        video_id=path.stem,
                        source_url=str(path),
                        title=path.stem,
                    )
                    inspection = {
                        "video": video.to_dict(),
                        "language_code": "",
                        "source_kind": "asr",
                        "source_path": transcript.source_path,
                        "text": transcript.text,
                        "normalized_text": workflow.normalizer.normalize(transcript.text),
                        "start_seconds": 0,
                        "end_seconds": self.seconds,
                        "asr_required": False,
                        "status": "ready_for_matching",
                    }
                    ready.extend(workflow._build_batch_items(video, inspection))  # noqa: SLF001
                    self.item_updated.emit(index, "识别完成", "本地 ASR", transcript.text)
                except Exception as exc:  # noqa: BLE001
                    self.item_updated.emit(index, "识别失败", "", str(exc))
            self.completed.emit(ready, pending, self.control.cancelled)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))

    @staticmethod
    def _video_from_url(source: str) -> YouTubeVideo:
        parts = urlsplit(source)
        video_id = (parse_qs(parts.query).get("v") or [""])[0]
        if not video_id and parts.path.startswith("/shorts/"):
            video_id = parts.path.removeprefix("/shorts/").strip("/")
        if not video_id and parts.netloc.lower() == "youtu.be":
            video_id = parts.path.strip("/")
        if not video_id:
            raise RuntimeError("无法从链接读取 YouTube 视频 ID。")
        return YouTubeVideo(video_id=video_id, source_url=source, title=video_id)


class _StandaloneCoverThread(QThread):
    """Download cover assets and optionally run the existing cover review model."""

    item_updated = Signal(int, str, str, object)
    completed = Signal(object, bool)
    failed = Signal(str)

    def __init__(self, items: list[dict[str, object]], download_cover: bool, detect_cover: bool, control: TaskControl, *, download_concurrency: int = 1) -> None:
        super().__init__()
        self.items = items
        self.download_cover = download_cover
        self.detect_cover = detect_cover
        self.control = control
        self.download_concurrency = download_concurrency

    def run(self) -> None:
        results: list[dict[str, object]] = []
        try:
            cover_service = YouTubeCoverService()
            review_service = YouTubeCoverReviewService()
            if self.download_cover and not self.detect_cover:
                videos = []
                for item in self.items:
                    raw = item.get("video") if isinstance(item.get("video"), dict) else item
                    if isinstance(raw, dict):
                        try:
                            videos.append((item, YouTubeVideo(**raw)))
                        except TypeError:
                            videos.append((item, None))
                    else:
                        videos.append((item, None))
                with ThreadPoolExecutor(max_workers=max(1, min(self.download_concurrency, 6))) as executor:
                    futures = {executor.submit(cover_service.download_cover, video): (index, item, video) for index, (item, video) in enumerate(videos) if video is not None}
                    for future in as_completed(futures):
                        index, item, video = futures[future]
                        downloaded = future.result()
                        if downloaded.error:
                            self.item_updated.emit(index, "封面失败", "", downloaded.error)
                            results.append({"video": video.to_dict(), "cover_path": "", "error": downloaded.error})
                        else:
                            self.item_updated.emit(index, "处理完成", downloaded.path, "")
                            results.append({"video": video.to_dict(), "cover_path": downloaded.path})
                self.completed.emit(results, False)
                return
            for index, item in enumerate(self.items):
                if not self.control.checkpoint():
                    break
                raw_video = item.get("video") if isinstance(item.get("video"), dict) else item
                if not isinstance(raw_video, dict):
                    self.item_updated.emit(index, "处理失败", "", "视频信息无效")
                    continue
                try:
                    video = YouTubeVideo(**raw_video)
                    cover_path = str(item.get("cover_path") or "")
                    if self.download_cover or not Path(cover_path).is_file():
                        self.item_updated.emit(index, "获取封面", "", "")
                        downloaded = cover_service.download_cover(video)
                        if downloaded.error:
                            self.item_updated.emit(index, "封面失败", "", downloaded.error)
                            results.append({"video": video.to_dict(), "cover_path": "", "error": downloaded.error})
                            continue
                        cover_path = downloaded.path
                    if self.detect_cover:
                        self.item_updated.emit(index, "检测封面", cover_path, "")
                        review = review_service.review_cover(video, cover_path)
                        self.item_updated.emit(
                            index,
                            "处理完成" if not review.error else "检测失败",
                            cover_path,
                            review.to_dict(),
                        )
                        results.append(
                            {
                                "video": video.to_dict(),
                                "cover_path": cover_path,
                                "review": review.to_dict(),
                            }
                        )
                    else:
                        self.item_updated.emit(index, "封面已获取", cover_path, "")
                        results.append({"video": video.to_dict(), "cover_path": cover_path})
                except Exception as exc:  # noqa: BLE001
                    self.item_updated.emit(index, "处理失败", "", str(exc))
                    results.append({"video": raw_video, "cover_path": "", "error": str(exc)})
            self.completed.emit(results, self.control.cancelled)
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowIcon(load_app_icon())
        self.setWindowTitle("YouTube 字幕核验助手")
        apply_responsive_window_geometry(
            self,
            preferred_width=1180,
            preferred_height=820,
            minimum_width=940,
            minimum_height=640,
        )
        self._videos: list[YouTubeVideo] = []
        self._ready_items: list[dict[str, object]] = []
        self._active_matching_items: list[dict[str, object]] = []
        self._pending_asr: list[dict[str, object]] = []
        self._active_video_ids: set[str] = set()
        self._completed_video_ids: set[str] = set()
        self._cover_paths: dict[str, str] = {}
        self._cover_review_results: list[CoverReviewResult] = []
        self._cover_review_refresh_pending = False
        self._matching_result_rows: list[dict[str, object]] = []
        self._match_paused = False
        self._task_control: TaskControl | None = None
        self._thread: QThread | None = None
        # The matching service owns its queue; keep it independent from local workers.
        self._matching_thread: _MatchThread | None = None
        self._parallel_cover_review_thread: _CoverReviewThread | None = None
        self._parallel_cover_review_control: TaskControl | None = None
        self._cover_collect_thread: _CoverCollectThread | None = None
        self._standalone_cover_progress_total = 0
        self._standalone_cover_progress_completed = 0
        self._evidence_thread: _EvidenceContextThread | None = None
        self._youtube_login_dialog: YouTubeLoginDialog | None = None
        self._browser_capture_dialog: BrowserAudioCaptureDialog | None = None
        self._asr_config_dialog: AsrConfigDialog | None = None
        self._llm_config_dialog: LlmConfigDialog | None = None
        self._matching_config_dialog: MatchingConfigDialog | None = None
        self._update_check_thread: _UpdateCheckThread | None = None
        self._update_download_thread: _UpdateDownloadThread | None = None
        self._update_progress_dialog: QProgressDialog | None = None
        self._api_config_service = ApiConfigService()
        self._standalone_download_results: list[dict[str, object]] = []
        self._task_spec: dict[str, object] = {}
        self._resume_auto_asr = False
        self._recovery_prompt_shown = False
        self._closing_for_recovery = False
        self._state_store = TaskStateStore()
        self._state_persist_timer = QTimer(self)
        self._state_persist_timer.setSingleShot(True)
        self._state_persist_timer.timeout.connect(self._persist_workspace_state)
        self._workspace_refresh_timer = QTimer(self)
        self._build_ui()
        self._apply_matching_service_config(self._api_config_service.get_matching_service_config())
        self._workspace_refresh_timer.timeout.connect(self._refresh_workspace_pages)
        self._workspace_refresh_timer.start(800)
        QTimer.singleShot(500, self._offer_detected_system_proxy)

    def _build_ui(self) -> None:
        root = QWidget(self)
        root.setObjectName("AppRoot")
        layout = QVBoxLayout(root)
        layout.setContentsMargins(18, 14, 18, 16)
        layout.setSpacing(10)

        top_bar = QFrame()
        top_bar.setObjectName("TopBar")
        top_layout = QVBoxLayout(top_bar)
        top_layout.setContentsMargins(18, 10, 18, 10)
        top_layout.setSpacing(7)
        top_main_row = QHBoxLayout()
        brand = QVBoxLayout()
        brand.setSpacing(1)
        brand_title = QLabel("YouTube 字幕核验助手")
        brand_title.setObjectName("BrandTitle")
        brand_subtitle = QLabel("采集 · 字幕 · 转写 · 匹配")
        brand_subtitle.setObjectName("BrandSubtitle")
        brand.addWidget(brand_title)
        brand.addWidget(brand_subtitle)
        top_main_row.addLayout(brand, 1)

        self.export_button = QPushButton("导出字幕 Excel")
        self.export_button.clicked.connect(self._export_subtitles)
        self.export_button.setEnabled(False)
        self.open_cover_folder_button = QPushButton("封面目录")
        self.open_cover_folder_button.clicked.connect(self._open_cover_directory)
        self.open_output_folder_button = QPushButton("打开输出目录")
        self.open_output_folder_button.clicked.connect(self._open_output_directory)
        self.cover_review_button = QPushButton("检测已下载封面（3路）")
        self.cover_review_button.setToolTip("仅检测已下载封面；字幕获取期间可与字幕任务同步执行，模型检测保持最多 3 路并发。")
        self.cover_review_button.clicked.connect(self._review_covers)
        self.cover_review_button.setEnabled(False)
        self.asr_config_button = QPushButton("ASR 配置")
        self.asr_config_button.clicked.connect(self._open_asr_config)
        self.llm_config_button = QPushButton("语言模型")
        self.llm_config_button.clicked.connect(self._open_llm_config)
        self.matching_config_button = QPushButton("匹配配置")
        self.matching_config_button.clicked.connect(self._open_matching_config)
        self.youtube_login_button = QPushButton("YouTube 登录")
        self.youtube_login_button.clicked.connect(self._open_youtube_login)
        self.update_button = QPushButton("检查更新")
        self.update_button.clicked.connect(self._check_for_updates)
        self.update_button.setToolTip(f"当前版本：{APP_VERSION}")
        self.clear_button = QPushButton("清空任务")
        self.clear_button.clicked.connect(self._clear)
        self.export_button.setProperty("secondary", True)
        self.open_cover_folder_button.setProperty("secondary", True)
        self.open_output_folder_button.setProperty("secondary", True)
        self.cover_review_button.setProperty("primary", True)
        self.asr_config_button.setProperty("secondary", True)
        self.llm_config_button.setProperty("secondary", True)
        self.matching_config_button.setProperty("secondary", True)
        self.update_button.setProperty("secondary", True)
        self.clear_button.setProperty("danger", True)
        top_main_row.addWidget(self.clear_button)
        self._top_tools_layout = QGridLayout()
        self._top_tools_layout.setHorizontalSpacing(8)
        self._top_tools_layout.setVerticalSpacing(7)
        self._top_tool_widgets = (
            self.export_button,
            self.open_cover_folder_button,
            self.open_output_folder_button,
            self.cover_review_button,
            self.asr_config_button,
            self.llm_config_button,
            self.matching_config_button,
            self.youtube_login_button,
            self.update_button,
        )
        top_layout.addLayout(top_main_row)
        top_layout.addLayout(self._top_tools_layout)
        layout.addWidget(top_bar)

        entry_card = QFrame()
        entry_card.setObjectName("EntryCard")
        entry_layout = QGridLayout(entry_card)
        entry_layout.setContentsMargins(18, 14, 18, 14)
        entry_layout.setHorizontalSpacing(10)
        entry_layout.setVerticalSpacing(10)
        entry_title = QLabel("开始一个频道任务")
        entry_title.setObjectName("SectionTitle")
        entry_hint = QLabel("输入频道或 Shorts 页面链接，先采集视频清单")
        entry_hint.setObjectName("SectionHint")
        entry_layout.addWidget(entry_title, 0, 0)
        entry_layout.addWidget(entry_hint, 0, 1, 1, 3)

        self.channel_input = QLineEdit()
        self.channel_input.setPlaceholderText("频道链接，例如 https://www.youtube.com/@channel-name")
        self.collect_button = QPushButton("采集频道视频")
        self.collect_button.clicked.connect(self._collect)
        self.collect_button.setProperty("primary", True)
        entry_layout.addWidget(QLabel("频道链接"), 1, 0)
        entry_layout.addWidget(self.channel_input, 1, 1, 1, 2)
        entry_layout.addWidget(self.collect_button, 1, 3)

        self.limit_spin = QSpinBox()
        self.limit_spin.setRange(0, 2000)
        self.limit_spin.setSpecialValueText("全部")
        self.seconds_combo = QComboBox()
        for label, value in (("前 1 分钟", 60), ("前 3 分钟", 180), ("前 5 分钟", 300)):
            self.seconds_combo.addItem(label, value)
        self.seconds_combo.setCurrentIndex(1)
        entry_layout.addWidget(QLabel("采集上限"), 2, 0)
        entry_layout.addWidget(self.limit_spin, 2, 1)
        entry_layout.addWidget(QLabel("字幕范围"), 2, 2)
        entry_layout.addWidget(self.seconds_combo, 2, 3)

        for widget in (
            self.channel_input,
            self.limit_spin,
            self.seconds_combo,
            self.collect_button,
        ):
            widget.setFixedHeight(32)
        entry_layout.setColumnStretch(1, 1)
        entry_layout.setColumnStretch(2, 2)
        layout.addWidget(entry_card)

        self.advanced_toggle_button = QPushButton("高级下载与代理设置")
        self.advanced_toggle_button.setProperty("secondary", True)
        self.advanced_toggle_button.clicked.connect(self._toggle_advanced_settings)
        self.advanced_settings_panel = QFrame()
        self.advanced_settings_panel.setObjectName("AdvancedPanel")
        advanced_layout = QGridLayout(self.advanced_settings_panel)
        advanced_layout.setContentsMargins(16, 12, 16, 12)
        advanced_layout.setHorizontalSpacing(10)
        advanced_layout.setVerticalSpacing(8)
        self.proxy_input = QLineEdit()
        self.proxy_input.setPlaceholderText("http://127.0.0.1:7897（可选）")
        try:
            self.proxy_input.setText(YOUTUBE_PROXY_CONFIG_PATH.read_text(encoding="utf-8").strip())
        except OSError:
            pass
        self.proxy_save_button = QPushButton("保存代理")
        self.proxy_save_button.clicked.connect(lambda: self._save_proxy(show_feedback=True))
        self.proxy_detect_button = QPushButton("检测本机代理")
        self.proxy_detect_button.setProperty("secondary", True)
        self.proxy_detect_button.clicked.connect(lambda: self._detect_system_proxy(interactive=True))
        for widget in (
            self.proxy_input,
            self.proxy_save_button,
            self.proxy_detect_button,
        ):
            widget.setFixedHeight(32)
        advanced_layout.addWidget(QLabel("下载代理"), 0, 0)
        advanced_layout.addWidget(self.proxy_input, 0, 1, 1, 2)
        advanced_layout.addWidget(self.proxy_save_button, 0, 3)
        advanced_layout.addWidget(self.proxy_detect_button, 1, 1)
        proxy_hint = QLabel("读取 Windows 系统代理、WinHTTP 与环境变量；发现后需确认才会保存。")
        proxy_hint.setObjectName("SectionHint")
        proxy_hint.setWordWrap(True)
        advanced_layout.addWidget(proxy_hint, 1, 2, 1, 2)
        advanced_layout.setColumnStretch(1, 2)
        advanced_layout.setColumnStretch(3, 1)
        self.advanced_settings_panel.setVisible(False)
        layout.addWidget(self.advanced_toggle_button, 0, Qt.AlignmentFlag.AlignLeft)
        layout.addWidget(self.advanced_settings_panel)

        workflow_card = QFrame()
        workflow_card.setObjectName("WorkflowCard")
        workflow_layout = QVBoxLayout(workflow_card)
        workflow_layout.setContentsMargins(16, 10, 16, 11)
        workflow_heading = QLabel("处理流程")
        workflow_heading.setObjectName("SectionTitle")
        workflow_layout.addWidget(workflow_heading)
        self._workflow_steps_layout = QGridLayout()
        self._workflow_steps_layout.setHorizontalSpacing(8)
        self._workflow_steps_layout.setVerticalSpacing(8)
        self._workflow_step_widgets: list[QFrame] = []
        self.cover_button = QPushButton("下载选中封面")
        self.cover_button.pressed.connect(self._on_cover_button_pressed)
        self.cover_button.clicked.connect(self._download_covers)
        self.cover_button.setEnabled(False)
        self.prepare_button = QPushButton("获取前段字幕")
        self.prepare_button.clicked.connect(self._prepare)
        self.prepare_button.setEnabled(False)
        self.asr_fallback_button = QPushButton("下载并转写兜底")
        self.asr_fallback_button.clicked.connect(self._prepare_asr_fallback)
        self.asr_fallback_button.setEnabled(False)
        self.auto_asr_fallback_check = QCheckBox("无直出字幕时自动下载音频并 ASR 兜底（稳定方案）")
        self.auto_asr_fallback_check.setToolTip("启用后，每条视频未获取直出字幕时会立即使用稳定下载方案转写；关闭后保留到待兜底列表手动处理。")
        self.skip_caption_probe_check = QCheckBox("已知无字幕时直接下载转写")
        self.skip_caption_probe_check.setToolTip(
            "适用于已确认没有字幕接口的视频；开启后跳过字幕探测，直接下载音频并进行 ASR。"
        )
        self.auto_cover_review_check = QCheckBox("获取字幕时同步检测已下载封面（3路）")
        self.auto_cover_review_check.setToolTip("仅检测已成功下载的封面；与字幕获取并行运行，模型检测最多 3 路。")
        self.asr_download_concurrency_combo = QComboBox()
        self.asr_transcribe_concurrency_combo = QComboBox()
        self.caption_concurrency_combo = QComboBox()
        # The detailed behavior remains in tooltips; compact labels prevent the
        # options row from pushing controls outside the narrow workbench.
        self.auto_asr_fallback_check.setText("自动 ASR 兜底")
        self.auto_cover_review_check.setText("同步封面检测")
        for checkbox in (
            self.auto_asr_fallback_check,
            self.skip_caption_probe_check,
            self.auto_cover_review_check,
        ):
            checkbox.setMinimumWidth(0)
            checkbox.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        for combo in (
            self.caption_concurrency_combo,
            self.asr_download_concurrency_combo,
            self.asr_transcribe_concurrency_combo,
        ):
            for lanes in (1, 2, 3):
                combo.addItem(f"{lanes} 路", lanes)
            combo.setCurrentIndex(0)
            combo.setToolTip("默认 1 路最稳定；提高并发会增加网络或 ASR 服务负载。")
        self.match_button = QPushButton("提交匹配")
        self.match_button.clicked.connect(self._match)
        self.match_button.setEnabled(False)
        for button in (
            self.cover_button,
            self.prepare_button,
            self.asr_fallback_button,
            self.match_button,
        ):
            button.setProperty("primary", True)

        def add_workflow_step(number: str, title: str, hint: str, action: QPushButton | None = None) -> None:
            step = QFrame()
            step.setObjectName("WorkflowStep")
            step_layout = QHBoxLayout(step)
            step_layout.setContentsMargins(10, 7, 10, 7)
            number_label = QLabel(number)
            number_label.setObjectName("WorkflowNumber")
            copy = QVBoxLayout()
            step_title = QLabel(title)
            step_title.setObjectName("WorkflowTitle")
            step_hint = QLabel(hint)
            step_hint.setObjectName("WorkflowHint")
            copy.addWidget(step_title)
            copy.addWidget(step_hint)
            step_layout.addWidget(number_label)
            step_layout.addLayout(copy, 1)
            if action is not None:
                action.setFixedHeight(28)
                step_layout.addWidget(action)
            self._workflow_step_widgets.append(step)

        add_workflow_step("01", "下载封面素材", "下载勾选视频的封面，右侧显示实时进度", self.cover_button)
        add_workflow_step("02", "获取字幕文件", "优先使用直出字幕", self.prepare_button)
        add_workflow_step("03", "ASR 语音识别", "自动兜底失败时可手动重试", self.asr_fallback_button)
        add_workflow_step("04", "提交匹配校验", "回传结果并查看", self.match_button)
        self._reflow_workflow_steps(compact=False)
        workflow_layout.addLayout(self._workflow_steps_layout)
        workflow_options = QHBoxLayout()
        workflow_options.addWidget(self.auto_asr_fallback_check)
        workflow_options.addWidget(self.skip_caption_probe_check)
        workflow_options.addWidget(QLabel("字幕"))
        workflow_options.addWidget(self.caption_concurrency_combo)
        workflow_options.addWidget(QLabel("下载"))
        workflow_options.addWidget(self.asr_download_concurrency_combo)
        workflow_options.addWidget(QLabel("转写"))
        workflow_options.addWidget(self.asr_transcribe_concurrency_combo)
        workflow_options.addWidget(self.auto_cover_review_check)
        workflow_options.addStretch(1)
        workflow_layout.addLayout(workflow_options)
        layout.addWidget(workflow_card)

        queue_card = QFrame()
        queue_card.setObjectName("QueueCard")
        self._queue_card = queue_card
        queue_card_layout = QVBoxLayout(queue_card)
        queue_card_layout.setContentsMargins(14, 12, 14, 14)
        self._queue_header_layout = QGridLayout()
        self._queue_header_layout.setHorizontalSpacing(8)
        self._queue_header_layout.setVerticalSpacing(7)
        queue_title = QLabel("视频队列")
        queue_title.setObjectName("SectionTitle")
        self._queue_title = queue_title
        self.select_all_button = QPushButton("全选")
        self.select_all_button.clicked.connect(self._select_all)
        self.invert_selection_button = QPushButton("反选")
        self.invert_selection_button.clicked.connect(self._invert_selection)
        self.select_count_spin = QSpinBox()
        self.select_count_spin.setRange(1, 2000)
        self.select_count_spin.setValue(10)
        self.select_count_spin.setSuffix(" 条")
        self.select_count_spin.setFixedWidth(78)
        self.select_count_spin.setFixedHeight(32)
        self.select_count_spin.setToolTip("输入要勾选的前 N 条视频")
        self.select_first_ten_button = QPushButton("勾选前 10 条")
        self.select_first_ten_button.clicked.connect(self._select_first_ten)
        self.select_count_spin.valueChanged.connect(
            lambda value: self.select_first_ten_button.setText(f"勾选前 {value} 条")
        )
        self._select_count_widget = QWidget()
        select_count_layout = QHBoxLayout(self._select_count_widget)
        select_count_layout.setContentsMargins(0, 0, 0, 0)
        select_count_layout.setSpacing(6)
        select_count_layout.addWidget(self.select_count_spin)
        select_count_layout.addWidget(self.select_first_ten_button)
        self.channel_selection_combo = QComboBox()
        self.channel_selection_combo.setMinimumWidth(0)
        self.channel_selection_combo.addItem("选择频道", "")
        self.select_channel_button = QPushButton("勾选该频道")
        self.select_channel_button.clicked.connect(self._select_channel)
        self.delete_selected_button = QPushButton("\u5220\u9664\u9009\u4e2d")
        self.delete_selected_button.setProperty("danger", True)
        self.delete_selected_button.clicked.connect(self._delete_selected_items)
        self.enqueue_fallback_button = QPushButton("\u52a0\u5165\u5f85\u515c\u5e95")
        self.enqueue_fallback_button.setProperty("secondary", True)
        self.enqueue_fallback_button.clicked.connect(self._enqueue_selected_for_fallback)
        self._queue_control_widgets = (
            self.select_all_button,
            self.invert_selection_button,
            self._select_count_widget,
            self.channel_selection_combo,
            self.select_channel_button,
            self.enqueue_fallback_button,
            self.delete_selected_button,
        )

        self.audio_strategy_combo = QComboBox()
        self.audio_strategy_combo.addItem("稳定分段下载", "download")
        self.audio_strategy_combo.setEnabled(False)
        self.browser_concurrency_combo = QComboBox()
        self.browser_concurrency_combo.addItem("浏览器并发 1 路", 1)
        self.browser_concurrency_combo.addItem("浏览器并发 2 路", 2)
        self.browser_concurrency_combo.addItem("浏览器并发 3 路", 3)
        self.browser_concurrency_combo.setEnabled(False)
        self.browser_concurrency_combo.setVisible(False)
        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(self._toggle_pause)
        self.pause_button.setEnabled(False)
        self.cancel_task_button = QPushButton("取消任务")
        self.cancel_task_button.clicked.connect(self._cancel_task)
        self.cancel_task_button.setEnabled(False)

        self.status_label = QLabel("请输入频道链接开始采集")
        self.status_label.setObjectName("StatusBar")
        self.status_label.setWordWrap(False)
        self.status_label.setMinimumWidth(0)
        self.status_label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Fixed)
        self.status_label.setFixedHeight(38)
        self.cover_progress_bar = QProgressBar()
        self.cover_progress_bar.setTextVisible(True)
        self.cover_progress_bar.setFixedHeight(24)
        self.cover_progress_bar.setRange(0, 1)
        self.cover_progress_bar.setValue(0)
        self.cover_progress_bar.setFormat("封面下载：等待开始")
        self.video_table = QTableWidget(0, 6)
        self.video_table.setHorizontalHeaderLabels(["选择", "标题", "视频 ID", "频道", "字幕状态", "封面状态"])
        self.video_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.video_table.setColumnWidth(0, 70)
        self.video_table.setColumnWidth(1, 170)
        self.video_table.setColumnWidth(2, 115)
        self.video_table.setColumnWidth(3, 125)
        self.video_table.setColumnWidth(5, 95)
        self.video_table.horizontalHeader().setStretchLastSection(False)
        self.video_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.fallback_table = QTableWidget(0, 5)
        self.fallback_table.setHorizontalHeaderLabels(["选择", "标题", "视频 ID", "频道", "原因"])
        self.fallback_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.fallback_table.horizontalHeader().setStretchLastSection(True)
        self.cover_review_table = QTableWidget(0, 6)
        self.cover_review_table.setHorizontalHeaderLabels(
            ["标题", "视频 ID", "检测结论", "风险标签", "置信度", "可见依据"]
        )
        self.cover_review_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.cover_review_table.setColumnCount(7)
        self.cover_review_table.setHorizontalHeaderLabels(
            [
                "\u9009\u62e9",
                "\u6807\u9898",
                "\u89c6\u9891 ID",
                "\u68c0\u6d4b\u7ed3\u8bba",
                "\u98ce\u9669\u6807\u7b7e",
                "\u7f6e\u4fe1\u5ea6",
                "\u53ef\u89c1\u4f9d\u636e",
            ]
        )
        self.cover_review_table.setColumnWidth(0, 180)
        self.cover_review_table.setColumnWidth(1, 115)
        self.cover_review_table.setColumnWidth(2, 90)
        self.cover_review_table.setColumnWidth(3, 150)
        self.cover_review_table.setColumnWidth(4, 70)
        self.cover_review_table.setColumnWidth(0, 70)
        self.cover_review_table.setColumnWidth(1, 180)
        self.cover_review_table.setColumnWidth(2, 115)
        self.cover_review_table.setColumnWidth(3, 90)
        self.cover_review_table.setColumnWidth(4, 150)
        self.cover_review_table.setColumnWidth(5, 70)
        self.cover_review_table.horizontalHeader().setStretchLastSection(True)
        self.task_tabs = QTabWidget()
        queue_page = QWidget()
        queue_page_layout = QVBoxLayout(queue_page)
        queue_page_layout.setContentsMargins(0, 0, 0, 0)
        queue_page_layout.addWidget(self.video_table)
        fallback_page = QWidget()
        fallback_layout = QVBoxLayout(fallback_page)
        fallback_layout.setContentsMargins(0, 0, 0, 0)
        fallback_layout.addWidget(self.fallback_table)
        cover_review_page = QWidget()
        cover_review_layout = QVBoxLayout(cover_review_page)
        cover_review_layout.setContentsMargins(0, 0, 0, 0)
        cover_review_layout.addWidget(self.cover_review_table)
        self.task_tabs.addTab(queue_page, "视频队列")
        self.task_tabs.addTab(fallback_page, "待兜底 (0)")
        self.task_tabs.addTab(cover_review_page, "封面检测 (0)")
        self.task_tabs.setDocumentMode(True)
        self.task_tabs.currentChanged.connect(self._on_task_tab_changed)
        self.task_tabs.setMinimumHeight(260)
        queue_card_layout.addLayout(self._queue_header_layout)
        queue_card_layout.addWidget(self.task_tabs, 1)

        self.result_edit = QPlainTextEdit()
        self.result_edit.setReadOnly(True)
        self.result_edit.setPlaceholderText("匹配结果将在这里显示")
        self.result_edit.setMinimumHeight(160)
        side_panel = QFrame()
        side_panel.setObjectName("SidePanel")
        self._side_panel = side_panel
        side_layout = QVBoxLayout(side_panel)
        side_layout.setContentsMargins(14, 12, 14, 14)
        side_title = QLabel("任务控制")
        side_title.setObjectName("SectionTitle")
        side_layout.addWidget(side_title)
        metrics_row = QHBoxLayout()
        self.side_total_metric = QLabel("队列\n0")
        self.side_ready_metric = QLabel("可匹配\n0")
        self.side_fallback_metric = QLabel("待兜底\n0")
        for metric in (self.side_total_metric, self.side_ready_metric, self.side_fallback_metric):
            metric.setObjectName("MetricCard")
            metric.setAlignment(Qt.AlignmentFlag.AlignCenter)
            metrics_row.addWidget(metric)
        side_layout.addLayout(metrics_row)
        side_layout.addWidget(self.status_label)
        side_layout.addWidget(self.cover_progress_bar)
        side_layout.addWidget(QLabel("音频方案"))
        side_layout.addWidget(self.audio_strategy_combo)
        side_layout.addWidget(self.browser_concurrency_combo)
        controls = QHBoxLayout()
        controls.addWidget(self.pause_button)
        controls.addWidget(self.cancel_task_button)
        side_layout.addLayout(controls)
        result_title = QLabel("匹配结果")
        result_title.setObjectName("SectionTitle")
        side_layout.addWidget(result_title)
        side_layout.addWidget(self.result_edit, 1)

        self.workspace_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.workspace_splitter.addWidget(queue_card)
        self.workspace_splitter.addWidget(side_panel)
        self.workspace_splitter.setChildrenCollapsible(False)
        self.workspace_splitter.setStretchFactor(0, 4)
        self.workspace_splitter.setStretchFactor(1, 2)
        self.workspace_splitter.setSizes([780, 360])
        layout.addWidget(self.workspace_splitter, 1)
        page_scroll = QScrollArea()
        page_scroll.setWidgetResizable(True)
        page_scroll.setFrameShape(QFrame.Shape.NoFrame)
        page_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        root.setMinimumWidth(0)
        page_scroll.setWidget(root)
        self._legacy_flow_page = page_scroll
        self._dashboard_page = DashboardPage(self._navigate_workspace, self)
        self._download_page = DownloadPage(
            start_download=self._start_standalone_download,
            pause_download=self._toggle_pause,
            cancel_download=self._cancel_task,
            parent=self,
        )
        self._cover_page = CoverPage(
            import_videos=self._import_cover_items,
            collect_channels=self._collect_cover_channels,
            start_review=self._start_cover_page_review,
            start_url_review=self._start_cover_page_url_review,
            remove_items=self._remove_cover_queue_items,
            start_cover=self._start_standalone_cover,
            pause_cover=self._toggle_pause,
            cancel_cover=self._cancel_task,
            export_reviews=self._export_cover_review_results,
            parent=self,
        )
        self._subtitle_page = SubtitlePage(
            start_transcription=self._start_standalone_transcription,
            import_downloads=self._import_download_results,
            pause_transcription=self._toggle_pause,
            cancel_transcription=self._cancel_task,
            parent=self,
        )
        self._matching_page = MatchingPage(
            import_ready_items=self._import_ready_items,
            import_files=self._import_matching_files,
            start_matching=self._start_standalone_matching,
            pause_matching=self._toggle_pause,
            cancel_matching=self._cancel_task,
            load_evidence_context=self._load_matching_evidence_context,
            export_results=self._export_matching_results,
            parent=self,
        )
        self._task_center_page = TaskCenterPage(
            lambda: self._navigate_workspace("full_flow"),
            self,
        )
        self._settings_page = SettingsHubPage(
            open_youtube_login=self._open_youtube_login,
            open_asr=self._open_asr_config,
            open_llm=self._open_llm_config,
            open_connection_settings=self._open_connection_settings,
            open_matching_config=self._open_matching_config,
            check_updates=self._check_for_updates,
            parent=self,
        )
        self._page_stack = QStackedWidget()
        self._page_indices = {
            "dashboard": self._page_stack.addWidget(self._dashboard_page),
            "full_flow": self._page_stack.addWidget(self._legacy_flow_page),
            "download": self._page_stack.addWidget(self._download_page),
            "cover": self._page_stack.addWidget(self._cover_page),
            "subtitles": self._page_stack.addWidget(self._subtitle_page),
            "matching": self._page_stack.addWidget(self._matching_page),
            "tasks": self._page_stack.addWidget(self._task_center_page),
            "settings": self._page_stack.addWidget(self._settings_page),
        }

        shell = QWidget(self)
        shell.setObjectName("WorkspaceShell")
        shell_layout = QHBoxLayout(shell)
        shell_layout.setContentsMargins(0, 0, 0, 0)
        shell_layout.setSpacing(0)
        self._navigation_rail = QFrame()
        self._navigation_rail.setObjectName("NavigationRail")
        nav_layout = QVBoxLayout(self._navigation_rail)
        nav_layout.setContentsMargins(14, 18, 14, 16)
        nav_layout.setSpacing(8)
        nav_brand = QLabel("YouTube\n字幕核验助手")
        nav_brand.setObjectName("NavBrand")
        nav_layout.addWidget(nav_brand)
        nav_layout.addSpacing(12)
        self._nav_buttons: dict[str, QPushButton] = {}
        self._nav_button_labels: dict[str, str] = {}
        for page_id, label in (
            ("dashboard", "工作台"),
            ("full_flow", "全流程处理"),
            ("download", "视频下载"),
            ("cover", "封面检测"),
            ("subtitles", "字幕转写"),
            ("matching", "匹配核验"),
            ("tasks", "任务中心"),
            ("settings", "设置"),
        ):
            button = QPushButton(label)
            button.setCheckable(True)
            button.setProperty("nav", True)
            button.clicked.connect(lambda _checked=False, target=page_id: self._navigate_workspace(target))
            nav_layout.addWidget(button)
            self._nav_buttons[page_id] = button
            self._nav_button_labels[page_id] = label
        nav_layout.addStretch(1)
        nav_footer = QLabel("稳定版工作区\n功能将按页逐步迁移")
        nav_footer.setObjectName("NavFooter")
        nav_footer.setWordWrap(True)
        nav_layout.addWidget(nav_footer)
        self._nav_brand = nav_brand
        self._nav_footer = nav_footer
        self._navigation_rail.setFixedWidth(184)
        shell_layout.addWidget(self._navigation_rail)
        shell_layout.addWidget(self._page_stack, 1)
        self.setCentralWidget(shell)
        self._navigate_workspace("dashboard")
        QTimer.singleShot(0, self._update_responsive_layout)

    def _toggle_advanced_settings(self) -> None:
        visible = not self.advanced_settings_panel.isVisible()
        self.advanced_settings_panel.setVisible(visible)
        self.advanced_toggle_button.setText(
            "收起高级下载与代理设置" if visible else "高级下载与代理设置"
        )

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        QTimer.singleShot(0, self._update_responsive_layout)

    def _update_responsive_layout(self) -> None:
        if not hasattr(self, "workspace_splitter"):
            return
        rail_compact = self.width() < 1100
        rail_width = 76 if rail_compact else 184
        self._navigation_rail.setFixedWidth(rail_width)
        self._nav_brand.setVisible(not rail_compact)
        self._nav_footer.setVisible(not rail_compact)
        for page_id, button in self._nav_buttons.items():
            label = self._nav_button_labels.get(page_id, button.text())
            button.setText(label[:2] if rail_compact else label)
            button.setToolTip(label)

        # Base the breakpoint on actual workbench width rather than monitor
        # resolution. This is stable across Windows display-scale settings.
        available_width = max(0, self.width() - rail_width)
        compact = available_width < 1000
        target_orientation = Qt.Orientation.Vertical if compact else Qt.Orientation.Horizontal
        if self.workspace_splitter.orientation() != target_orientation:
            self.workspace_splitter.setOrientation(target_orientation)
            self.workspace_splitter.setSizes([520, 410] if compact else [780, 360])
        # On narrow screens the full-flow panels stack vertically. Preserve a
        # useful table and result area; the parent scroll area handles overflow.
        self._queue_card.setMinimumHeight(430 if compact else 0)
        self._side_panel.setMinimumHeight(370 if compact else 0)
        self._reflow_workflow_steps(compact=compact)
        self._reflow_queue_header(compact=compact)
        self._reflow_top_tools(compact=compact)

    def _reflow_top_tools(self, *, compact: bool) -> None:
        if not hasattr(self, "_top_tools_layout"):
            return
        while self._top_tools_layout.count():
            self._top_tools_layout.takeAt(0)
        for column in range(max(1, len(self._top_tool_widgets))):
            self._top_tools_layout.setColumnStretch(column, 0)
        columns = 3 if compact else len(self._top_tool_widgets)
        for column in range(columns):
            self._top_tools_layout.setColumnStretch(column, 1 if compact else 0)
        for index, button in enumerate(self._top_tool_widgets):
            row, column = divmod(index, columns)
            self._top_tools_layout.addWidget(button, row, column)

    def _reflow_queue_header(self, *, compact: bool) -> None:
        if not hasattr(self, "_queue_header_layout"):
            return
        while self._queue_header_layout.count():
            self._queue_header_layout.takeAt(0)
        self._queue_header_layout.setColumnStretch(0, 0)
        self._queue_header_layout.setColumnStretch(1, 0)
        self._queue_header_layout.setColumnStretch(2, 0)
        self._queue_header_layout.setColumnStretch(3, 0)
        self._queue_header_layout.setColumnStretch(4, 0)
        self._queue_header_layout.addWidget(self._queue_title, 0, 0)
        if compact:
            for column in range(3):
                self._queue_header_layout.setColumnStretch(column, 1)
            for column, widget in enumerate(self._queue_control_widgets[:3]):
                self._queue_header_layout.addWidget(widget, 1, column)
            self._queue_header_layout.addWidget(self.channel_selection_combo, 2, 0, 1, 2)
            self._queue_header_layout.addWidget(self.select_channel_button, 2, 2)
            self._queue_header_layout.addWidget(self.enqueue_fallback_button, 3, 0)
            self._queue_header_layout.addWidget(self.delete_selected_button, 3, 1)
        else:
            self._queue_header_layout.setColumnStretch(1, 1)
            for column, widget in enumerate(self._queue_control_widgets[:-1], start=2):
                self._queue_header_layout.addWidget(widget, 0, column)
            self._queue_header_layout.addWidget(self.delete_selected_button, 0, 1, Qt.AlignmentFlag.AlignRight)

    @staticmethod
    def _short_title(title: str, limit: int = 38) -> str:
        title = str(title or "").strip()
        if len(title) <= limit:
            return title
        return f"{title[:limit]}..."

    def _reflow_workflow_steps(self, *, compact: bool) -> None:
        if not hasattr(self, "_workflow_steps_layout"):
            return
        while self._workflow_steps_layout.count():
            self._workflow_steps_layout.takeAt(0)
        for column in range(4):
            self._workflow_steps_layout.setColumnStretch(column, 0)
        for column in range(2 if compact else 4):
            self._workflow_steps_layout.setColumnStretch(column, 1)
        for index, step in enumerate(self._workflow_step_widgets):
            row, column = (divmod(index, 2) if compact else (0, index))
            self._workflow_steps_layout.addWidget(step, row, column)

    def _set_busy(self, busy: bool, *, scope: str = "local") -> None:
        """Lock only the task lane that is running.

        Matching is a server-side queued task, so its polling thread must not
        disable local collection, downloading, or subtitle preparation.
        """
        local_busy = bool(busy) if scope == "local" else self._thread_is_running()
        matching_busy = bool(busy) if scope == "matching" else self._matching_thread_is_running()
        parallel_cover_running = self._parallel_cover_review_is_running()
        self._refresh_dashboard_metrics()
        self.collect_button.setEnabled(not local_busy)
        if hasattr(self, "_download_page"):
            self._download_page.set_busy(local_busy)
        if hasattr(self, "_cover_page"):
            self._cover_page.set_busy(
                local_busy or parallel_cover_running,
                allow_review=local_busy and self._task_control is not None and self._task_control.paused,
            )
        if hasattr(self, "_subtitle_page"):
            self._subtitle_page.set_busy(local_busy)
        if hasattr(self, "_matching_page"):
            self._matching_page.set_busy(matching_busy, submission_blocked=local_busy)
        subtitle_preparing = isinstance(self._thread, _PrepareThread)
        self.cover_button.setEnabled(not local_busy and bool(self._videos))
        self.cover_review_button.setEnabled(
            not parallel_cover_running
            and (not local_busy or subtitle_preparing)
            and any(video.video_id in self._cover_paths for video in self._selected_videos())
        )
        self.prepare_button.setEnabled(not local_busy and bool(self._videos))
        self.export_button.setEnabled(not local_busy and bool(self._ready_items))
        self.asr_fallback_button.setEnabled(
            not local_busy and bool(self._pending_asr or self._selected_direct_asr_videos())
        )
        self.auto_asr_fallback_check.setEnabled(not local_busy)
        self.skip_caption_probe_check.setEnabled(not local_busy)
        self.caption_concurrency_combo.setEnabled(not local_busy)
        self.asr_download_concurrency_combo.setEnabled(not local_busy)
        self.asr_transcribe_concurrency_combo.setEnabled(not local_busy)
        self.auto_cover_review_check.setEnabled(not local_busy)
        self.match_button.setEnabled(not local_busy and not matching_busy and bool(self._ready_items))
        for button in (
            self.select_all_button,
            self.invert_selection_button,
            self.select_first_ten_button,
            self.select_channel_button,
            self.enqueue_fallback_button,
            self.delete_selected_button,
        ):
            button.setEnabled(not local_busy and bool(self._videos))
        self.delete_selected_button.setEnabled(
            not local_busy and bool(self._selection_table().rowCount())
        )
        self.channel_selection_combo.setEnabled(not local_busy and bool(self._videos))
        self.select_count_spin.setEnabled(not local_busy and bool(self._videos))
        self.audio_strategy_combo.setEnabled(False)
        self.browser_concurrency_combo.setEnabled(
            not local_busy and self.audio_strategy_combo.currentData() == "browser"
        )
        local_task_active = local_busy and self._task_control is not None and not self._task_control.cancelled
        self.pause_button.setEnabled(local_task_active or matching_busy)
        self.cancel_task_button.setEnabled(local_task_active or matching_busy)
        self.clear_button.setEnabled(not local_busy and not matching_busy)
        self.asr_config_button.setEnabled(not local_busy)
        self.llm_config_button.setEnabled(not local_busy)
        self.matching_config_button.setEnabled(not local_busy)
        self.youtube_login_button.setEnabled(not local_busy)
        if not local_busy and not matching_busy:
            QTimer.singleShot(120, self._persist_workspace_state)

    def _thread_is_running(self) -> bool:
        """Return worker state without touching a Qt thread wrapper already deleted by Qt."""
        thread = self._thread
        if thread is None:
            return False
        try:
            return bool(thread.isRunning())
        except RuntimeError:
            # Existing workers use deleteLater() after completion. Keeping that
            # Python wrapper and probing it later raises here, which previously
            # aborted the next task before its thread could be started.
            LOGGER.info("Discarding deleted worker thread reference.")
            self._thread = None
            return False

    def _matching_thread_is_running(self) -> bool:
        thread = self._matching_thread
        if thread is None:
            return False
        try:
            return bool(thread.isRunning())
        except RuntimeError:
            LOGGER.info("Discarding deleted matching worker thread reference.")
            self._matching_thread = None
            return False

    def _finish_matching_thread(self, thread: _MatchThread) -> None:
        if self._matching_thread is thread:
            self._matching_thread = None
        thread.deleteLater()

    def _parallel_cover_review_is_running(self) -> bool:
        thread = self._parallel_cover_review_thread
        if thread is None:
            return False
        try:
            return bool(thread.isRunning())
        except RuntimeError:
            self._parallel_cover_review_thread = None
            self._parallel_cover_review_control = None
            return False

    def _check_for_updates(self) -> None:
        if self._update_check_thread is not None and self._update_check_thread.isRunning():
            return
        self.update_button.setEnabled(False)
        self.status_label.setText("正在检查软件更新...")
        thread = _UpdateCheckThread()
        self._update_check_thread = thread
        thread.completed.connect(self._on_update_check_completed)
        thread.failed.connect(self._on_update_check_failed)
        thread.finished.connect(self._finish_update_check_thread)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def _on_update_check_completed(self, update: AvailableUpdate | None) -> None:
        self.update_button.setEnabled(True)
        if update is None:
            self.status_label.setText(f"当前已是最新版本（{APP_VERSION}）。")
            QMessageBox.information(self, "检查更新", f"当前已是最新版本：{APP_VERSION}")
            return
        size_text = "大小将在下载时显示"
        if update.asset.size_bytes:
            size_text = f"约 {update.asset.size_bytes / 1_048_576:.0f} MB"
        message = (
            f"发现新版本：{update.version}\n"
            f"当前版本：{APP_VERSION}\n"
            f"更新包：{size_text}\n\n"
            "下载完成并校验后将自动更新和重启。runtime、output 中的配置与任务数据会保留。"
        )
        answer = QMessageBox.question(
            self,
            "发现新版本",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if answer == QMessageBox.StandardButton.Yes:
            self._download_update(update)
        else:
            self.status_label.setText(f"已取消更新到 {update.version}。")

    def _on_update_check_failed(self, message: str) -> None:
        self.update_button.setEnabled(True)
        self.status_label.setText("检查更新失败。")
        QMessageBox.warning(self, "检查更新失败", message)

    def _download_update(self, update: AvailableUpdate) -> None:
        self.update_button.setEnabled(False)
        dialog = QProgressDialog("正在下载并校验更新包...", "", 0, 0, self)
        dialog.setWindowTitle("正在更新")
        dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
        dialog.setAutoClose(False)
        dialog.setAutoReset(False)
        dialog.setCancelButton(None)
        dialog.setMinimumDuration(0)
        dialog.show()
        self._update_progress_dialog = dialog
        thread = _UpdateDownloadThread(update)
        self._update_download_thread = thread
        thread.progress.connect(self._on_update_download_progress)
        thread.completed.connect(self._on_update_download_completed)
        thread.failed.connect(self._on_update_download_failed)
        thread.finished.connect(self._finish_update_download_thread)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def _on_update_download_progress(self, received: int, total: int) -> None:
        dialog = self._update_progress_dialog
        if dialog is None:
            return
        if total > 0:
            dialog.setRange(0, total)
            dialog.setValue(min(received, total))
            dialog.setLabelText(f"正在下载更新包：{received / 1_048_576:.1f} / {total / 1_048_576:.1f} MB")
        else:
            dialog.setRange(0, 0)
            dialog.setLabelText(f"正在下载更新包：{received / 1_048_576:.1f} MB")

    def _close_update_progress_dialog(self) -> None:
        dialog = self._update_progress_dialog
        self._update_progress_dialog = None
        if dialog is not None:
            dialog.close()
            dialog.deleteLater()

    def _finish_update_check_thread(self) -> None:
        self._update_check_thread = None

    def _finish_update_download_thread(self) -> None:
        self._update_download_thread = None

    def _on_update_download_completed(self, archive: Path) -> None:
        self._close_update_progress_dialog()
        try:
            self._launch_updater(Path(archive))
        except Exception as exc:  # noqa: BLE001
            self.update_button.setEnabled(True)
            self.status_label.setText("更新器启动失败。")
            QMessageBox.critical(self, "更新失败", str(exc))

    def _on_update_download_failed(self, message: str) -> None:
        self._close_update_progress_dialog()
        self.update_button.setEnabled(True)
        self.status_label.setText("更新下载或校验失败。")
        QMessageBox.warning(self, "更新失败", message)

    def _launch_updater(self, archive: Path) -> None:
        if not getattr(sys, "frozen", False):
            raise RuntimeError("应用内更新仅适用于已安装的软件版本，请先使用发布安装包安装。")
        executable = Path(sys.executable).resolve()
        if sys.platform == "win32":
            source_updater = executable.parent / "YouTubeSubtitleVerifier-Updater.exe"
            install_root = executable.parent
            executable_name = executable.name
            platform_name = "windows"
        elif sys.platform == "darwin":
            source_updater = executable.parent / "YouTubeSubtitleVerifier-Updater"
            app_bundle = next((parent for parent in executable.parents if parent.suffix == ".app"), None)
            if app_bundle is None:
                raise RuntimeError("无法确定 macOS 应用安装目录。")
            install_root = app_bundle
            executable_name = ""
            platform_name = "macos"
        else:
            raise RuntimeError("当前系统不支持应用内更新。")
        if not source_updater.is_file():
            raise RuntimeError("当前安装包缺少更新组件，请先下载安装一次最新完整安装包。")
        updater_dir = Path(tempfile.mkdtemp(prefix="youtube-subtitle-verifier-updater-"))
        updater_copy = updater_dir / source_updater.name
        shutil.copy2(source_updater, updater_copy)
        if sys.platform != "win32":
            updater_copy.chmod(updater_copy.stat().st_mode | 0o111)
        command = [
            str(updater_copy),
            "--platform", platform_name,
            "--pid", str(os.getpid()),
            "--archive", str(archive),
            "--install-root", str(install_root),
        ]
        if executable_name:
            command.extend(["--executable-name", executable_name])
        subprocess.Popen(command, cwd=str(updater_dir))
        self._persist_workspace_state()
        self.status_label.setText("更新器已启动，软件将在更新完成后自动重启。")
        QTimer.singleShot(150, QApplication.instance().quit)

    def closeEvent(self, event) -> None:  # type: ignore[override]
        local_thread = self._thread
        if local_thread is not None and local_thread.isRunning():
            # Keep the durable snapshot active after cancellation so the next
            # launch can offer recovery.
            self._closing_for_recovery = True
        if self._parallel_cover_review_control is not None:
            self._parallel_cover_review_control.cancel()
        parallel_cover_thread = self._parallel_cover_review_thread
        if parallel_cover_thread is not None and parallel_cover_thread.isRunning():
            if not parallel_cover_thread.wait(5000):
                choice = QMessageBox.question(
                    self,
                    "任务仍在停止",
                    "封面检测正在结束网络请求。\n任务状态已保存。是否强制退出？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if choice != QMessageBox.StandardButton.Yes:
                    event.ignore()
                    return

        # Do not let Qt destroy a local worker while its run() method still
        # owns network/process resources. This was the source of the native
        # Qt6Core crash observed after restoring a long ASR task.
        if local_thread is not None and local_thread.isRunning():
            if self._task_control is not None:
                self._task_control.cancel()
            self._persist_workspace_state()
            if not local_thread.wait(5000):
                choice = QMessageBox.question(
                    self,
                    "任务仍在停止",
                    "当前任务正在结束网络请求。\n任务状态已保存。是否强制退出？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if choice != QMessageBox.StandardButton.Yes:
                    event.ignore()
                    return
        if self._matching_thread_is_running():
            self._matching_thread.request_control("cancel")
        self._persist_workspace_state()
        super().closeEvent(event)

    def _refresh_dashboard_metrics(self) -> None:
        ready_ids = {
            str(item.get("source_video_id") or "")
            for item in self._ready_items
            if str(item.get("source_video_id") or "")
        }
        if hasattr(self, "side_total_metric"):
            self.side_total_metric.setText(f"队列\n{len(self._videos)}")
            self.side_ready_metric.setText(f"可匹配\n{len(ready_ids)}")
            self.side_fallback_metric.setText(f"待兜底\n{len(self._pending_asr)}")
        self._refresh_workspace_pages()

    def _workspace_snapshot(self) -> dict[str, object]:
        task_kind = ""
        if isinstance(self._thread, _CollectThread):
            task_kind = "视频采集"
        elif isinstance(self._thread, _CoverThread):
            task_kind = "封面下载"
        elif isinstance(self._thread, _CoverReviewThread):
            task_kind = "封面检测"
        elif isinstance(self._thread, _PrepareThread):
            task_kind = "字幕获取"
        elif isinstance(self._thread, _AsrFallbackThread):
            task_kind = "ASR 兜底"
        elif isinstance(self._thread, _MatchThread):
            task_kind = "匹配核验"
        elif isinstance(self._thread, _StandaloneDownloadThread):
            task_kind = "视频下载"
        elif isinstance(self._thread, _StandaloneSubtitleThread):
            task_kind = "字幕转写"
        elif isinstance(self._thread, _StandaloneCoverThread):
            task_kind = "封面检测"
        matching_active = self._matching_thread_is_running()
        if matching_active:
            task_kind = f"{task_kind} + 匹配核验" if task_kind else "匹配核验"
        parallel_cover_active = self._parallel_cover_review_is_running()
        if parallel_cover_active:
            task_kind = f"{task_kind} + 封面检测" if task_kind else "封面检测"
        active = self._thread_is_running() or matching_active or parallel_cover_active
        # Do not erase the recovery snapshot when a worker acknowledges
        # cancellation while the window is closing.
        if self._closing_for_recovery and (self._videos or self._pending_asr):
            active = True
        ready_ids = {
            str(item.get("source_video_id") or "")
            for item in self._ready_items
            if str(item.get("source_video_id") or "")
        }
        return {
            "videos": len(self._videos),
            "downloads": self._download_page.queue_table.rowCount() if hasattr(self, "_download_page") else 0,
            "download_completed": sum(
                1 for result in self._standalone_download_results if bool(result.get("ok"))
            ),
            "ready": len(ready_ids),
            "fallback": len(self._pending_asr),
            "reviews": len(self._cover_review_results),
            "active": active,
            "task_kind": task_kind,
            "status": self.status_label.text() if hasattr(self, "status_label") else "当前没有运行任务",
        }

    def _refresh_workspace_pages(self) -> None:
        if not hasattr(self, "_dashboard_page"):
            return
        snapshot = self._workspace_snapshot()
        self._dashboard_page.refresh(snapshot)
        self._task_center_page.refresh(snapshot)
        if snapshot.get("active") and not self._state_persist_timer.isActive():
            self._state_persist_timer.start(1200)

    def _persist_workspace_state(self) -> None:
        if not hasattr(self, "_dashboard_page"):
            return
        # A newly launched window has no local rows yet. Do not let its close
        # event overwrite a recoverable snapshot before the startup prompt has
        # been displayed and answered.
        if not self._recovery_prompt_shown and not (self._videos or self._pending_asr):
            existing_state = self._state_store.load()
            current_cover_items = self._cover_page.items if hasattr(self, "_cover_page") else []
            if isinstance(existing_state, dict) and has_recoverable_work(existing_state) and not current_cover_items:
                LOGGER.info("Preserving recoverable workspace before startup recovery prompt.")
                return
        snapshot = self._workspace_snapshot()
        download_jobs: list[dict[str, str]] = []
        if hasattr(self, "_download_page"):
            for row in range(self._download_page.queue_table.rowCount()):
                values = []
                for column in (0, 1, 2, 3, 4):
                    item = self._download_page.queue_table.item(row, column)
                    values.append(item.text() if item is not None else "")
                download_jobs.append(
                    {
                        "source": values[0],
                        "content": values[1],
                        "range": values[2],
                        "status": values[3],
                        "output": values[4],
                    }
                )
        subtitle_jobs: list[dict[str, str]] = []
        if hasattr(self, "_subtitle_page"):
            for row, source in enumerate(self._subtitle_page._sources):  # noqa: SLF001
                status_item = self._subtitle_page.queue_table.item(row, 3)
                subtitle_jobs.append(
                    {
                        "source": source,
                        "status": status_item.text() if status_item is not None else "等待识别",
                    }
                )
        payload = {
            "version": 1,
            "active": bool(snapshot.get("active")),
            "task_kind": str(snapshot.get("task_kind") or ""),
            "status": str(snapshot.get("status") or ""),
            "videos": [video.to_dict() for video in self._videos],
            "ready_items": self._ready_items,
            "matching_result_rows": self._matching_result_rows,
            "pending_asr": self._pending_asr,
            "cover_paths": self._cover_paths,
            "cover_review_results": [result.to_dict() for result in self._cover_review_results],
            "cover_queue_items": self._cover_page.items if hasattr(self, "_cover_page") else [],
            "download_results": self._standalone_download_results,
            "download_jobs": download_jobs,
            "subtitle_jobs": subtitle_jobs,
            "task_spec": dict(self._task_spec),
        }
        try:
            self._state_store.save(payload)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Unable to persist workspace state: %s", exc)

    def _offer_workspace_recovery(self) -> None:
        if self._recovery_prompt_shown:
            return
        self._recovery_prompt_shown = True
        state = self._state_store.load()
        if not isinstance(state, dict):
            return
        status = str(state.get("status") or "")
        if not has_recoverable_work(state):
            return
        task_kind = str(state.get("task_kind") or "批处理")
        status = status or "上次任务未完成"
        cover_queue_count = len(state.get("cover_queue_items")) if isinstance(state.get("cover_queue_items"), list) else 0
        video_count = len(state.get("videos")) if isinstance(state.get("videos"), list) else 0
        count_hint = f"封面队列 {cover_queue_count} 条" if cover_queue_count else f"视频 {video_count} 条"
        answer = QMessageBox.question(
            self,
            "检测到未完成任务",
            f"检测到上次未完成的{task_kind}任务。\n当前状态：{status}\n{count_hint}\n是否恢复上次的素材和处理结果？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if answer != QMessageBox.StandardButton.Yes:
            state["active"] = False
            state["status"] = "用户选择不恢复"
            self._state_store.save(state)
            return

        # Acknowledge recovery before rebuilding the UI. The next snapshot
        # must describe an active recovery instead of the stale prompt.
        state["active"] = True
        state["status"] = "正在恢复上次未完成任务..."
        self._state_store.save(state)

        videos: list[YouTubeVideo] = []
        for raw in state.get("videos") if isinstance(state.get("videos"), list) else []:
            if not isinstance(raw, dict):
                continue
            try:
                videos.append(YouTubeVideo(**raw))
            except TypeError:
                continue
        if videos:
            self._on_collected(videos)
        raw_ready = state.get("ready_items")
        raw_pending = state.get("pending_asr")
        self._ready_items = VerificationWorkflow().coalesce_video_match_items(
            list(raw_ready) if isinstance(raw_ready, list) else []
        )
        raw_matching_rows = state.get("matching_result_rows")
        self._matching_result_rows = [
            dict(row) for row in raw_matching_rows
            if isinstance(row, dict)
        ] if isinstance(raw_matching_rows, list) else []
        self._pending_asr = list(raw_pending) if isinstance(raw_pending, list) else []
        raw_cover_paths = state.get("cover_paths")
        self._cover_paths = dict(raw_cover_paths) if isinstance(raw_cover_paths, dict) else {}
        raw_cover_queue = state.get("cover_queue_items")
        if hasattr(self, "_cover_page") and isinstance(raw_cover_queue, list):
            recovered_cover_items = []
            for raw_item in raw_cover_queue:
                if not isinstance(raw_item, dict):
                    continue
                item = dict(raw_item)
                raw_video = item.get("video") if isinstance(item.get("video"), dict) else item
                video_id = str(raw_video.get("video_id") or "") if isinstance(raw_video, dict) else ""
                if video_id and not str(item.get("cover_path") or "").strip():
                    for suffix in (".jpg", ".jpeg", ".png", ".webp"):
                        candidate = COVER_DIR / f"{video_id}{suffix}"
                        if candidate.is_file() and candidate.stat().st_size >= 1024:
                            item["cover_path"] = str(candidate)
                            item["cover_status"] = "downloaded"
                            break
                # A process restart cannot have an active worker. Convert
                # transient states left in the snapshot back to retryable.
                if not str(item.get("cover_path") or "").strip() and item.get("cover_status") in {"downloading", "下载中"}:
                    item["cover_status"] = "pending"
                if item.get("review_status") in {"reviewing", "检测中"}:
                    item["review_status"] = "pending"
                recovered_cover_items.append(item)
            self._cover_page.set_items(recovered_cover_items)
        self._standalone_download_results = (
            list(state.get("download_results"))
            if isinstance(state.get("download_results"), list)
            else []
        )
        self._refresh_downloaded_subtitle_import()
        restored_reviews: list[CoverReviewResult] = []
        for raw in state.get("cover_review_results") if isinstance(state.get("cover_review_results"), list) else []:
            if not isinstance(raw, dict):
                continue
            try:
                restored_reviews.append(
                    CoverReviewResult(
                        video_id=str(raw.get("video_id") or ""),
                        title=str(raw.get("title") or ""),
                        cover_path=str(raw.get("cover_path") or ""),
                        overall_risk=str(raw.get("overall_risk") or "unknown"),
                        risk_tags=tuple(str(tag) for tag in raw.get("risk_tags") or []),
                            summary=str(raw.get("summary") or ""),
                            evidence=str(raw.get("evidence") or ""),
                            confidence=float(raw.get("confidence") or 0),
                            model_response=str(raw.get("model_response") or ""),
                            error=str(raw.get("error") or ""),
                    )
                )
            except (TypeError, ValueError):
                continue
        self._cover_review_results = restored_reviews
        self._refresh_fallback_table()
        self._refresh_table_statuses()
        self._refresh_cover_review_table()
        self._matching_page.set_items(self._ready_items)
        if self._matching_result_rows:
            self._matching_page.set_result_rows(self._matching_result_rows)
        self.status_label.setText(
            f"已恢复上次任务：视频 {len(self._videos)} 条，可匹配视频字幕 {len(self._ready_items)} 条。"
        )
        self._set_busy(False)
        self._resume_recovered_task(state)

    def _resume_recovered_task(self, state: dict[str, object]) -> None:
        """Resume unfinished subtitle work without repeating completed items."""
        task_kind = str(state.get("task_kind") or "")
        raw_spec = state.get("task_spec")
        spec = dict(raw_spec) if isinstance(raw_spec, dict) else {}
        if spec.get("kind") == "asr_fallback" or "ASR" in task_kind:
            if self._pending_asr:
                self.status_label.setText(
                    f"恢复 ASR 兜底：剩余 {len(self._pending_asr)} 条。"
                )
                self._start_asr_thread(list(self._pending_asr), {})
            else:
                self._persist_workspace_state()
            return
        if spec.get("kind") != "prepare" and "字幕获取" not in task_kind:
            return

        ready_ids = {
            str(item.get("source_video_id") or "")
            for item in self._ready_items
            if str(item.get("source_video_id") or "")
        }
        pending_ids = {
            str((item.get("video") or {}).get("video_id") or "")
            for item in self._pending_asr
            if isinstance(item, dict)
        }
        processed_ids = ready_ids | pending_ids
        remaining = [
            video for video in self._videos
            if video.video_id not in processed_ids
        ]
        self._resume_auto_asr = bool(spec.get("allow_asr_fallback"))

        if remaining:
            remaining_ids = {video.video_id for video in remaining}
            for row, video in enumerate(self._videos):
                item = self.video_table.item(row, 0)
                if item is not None:
                    item.setCheckState(
                        Qt.CheckState.Checked
                        if video.video_id in remaining_ids
                        else Qt.CheckState.Unchecked
                    )
            self.status_label.setText(
                f"恢复字幕任务：剩余 {len(remaining)} 条，已保留 {len(processed_ids)} 条结果。"
            )
            self._prepare()
            return

        if self._resume_auto_asr and self._pending_asr:
            self.status_label.setText(
                f"恢复 ASR 兜底：剩余 {len(self._pending_asr)} 条。"
            )
            self._start_asr_thread(list(self._pending_asr), {})
            return
        self._resume_auto_asr = False

    def _navigate_workspace(self, page_id: str) -> None:
        if not hasattr(self, "_page_stack"):
            return
        focus_target = page_id
        if page_id in {"collect"}:
            page_id = "full_flow"
        index = self._page_indices.get(page_id, self._page_indices["dashboard"])
        self._page_stack.setCurrentIndex(index)
        if page_id == "matching" and not self._matching_page._items:  # noqa: SLF001
            self._matching_page.set_items(self._ready_items)
        for current_id, button in self._nav_buttons.items():
            active = current_id == page_id
            button.setChecked(active)
            button.setProperty("nav_active", active)
            button.style().unpolish(button)
            button.style().polish(button)
        if page_id != "full_flow":
            return
        if focus_target == "collect":
            self._legacy_flow_page.ensureWidgetVisible(self.channel_input)
            self.channel_input.setFocus()
        elif focus_target == "download":
            self._legacy_flow_page.ensureWidgetVisible(self.asr_fallback_button)
            self.asr_fallback_button.setFocus()
        elif focus_target == "subtitles":
            self._legacy_flow_page.ensureWidgetVisible(self.prepare_button)
            self.prepare_button.setFocus()
        elif focus_target == "matching":
            self._legacy_flow_page.ensureWidgetVisible(self.match_button)
            self.match_button.setFocus()
        elif focus_target == "cover":
            self._legacy_flow_page.ensureWidgetVisible(self.cover_button)
            self.cover_button.setFocus()

    def _open_connection_settings(self) -> None:
        self._navigate_workspace("full_flow")
        if not self.advanced_settings_panel.isVisible():
            self._toggle_advanced_settings()
        self._legacy_flow_page.ensureWidgetVisible(self.advanced_settings_panel)

    def _refresh_audio_strategy_controls(self) -> None:
        self.browser_concurrency_combo.setEnabled(self.audio_strategy_combo.currentData() == "browser")

    def _import_download_results(self) -> list[str]:
        paths: list[str] = []
        audio_suffixes = {".m4a", ".mp3", ".wav", ".aac", ".ogg", ".flac", ".opus"}
        for result in self._standalone_download_results:
            if not result.get("ok"):
                continue
            raw_paths = result.get("paths")
            if isinstance(raw_paths, list):
                valid_paths = [Path(str(path)) for path in raw_paths if str(path).strip() and Path(str(path)).is_file()]
                if not valid_paths:
                    continue
                preferred = next((path for path in valid_paths if path.suffix.lower() in audio_suffixes), valid_paths[0])
                paths.append(str(preferred))
        return paths

    def _refresh_downloaded_subtitle_import(self) -> None:
        if hasattr(self, "_subtitle_page"):
            self._subtitle_page.set_downloaded_material_count(len(self._import_download_results()))

    def _import_cover_items(self) -> list[dict[str, object]]:
        videos = self._selected_videos() if self._videos else []
        if not videos:
            videos = list(self._videos)
        reviews_by_video_id = {
            result.video_id: result.to_dict()
            for result in self._cover_review_results
            if result.video_id
        }
        return [
            {
                "video": video.to_dict(),
                "cover_path": str(self._cover_paths.get(video.video_id) or ""),
                "review": reviews_by_video_id.get(video.video_id, {}),
            }
            for video in videos
        ]

    def _collect_cover_channels(self, urls: list[str], limit: int) -> None:
        if self._thread_is_running() or self._cover_collect_thread is not None and self._cover_collect_thread.isRunning():
            self._cover_page.set_status("已有任务正在运行，请等待完成或先取消。")
            return
        self._cover_page.set_busy(True)
        self._cover_page.set_status(f"正在采集频道：0/{len(urls)}...")
        self._task_control = TaskControl()
        thread = _CoverCollectThread(urls, limit, self._task_control)
        self._cover_collect_thread = thread
        thread.progress.connect(
            lambda index, total, url, added, error: self._cover_page.set_status(
                f"采集频道：{index}/{total}，本频道新增 {added} 条" + (f"；失败：{error}" if error else "")
            )
        )
        thread.batch.connect(self._on_cover_channel_batch)
        thread.succeeded.connect(self._on_cover_channels_collected)
        thread.failed.connect(lambda error: self._cover_page.set_status(f"频道采集失败：{error}"))
        thread.finished.connect(lambda: self._finish_cover_collection(thread))
        thread.start()

    def _finish_cover_collection(self, thread: _CoverCollectThread) -> None:
        if self._cover_collect_thread is thread:
            self._cover_collect_thread = None
        self._task_control = None
        self._cover_page.set_busy(False)

    def _on_cover_channels_collected(self, items: list[dict[str, object]]) -> None:
        if not self._cover_page.items:
            self._cover_page.set_items(items)
        self._cover_page.set_status(f"频道采集完成：共获得 {len(items)} 条视频，可选择后下载封面。")

    def _on_cover_channel_batch(self, items: list[dict[str, object]]) -> None:
        if not isinstance(items, list) or not items:
            return
        self._cover_page.append_items([item for item in items if isinstance(item, dict)])
        self._persist_workspace_state()

    def _start_standalone_cover(
        self,
        items: list[dict[str, object]],
        download_cover: bool,
        detect_cover: bool,
    ) -> None:
        if self._thread_is_running():
            self._cover_page.set_status("已有封面任务正在运行，请先暂停或取消当前任务。")
            QMessageBox.information(self, "任务进行中", "请等待当前任务完成，或先暂停/取消当前任务。")
            return
        self._task_control = TaskControl()
        self._standalone_cover_progress_total = len(items)
        self._standalone_cover_progress_completed = 0
        self._set_busy(True)
        self.status_label.setText(f"正在处理 {len(items)} 条封面素材...")
        self._cover_page.set_status(f"正在处理 0/{len(items)} 条封面素材...")
        self._set_cover_progress(0, len(items), "封面处理")
        self._thread = _StandaloneCoverThread(
            items, download_cover, detect_cover, self._task_control,
            download_concurrency=int(self._cover_page.download_concurrency_combo.currentData() or 1),
        )
        self._thread.item_updated.connect(self._on_standalone_cover_item)
        self._thread.completed.connect(self._on_standalone_cover_completed)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _start_cover_page_review(self, items: list[dict[str, object]]) -> None:
        videos: list[YouTubeVideo] = []
        for item in items:
            raw = item.get("video") if isinstance(item.get("video"), dict) else item
            cover_path = str(item.get("cover_path") or "").strip()
            if not isinstance(raw, dict) or not cover_path:
                continue
            try:
                video = YouTubeVideo(**raw)
                # Restored independent cover queues may contain the durable
                # path even when the legacy global map was not rebuilt.
                self._cover_paths[video.video_id] = cover_path
                videos.append(video)
            except TypeError:
                continue
        if not videos:
            self._cover_page.set_status("没有可检测的已下载封面。")
            return
        if self._parallel_cover_review_is_running():
            self._cover_page.set_status("已有封面检测任务正在运行。")
            return
        self._start_parallel_cover_review(
            videos,
            concurrency=self._cover_page.review_concurrency_combo.currentData(),
        )

    def _start_cover_page_url_review(self, items: list[dict[str, object]]) -> None:
        videos: list[YouTubeVideo] = []
        for item in items:
            raw = item.get("video") if isinstance(item.get("video"), dict) else item
            if not isinstance(raw, dict):
                continue
            try:
                video = YouTubeVideo(**raw)
            except TypeError:
                continue
            self._cover_paths[video.video_id] = str(item.get("cover_path") or "")
            videos.append(video)
        if not videos:
            self._cover_page.set_status("没有可进行 URL 优先检测的项目。")
            return
        if self._parallel_cover_review_is_running():
            self._cover_page.set_status("已有封面检测任务正在运行。")
            return
        self._start_parallel_cover_review(videos, prefer_url=True)

    def _remove_cover_queue_items(self, video_ids: set[str]) -> None:
        ids = {str(video_id or "") for video_id in video_ids if str(video_id or "")}
        if not ids:
            return
        self._cover_paths = {
            video_id: path for video_id, path in self._cover_paths.items() if video_id not in ids
        }
        self._cover_review_results = [
            result for result in self._cover_review_results if result.video_id not in ids
        ]
        self._refresh_cover_review_table()
        self._persist_workspace_state()

    def _on_standalone_cover_item(self, row: int, status: str, cover_path: str, result: object) -> None:
        self._cover_page.update_job(row, status, cover_path, result)
        self._cover_page.refresh_stage_views()
        # Persist each completed item (debounced) so a restart never falls
        # back to an old 2971/463-style snapshot.
        if not self._state_persist_timer.isActive():
            self._state_persist_timer.start(500)
        self._standalone_cover_progress_completed += 1
        total = max(1, self._standalone_cover_progress_total)
        finished = min(self._standalone_cover_progress_completed, total)
        message = f"封面处理：{finished}/{total}，当前第 {row + 1} 条 {status}"
        self.status_label.setText(message)
        self._cover_page.set_status(message)
        self._set_cover_progress(finished, total, f"封面处理 · 当前第 {row + 1} 条")
        self._refresh_workspace_pages()

    def _on_standalone_cover_completed(self, results: list[dict[str, object]], cancelled: bool) -> None:
        for result in results:
            raw_video = result.get("video")
            if not isinstance(raw_video, dict):
                continue
            video_id = str(raw_video.get("video_id") or "")
            cover_path = str(result.get("cover_path") or "")
            if video_id and cover_path:
                self._cover_paths[video_id] = cover_path
            review = result.get("review")
            if isinstance(review, dict):
                try:
                    self._cover_review_results = [
                        current for current in self._cover_review_results
                        if current.video_id != video_id
                    ]
                    self._cover_review_results.append(
                        CoverReviewResult(
                            video_id=video_id,
                            title=str(review.get("title") or raw_video.get("title") or ""),
                            cover_path=str(review.get("cover_path") or cover_path),
                            overall_risk=str(review.get("overall_risk") or "unknown"),
                            risk_tags=tuple(str(tag) for tag in review.get("risk_tags") or []),
                            summary=str(review.get("summary") or ""),
                            evidence=str(review.get("evidence") or ""),
                            confidence=float(review.get("confidence") or 0),
                            model_response=str(review.get("model_response") or ""),
                            error=str(review.get("error") or ""),
                        )
                    )
                except (TypeError, ValueError):
                    pass
        self._refresh_cover_review_table()
        self._cover_page.refresh_stage_views()
        self._cover_page.set_items(self._cover_page.items)
        self._task_control = None
        self._cover_page.pause_button.setText("暂停")
        message = (
            f"封面任务已取消，已完成 {len(results)} 条。"
            if cancelled
            else f"封面处理完成：已处理 {len(results)} 条。"
        )
        self.status_label.setText(message)
        self._cover_page.set_status(message)
        self._set_cover_progress(len(results), len(results), "封面处理完成" if not cancelled else "封面处理已取消")
        self._set_busy(False)

    def _start_standalone_transcription(
        self,
        sources: list[str],
        seconds: int,
        allow_asr_fallback: bool,
        skip_caption_probe: bool = False,
    ) -> None:
        if self._thread_is_running():
            QMessageBox.information(self, "任务进行中", "请等待当前任务完成，或先暂停/取消当前任务。")
            return
        self._task_control = TaskControl()
        self._set_busy(True)
        mode = "直接下载并 ASR 转写" if skip_caption_probe else "获取字幕并按需转写"
        self.status_label.setText(f"正在处理 {len(sources)} 条字幕素材（{mode}）...")
        self._subtitle_page.set_status(f"正在处理 0/{len(sources)} 条素材...")
        self._thread = _StandaloneSubtitleThread(
            sources,
            seconds,
            allow_asr_fallback,
            self._task_control,
            skip_caption_probe=skip_caption_probe,
        )
        self._thread.item_updated.connect(self._on_standalone_subtitle_item)
        self._thread.completed.connect(self._on_standalone_subtitle_completed)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _on_standalone_subtitle_item(self, row: int, status: str, source_kind: str, text: str) -> None:
        self._subtitle_page.update_job(row, status, source_kind, text)
        total = self._subtitle_page.queue_table.rowCount()
        finished = sum(
            1
            for index in range(total)
            if (item := self._subtitle_page.queue_table.item(index, 3)) is not None
            and item.text() in {"识别完成", "识别失败", "待 ASR 兜底"}
        )
        message = f"字幕转写：{finished}/{total}，当前第 {row + 1} 条 {status}"
        self.status_label.setText(message)
        self._subtitle_page.set_status(message)
        self._refresh_workspace_pages()

    def _on_standalone_subtitle_completed(
        self,
        ready: list[dict[str, object]],
        pending: list[dict[str, object]],
        cancelled: bool,
    ) -> None:
        self._ready_items = VerificationWorkflow().coalesce_video_match_items([*self._ready_items, *ready])
        self._pending_asr.extend(pending)
        self._matching_page.set_items(self._ready_items)
        self._task_control = None
        self._subtitle_page.pause_button.setText("暂停")
        if cancelled:
            message = f"字幕任务已取消，已生成 {len(ready)} 条可匹配视频字幕。"
        else:
            message = f"字幕处理完成：生成 {len(ready)} 条可匹配视频字幕，{len(pending)} 条待 ASR 兜底。"
        self.status_label.setText(message)
        self._subtitle_page.set_status(message)
        self._set_busy(False)

    def _import_ready_items(self) -> list[dict[str, object]]:
        self._ready_items = VerificationWorkflow().coalesce_video_match_items(self._ready_items)
        return list(self._ready_items)

    def _import_matching_files(self) -> list[dict[str, object]]:
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "导入字幕文件",
            "",
            "字幕文件 (*.txt *.xlsx);;所有文件 (*.*)",
        )
        items: list[dict[str, object]] = []
        for raw_path in paths:
            path = Path(raw_path)
            if path.suffix.lower() == ".txt":
                try:
                    text = path.read_text(encoding="utf-8", errors="replace").strip()
                except OSError:
                    text = ""
                if text:
                    items.append(
                        {
                            "source_ref": str(path),
                            "source_platform": "local",
                            "source_display_title": path.stem,
                            "source_video_id": path.stem,
                            "source_caption_source": "txt",
                            "query_text": text,
                            "source_text_original": text,
                        }
                    )
                continue
            try:
                from openpyxl import load_workbook

                sheet = load_workbook(path, read_only=True, data_only=True).active
                rows = list(sheet.iter_rows(values_only=True))
            except Exception as exc:  # noqa: BLE001
                LOGGER.warning("Unable to import matching workbook path=%s error=%s", path, exc)
                continue
            if not rows:
                continue
            headers = [str(value or "").strip() for value in rows[0]]
            header_index = {header: index for index, header in enumerate(headers) if header}

            def cell(row: tuple[object, ...], *names: str) -> str:
                for name in names:
                    index = header_index.get(name)
                    if index is not None and index < len(row):
                        return str(row[index] or "").strip()
                return ""

            for row in rows[1:]:
                if not isinstance(row, tuple):
                    row = tuple(row)
                text = cell(row, "完整字幕", "source_text_original", "字幕文本")
                if not text:
                    continue
                source_ref = cell(row, "视频链接", "source_ref") or str(path)
                items.append(
                    {
                        "source_ref": source_ref,
                        "source_description": source_ref,
                        "source_platform": "YouTube" if "youtube" in source_ref.lower() else "local",
                        "source_display_title": cell(row, "视频标题", "source_display_title") or path.stem,
                        "source_video_id": cell(row, "视频 ID", "source_video_id") or path.stem,
                        "source_channel": cell(row, "频道", "source_channel"),
                        "source_caption_language": cell(row, "字幕语言", "source_caption_language"),
                        "source_caption_source": cell(row, "字幕来源", "source_caption_source") or "xlsx",
                        "query_text": text,
                        "source_text_original": text,
                    }
                )
        return items

    def _start_standalone_matching(
        self,
        items: list[dict[str, object]],
        top_k: int,
    ) -> None:
        credentials = self._matching_service_credentials()
        if credentials is None:
            self._matching_page.set_status("请先在“匹配配置”中完成服务配置。")
            return
        server, username, password = credentials
        if self._thread_is_running():
            QMessageBox.information(self, "任务进行中", "请等待本地字幕、下载或封面任务完成后再提交匹配。")
            return
        if self._matching_thread_is_running():
            QMessageBox.information(self, "任务进行中", "请等待当前任务完成，或先暂停/取消当前任务。")
            return
        self._match_paused = False
        self._set_busy(True, scope="matching")
        items = VerificationWorkflow().coalesce_video_match_items(items)
        source_video_count = len(items)
        submit_message = f"正在提交 {source_video_count} 条视频字幕到匹配服务..."
        self.status_label.setText(submit_message)
        self._matching_page.set_status(submit_message)
        self._active_matching_items = [dict(item) for item in items]
        thread = _MatchThread(items, server, username, password, top_k=top_k)
        self._matching_thread = thread
        thread.task_created.connect(self._on_match_task_created)
        thread.progress.connect(self._on_match_progress)
        thread.succeeded.connect(self._on_matched)
        thread.failed.connect(self._on_matching_worker_failed)
        thread.finished.connect(lambda: self._finish_matching_thread(thread))
        thread.start()

    @staticmethod
    def _matching_source_video_count(items: list[dict[str, object]]) -> int:
        source_keys: set[str] = set()
        for index, item in enumerate(items, start=1):
            key = str(item.get("source_video_id") or item.get("source_ref") or f"item-{index}").strip()
            source_keys.add(key.split("#segment-", 1)[0] or f"item-{index}")
        return len(source_keys)

    def _load_matching_evidence_context(self, window_uid: str) -> None:
        page = self._matching_page
        credentials = self._matching_service_credentials()
        if credentials is None:
            page.set_status("请先在“匹配配置”中完成服务配置，再加载扩展字幕上下文。")
            page.set_evidence_context_error()
            return
        server, username, password = credentials
        if self._evidence_thread is not None and self._evidence_thread.isRunning():
            page.set_status("正在加载另一条匹配字幕上下文，请稍后重试。")
            page.set_evidence_context_error()
            return
        page.set_status("正在加载匹配库的扩展字幕上下文...")
        thread = _EvidenceContextThread(server, username, password, window_uid)
        self._evidence_thread = thread
        thread.succeeded.connect(self._on_matching_evidence_context_loaded)
        thread.failed.connect(self._on_matching_evidence_context_failed)
        # Do not retain a deleted QThread wrapper: a second click previously
        # could access an already-deleted object and make the action appear dead.
        thread.finished.connect(lambda: self._finish_evidence_context_thread(thread))
        thread.start()

    def _finish_evidence_context_thread(self, thread: _EvidenceContextThread) -> None:
        if self._evidence_thread is thread:
            self._evidence_thread = None
        thread.deleteLater()

    def _on_matching_evidence_context_loaded(self, window_uid: str, text: str) -> None:
        self._matching_page.set_evidence_context(window_uid, text)
        self._matching_page.set_status("已加载匹配库扩展字幕上下文。")

    def _on_matching_evidence_context_failed(self, message: str) -> None:
        self._matching_page.set_evidence_context_error()
        self._matching_page.set_status(f"扩展字幕上下文加载失败：{message}")

    def _collect(self) -> None:
        self._save_proxy()
        url = self.channel_input.text().strip()
        if not url:
            QMessageBox.warning(self, "缺少链接", "请输入 YouTube 频道链接。")
            return
        self._set_busy(True)
        self.status_label.setText("正在采集公开视频目录...")
        self._thread = _CollectThread(url, self.limit_spin.value())
        self._thread.succeeded.connect(self._on_collected)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _on_collected(self, videos: list[YouTubeVideo]) -> None:
        existing_ids = {video.video_id for video in self._videos}
        new_videos = [video for video in videos if video.video_id not in existing_ids]
        start_row = self.video_table.rowCount()
        if new_videos:
            for row in range(start_row):
                item = self.video_table.item(row, 0)
                if item is not None:
                    item.setCheckState(Qt.CheckState.Unchecked)
        self._videos.extend(new_videos)
        self.video_table.setRowCount(start_row + len(new_videos))
        for row, video in enumerate(new_videos, start=start_row):
            check = QTableWidgetItem()
            check.setCheckState(Qt.CheckState.Checked)
            self.video_table.setItem(row, 0, check)
            self.video_table.setItem(row, 1, QTableWidgetItem(video.title))
            self.video_table.setItem(row, 2, QTableWidgetItem(video.video_id))
            self.video_table.setItem(row, 3, QTableWidgetItem(video.channel or "未知频道"))
            self.video_table.setItem(row, 4, QTableWidgetItem("待获取字幕"))
            self.video_table.setItem(row, 5, QTableWidgetItem("待下载"))
        self.status_label.setText(f"已采集 {len(self._videos)} 条视频，请确认选择后获取前段字幕。")
        self._refresh_channel_selector()
        self._set_busy(False)

    def _refresh_channel_selector(self) -> None:
        selected_channel = str(self.channel_selection_combo.currentData() or "")
        channels = sorted({video.channel.strip() for video in self._videos if video.channel.strip()})
        self.channel_selection_combo.blockSignals(True)
        self.channel_selection_combo.clear()
        self.channel_selection_combo.addItem("选择频道", "")
        for channel in channels:
            self.channel_selection_combo.addItem(channel, channel)
        index = self.channel_selection_combo.findData(selected_channel)
        self.channel_selection_combo.setCurrentIndex(index if index >= 0 else 0)
        self.channel_selection_combo.blockSignals(False)

    def _selection_table(self) -> QTableWidget:
        if self.task_tabs.currentIndex() == 1:
            return self.fallback_table
        if self.task_tabs.currentIndex() == 2:
            return self.cover_review_table
        return self.video_table

    def _on_task_tab_changed(self, index: int) -> None:
        labels = {
            0: "\u5220\u9664\u9009\u4e2d\u89c6\u9891",
            1: "\u5220\u9664\u9009\u4e2d\u5f85\u515c\u5e95",
            2: "\u5220\u9664\u9009\u4e2d\u68c0\u6d4b\u8bb0\u5f55",
        }
        self.delete_selected_button.setText(labels.get(index, "\u5220\u9664\u9009\u4e2d"))

    def _set_checked_rows(self, checked_rows: set[int]) -> None:
        table = self._selection_table()
        for row in range(table.rowCount()):
            item = table.item(row, 0)
            if item is not None:
                item.setCheckState(
                    Qt.CheckState.Checked if row in checked_rows else Qt.CheckState.Unchecked
                )

    def _select_all(self) -> None:
        table = self._selection_table()
        self._set_checked_rows(set(range(table.rowCount())))

    def _invert_selection(self) -> None:
        table = self._selection_table()
        for row in range(table.rowCount()):
            item = table.item(row, 0)
            if item is not None:
                item.setCheckState(
                    Qt.CheckState.Unchecked
                    if item.checkState() == Qt.CheckState.Checked
                    else Qt.CheckState.Checked
                )

    def _select_first_ten(self) -> None:
        table = self._selection_table()
        count = max(1, int(self.select_count_spin.value()))
        self._set_checked_rows(set(range(min(count, table.rowCount()))))

    def _select_channel(self) -> None:
        channel = str(self.channel_selection_combo.currentData() or "").strip()
        if not channel:
            QMessageBox.information(self, "请选择频道", "请先从下拉列表中选择一个频道。")
            return
        table = self._selection_table()
        if table is self.fallback_table:
            self._set_checked_rows({
                row for row in range(table.rowCount())
                if table.item(row, 3) is not None and table.item(row, 3).text().strip() == channel
            })
            return
        if table is self.cover_review_table:
            channel_video_ids = {
                video.video_id for video in self._videos if video.channel.strip() == channel
            }
            self._set_checked_rows({
                row for row in range(table.rowCount())
                if table.item(row, 2) is not None
                and table.item(row, 2).text().strip() in channel_video_ids
            })
            return
        self._set_checked_rows({
            row for row, video in enumerate(self._videos) if video.channel.strip() == channel
        })

    def _enqueue_selected_for_fallback(self) -> None:
        """Move selected videos into the manual ASR fallback queue."""
        if self._thread_is_running() or self._matching_thread_is_running() or self._parallel_cover_review_is_running():
            QMessageBox.information(
                self,
                "\u4efb\u52a1\u8fdb\u884c\u4e2d",
                "\u5f53\u524d\u6709\u4efb\u52a1\u6b63\u5728\u4f7f\u7528\u5217\u8868\u6570\u636e\uff0c\u8bf7\u7b49\u4efb\u52a1\u5b8c\u6210\u540e\u518d\u64cd\u4f5c\u3002",
            )
            return
        videos = self._selected_videos()
        if not videos:
            QMessageBox.information(self, "\u672a\u9009\u62e9", "\u8bf7\u5148\u5728\u89c6\u9891\u961f\u5217\u4e2d\u52fe\u9009\u8981\u8f6c\u5165\u5f85\u515c\u5e95\u7684\u89c6\u9891\u3002")
            return

        selected_ids = {video.video_id for video in videos}
        pending_by_id = {
            str((item.get("video") or {}).get("video_id") or ""): item
            for item in self._pending_asr
        }
        added_count = 0
        for video in videos:
            if video.video_id not in pending_by_id:
                pending_by_id[video.video_id] = {
                    "video": video.to_dict(),
                    "inspection": {
                        "status": "asr_required",
                        "asr_status": "pending",
                        "asr_error": "\u7528\u6237\u624b\u52a8\u9009\u62e9\u8fdb\u5165\u5f85\u515c\u5e95",
                    },
                }
                added_count += 1
        self._pending_asr = list(pending_by_id.values())
        self._ready_items = [
            item for item in self._ready_items
            if str(item.get("source_video_id") or "") not in selected_ids
        ]
        self._refresh_fallback_table()
        for row in range(self.fallback_table.rowCount()):
            video_id_item = self.fallback_table.item(row, 2)
            if video_id_item is not None and video_id_item.text().strip() in selected_ids:
                check = self.fallback_table.item(row, 0)
                if check is not None:
                    check.setCheckState(Qt.CheckState.Checked)
        self._refresh_table_statuses()
        self._matching_page.set_items(self._ready_items)
        self._persist_workspace_state()
        self.task_tabs.setCurrentIndex(1)
        self.status_label.setText(
            f"\u5df2\u5c06 {added_count} \u6761\u89c6\u9891\u52a0\u5165\u5f85\u515c\u5e95\uff0c\u53ef\u5728\u5f85\u515c\u5e95\u9875\u70b9\u51fb\u4e0b\u8f7d\u5e76\u8f6c\u5199\u3002"
            if added_count
            else "\u9009\u4e2d\u89c6\u9891\u5df2\u5728\u5f85\u515c\u5e95\u5217\u8868\u4e2d\u3002"
        )

    def _delete_selected_items(self) -> None:
        """Remove selected workspace records without deleting source files."""
        if self._thread_is_running() or self._matching_thread_is_running() or self._parallel_cover_review_is_running():
            QMessageBox.information(
                self,
                "\u4efb\u52a1\u8fdb\u884c\u4e2d",
                "\u5f53\u524d\u6709\u4efb\u52a1\u6b63\u5728\u4f7f\u7528\u5217\u8868\u6570\u636e\uff0c\u8bf7\u7b49\u4efb\u52a1\u5b8c\u6210\u6216\u53d6\u6d88\u540e\u518d\u5220\u9664\u3002",
            )
            return

        index = self.task_tabs.currentIndex()
        table = self._selection_table()
        if index == 0:
            selected_rows = {
                row for row in range(table.rowCount())
                if table.item(row, 0) is not None
                and table.item(row, 0).checkState() == Qt.CheckState.Checked
            }
            selected_videos = [
                video for row, video in enumerate(self._videos) if row in selected_rows
            ]
            selected_ids = {video.video_id for video in selected_videos}
            if not selected_ids:
                QMessageBox.information(self, "\u672a\u9009\u62e9", "\u8bf7\u5148\u52fe\u9009\u8981\u5220\u9664\u7684\u89c6\u9891\u3002")
                return
            selected_urls = {video.source_url for video in selected_videos}
            prompt = f"\u5c06\u4ece\u5de5\u4f5c\u533a\u79fb\u9664 {len(selected_ids)} \u6761\u89c6\u9891\u53ca\u5176\u5173\u8054\u7684\u5b57\u5e55\u3001\u5c01\u9762\u68c0\u6d4b\u8bb0\u5f55\u548c\u5339\u914d\u8bb0\u5f55\u3002\n\u539f\u59cb\u89c6\u9891\u3001\u97f3\u9891\u548c\u5c01\u9762\u6587\u4ef6\u4e0d\u4f1a\u88ab\u5220\u9664\u3002\n\u662f\u5426\u7ee7\u7eed\uff1f"
            if QMessageBox.question(
                self,
                "\u786e\u8ba4\u5220\u9664\u89c6\u9891",
                prompt,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            ) != QMessageBox.StandardButton.Yes:
                return

            self._videos = [video for video in self._videos if video.video_id not in selected_ids]
            self._ready_items = [
                item for item in self._ready_items
                if str(item.get("source_video_id") or "") not in selected_ids
            ]
            self._pending_asr = [
                item for item in self._pending_asr
                if str((item.get("video") or {}).get("video_id") or "") not in selected_ids
            ]
            self._cover_paths = {
                video_id: path for video_id, path in self._cover_paths.items()
                if video_id not in selected_ids
            }
            self._cover_review_results = [
                result for result in self._cover_review_results
                if result.video_id not in selected_ids
            ]
            self._matching_result_rows = [
                row for row in self._matching_result_rows
                if str(row.get("source_video_id") or row.get("video_id") or "") not in selected_ids
            ]
            self._active_matching_items = [
                item for item in self._active_matching_items
                if str(item.get("source_video_id") or "") not in selected_ids
            ]
            self._active_video_ids.difference_update(selected_ids)
            self._completed_video_ids.difference_update(selected_ids)
            self._standalone_download_results = [
                result for result in self._standalone_download_results
                if str(result.get("url") or "") not in selected_urls
            ]
            for row in sorted(selected_rows, reverse=True):
                self.video_table.removeRow(row)
            if not self._videos:
                self._task_spec = {}
                self.result_edit.clear()
            self._refresh_fallback_table()
            self._refresh_cover_review_table()
            self._refresh_table_statuses()
            self._matching_page.set_items(self._ready_items)
            self._matching_page.set_result_rows(self._matching_result_rows)
            self._refresh_downloaded_subtitle_import()
            self._refresh_channel_selector()
            self._persist_workspace_state()
            self.status_label.setText(f"\u5df2\u5220\u9664 {len(selected_ids)} \u6761\u89c6\u9891\uff0c\u5176\u4ed6\u539f\u59cb\u6587\u4ef6\u4fdd\u7559\u3002")
            self._set_busy(False)
            return

        selected_ids = {
            str(table.item(row, 2).text() or "")
            for row in range(table.rowCount())
            if table.item(row, 0) is not None
            and table.item(row, 0).checkState() == Qt.CheckState.Checked
            and table.item(row, 2) is not None
        }
        selected_ids.discard("")
        if not selected_ids:
            QMessageBox.information(self, "\u672a\u9009\u62e9", "\u8bf7\u5148\u52fe\u9009\u8981\u5220\u9664\u7684\u6761\u76ee\u3002")
            return

        if index == 1:
            prompt = f"\u5c06\u4ece\u5f85\u515c\u5e95\u5217\u8868\u79fb\u9664 {len(selected_ids)} \u6761\u8bb0\u5f55\uff0c\u540e\u7eed\u53ef\u91cd\u65b0\u9009\u62e9\u8be5\u89c6\u9891\u8fdb\u884c\u5904\u7406\u3002\n\u662f\u5426\u7ee7\u7eed\uff1f"
            if QMessageBox.question(
                self,
                "\u786e\u8ba4\u5220\u9664\u5f85\u515c\u5e95\u8bb0\u5f55",
                prompt,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            ) != QMessageBox.StandardButton.Yes:
                return
            self._pending_asr = [
                item for item in self._pending_asr
                if str((item.get("video") or {}).get("video_id") or "") not in selected_ids
            ]
            for row, video in enumerate(self._videos):
                if video.video_id in selected_ids:
                    self.video_table.setItem(row, 4, QTableWidgetItem("\u5df2\u79fb\u9664\u515c\u5e95\uff0c\u5f85\u91cd\u65b0\u5904\u7406"))
            self._refresh_fallback_table()
            self._refresh_table_statuses()
            self._persist_workspace_state()
            self.status_label.setText(f"\u5df2\u5220\u9664 {len(selected_ids)} \u6761\u5f85\u515c\u5e95\u8bb0\u5f55\u3002")
            self._set_busy(False)
            return

        prompt = f"\u5c06\u4ece\u5c01\u9762\u68c0\u6d4b\u5217\u8868\u79fb\u9664 {len(selected_ids)} \u6761\u68c0\u6d4b\u8bb0\u5f55\uff0c\u5df2\u4e0b\u8f7d\u7684\u5c01\u9762\u6587\u4ef6\u4ecd\u4f1a\u4fdd\u7559\u5e76\u53ef\u590d\u7528\u3002\n\u662f\u5426\u7ee7\u7eed\uff1f"
        if QMessageBox.question(
            self,
            "\u786e\u8ba4\u5220\u9664\u68c0\u6d4b\u8bb0\u5f55",
            prompt,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return
        self._cover_review_results = [
            result for result in self._cover_review_results if result.video_id not in selected_ids
        ]
        for row, video in enumerate(self._videos):
            if video.video_id in selected_ids:
                status = "\u5df2\u4e0b\u8f7d\uff0c\u5f85\u68c0\u6d4b" if video.video_id in self._cover_paths else "\u5f85\u4e0b\u8f7d"
                self.video_table.setItem(row, 5, QTableWidgetItem(status))
        self._refresh_cover_review_table()
        self._persist_workspace_state()
        self.status_label.setText(f"\u5df2\u5220\u9664 {len(selected_ids)} \u6761\u5c01\u9762\u68c0\u6d4b\u8bb0\u5f55\uff0c\u5c01\u9762\u6587\u4ef6\u4ecd\u4fdd\u7559\u3002")
        self._set_busy(False)

    def _toggle_pause(self) -> None:
        sender = self.sender()
        if self._parallel_cover_review_is_running() and sender is self._cover_page.pause_button:
            control = self._parallel_cover_review_control
            if control is not None:
                if control.paused:
                    control.resume()
                    self._cover_page.pause_button.setText("暂停")
                    self._cover_page.set_status("已请求继续封面检测。")
                else:
                    control.pause()
                    self._cover_page.pause_button.setText("继续")
                    self._cover_page.set_status("已请求暂停封面检测，将在当前请求结束后暂停。")
            return
        matching_page_control = sender in {
            self._matching_page.pause_button,
            self._matching_page.cancel_button,
        }
        matching_control = matching_page_control or (
            sender is self.pause_button
            and self._matching_thread_is_running()
            and not self._thread_is_running()
        )
        if matching_control and self._matching_thread_is_running():
            thread = self._matching_thread
            if self._match_paused:
                thread.request_control("resume")
                self._match_paused = False
                self.pause_button.setText("暂停")
                self._matching_page.pause_button.setText("暂停")
                self.status_label.setText("已请求继续视频级匹配，将在当前视频请求结束后继续。")
                self._matching_page.set_status("已请求继续视频级匹配，将在当前视频请求结束后继续。")
            else:
                thread.request_control("pause")
                self._match_paused = True
                self.pause_button.setText("继续")
                self._matching_page.pause_button.setText("继续")
                self.status_label.setText("已请求暂停视频级匹配，将在当前视频请求结束后暂停。")
                self._matching_page.set_status("已请求暂停视频级匹配，将在当前视频请求结束后暂停。")
            return
        if self._task_control is None or self._task_control.cancelled:
            return
        if self._task_control.paused:
            self._task_control.resume()
            self.pause_button.setText("暂停")
            self.status_label.setText("任务已继续，将在当前视频处理完成后更新进度。")
            if isinstance(self._thread, _StandaloneDownloadThread):
                self._download_page.pause_button.setText("暂停")
                self._download_page.set_status("下载任务已继续。")
            if isinstance(self._thread, _StandaloneSubtitleThread):
                self._subtitle_page.pause_button.setText("暂停")
                self._subtitle_page.set_status("字幕任务已继续。")
        else:
            self._task_control.pause()
            self.pause_button.setText("继续")
            self.status_label.setText("将在当前视频处理完成后暂停。")
            if isinstance(self._thread, _StandaloneDownloadThread):
                self._download_page.pause_button.setText("继续")
                self._download_page.set_status("将在当前下载检查点暂停。")
            if isinstance(self._thread, _StandaloneSubtitleThread):
                self._subtitle_page.pause_button.setText("继续")
                self._subtitle_page.set_status("将在当前字幕素材处理完成后暂停。")

    def _cancel_task(self) -> None:
        sender = self.sender()
        matching_page_control = sender in {
            self._matching_page.pause_button,
            self._matching_page.cancel_button,
        }
        if self._parallel_cover_review_control is not None and self._parallel_cover_review_is_running():
            if not matching_page_control:
                self._parallel_cover_review_control.cancel()
        matching_control = matching_page_control or (
            sender is self.cancel_task_button
            and self._matching_thread_is_running()
            and not self._thread_is_running()
        )
        if matching_control and self._matching_thread_is_running():
            self._matching_thread.request_control("cancel")
            self.pause_button.setEnabled(False)
            self.cancel_task_button.setEnabled(False)
            self._matching_page.pause_button.setEnabled(False)
            self._matching_page.cancel_button.setEnabled(False)
            self.status_label.setText("正在取消视频级匹配，当前视频请求结束后停止。")
            self._matching_page.set_status("正在取消视频级匹配，当前视频请求结束后停止。")
            return
        if self._task_control is None:
            return
        self._task_control.cancel()
        if self._browser_capture_dialog is not None and self._browser_capture_dialog.isVisible():
            self._browser_capture_dialog.reject()
        self.pause_button.setEnabled(False)
        self.cancel_task_button.setEnabled(False)
        self.status_label.setText("正在取消任务，当前视频完成后将停止。")
        if isinstance(self._thread, _StandaloneDownloadThread):
            self._download_page.set_status("正在取消下载任务，当前请求结束后停止。")
        if isinstance(self._thread, _StandaloneSubtitleThread):
            self._subtitle_page.set_status("正在取消字幕任务，当前素材处理结束后停止。")

    def _start_standalone_download(
        self,
        urls: list[str],
        content_kind: str,
        duration_seconds: int,
        audio_concurrency: int,
    ) -> None:
        if self._thread_is_running():
            QMessageBox.information(self, "任务进行中", "请等待当前任务完成，或先暂停/取消当前任务。")
            return
        self._save_proxy()
        self._task_control = TaskControl()
        self._standalone_download_results = []
        self._refresh_downloaded_subtitle_import()
        self._set_busy(True)
        self.status_label.setText(f"正在下载 {len(urls)} 条视频素材...")
        self._download_page.set_status(f"正在下载 0/{len(urls)} 条素材...")
        self._thread = _StandaloneDownloadThread(
            urls,
            content_kind=content_kind,
            duration_seconds=duration_seconds,
            audio_concurrency=audio_concurrency,
            control=self._task_control,
        )
        self._thread.item_updated.connect(self._on_standalone_download_item)
        self._thread.completed.connect(self._on_standalone_download_completed)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _on_standalone_download_item(self, row: int, status: str, output: str) -> None:
        self._download_page.update_job(row, status, output)
        total = self._download_page.queue_table.rowCount()
        completed = sum(
            1
            for index in range(total)
            if (item := self._download_page.queue_table.item(index, 3)) is not None
            and item.text() in {"下载完成", "下载失败"}
        )
        self.status_label.setText(f"视频下载：{completed}/{total}，当前第 {row + 1} 条 {status}")
        self._download_page.set_status(f"视频下载：{completed}/{total}，当前第 {row + 1} 条 {status}")
        self._refresh_workspace_pages()

    def _on_standalone_download_completed(self, results: list[dict[str, object]], cancelled: bool) -> None:
        self._standalone_download_results = results
        self._refresh_downloaded_subtitle_import()
        success_count = sum(1 for result in results if bool(result.get("ok")))
        failed_count = sum(1 for result in results if not bool(result.get("ok")))
        self._task_control = None
        self.pause_button.setText("暂停")
        self._download_page.pause_button.setText("暂停")
        if cancelled:
            message = f"下载任务已取消，已完成 {success_count} 条，失败 {failed_count} 条。"
        else:
            message = f"下载完成：成功 {success_count} 条，失败 {failed_count} 条。"
        self.status_label.setText(message)
        self._download_page.set_status(message)
        self._set_busy(False)

    def _selected_videos(self) -> list[YouTubeVideo]:
        selected: list[YouTubeVideo] = []
        for row, video in enumerate(self._videos):
            item = self.video_table.item(row, 0)
            if item is not None and item.checkState() == Qt.CheckState.Checked:
                selected.append(video)
        return selected

    def _download_covers(self) -> None:
        LOGGER.info("Cover download button clicked.")
        self.status_label.setText("已接收封面下载操作，正在准备任务...")
        selected_count = len(self._selected_videos())
        self._set_cover_progress(0, selected_count or 1, "封面下载 · 正在准备")
        try:
            self._save_proxy()
        except Exception as exc:  # noqa: BLE001
            LOGGER.exception("Unable to save proxy before cover download.")
            self.status_label.setText(f"封面下载未启动：保存代理失败（{type(exc).__name__}）")
            QMessageBox.critical(self, "封面下载未启动", f"保存下载代理时发生错误：\n{exc}")
            return
        videos = self._selected_videos()
        if not videos:
            QMessageBox.warning(self, "没有选择", "请至少勾选一条视频后下载封面。")
            return
        LOGGER.info("Cover download task requested. selected_count=%s", len(videos))
        selected_ids = {item.video_id for item in videos}
        for row, video in enumerate(self._videos):
            if video.video_id in selected_ids:
                self.video_table.setItem(row, 5, QTableWidgetItem("等待下载"))
        self._task_control = TaskControl()
        self._set_busy(True)
        self.status_label.setText(f"正在下载公开视频封面：0/{len(videos)}")
        self._set_cover_progress(0, len(videos), "封面下载")
        self._thread = _CoverThread(videos, self._task_control)
        self._thread.started.connect(self._on_cover_started)
        self._thread.progress.connect(self._on_cover_progress)
        self._thread.succeeded.connect(self._on_covers_downloaded)
        self._thread.cancelled.connect(self._on_covers_cancelled)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _on_cover_button_pressed(self) -> None:
        selected_count = len(self._selected_videos())
        LOGGER.info("Cover download button pressed. selected_count=%s", selected_count)
        self.status_label.setText("已接收封面下载操作，请稍候...")
        self._set_cover_progress(0, selected_count or 1, "封面下载 · 正在准备")

    def _on_cover_started(self, index: int, total: int, video: YouTubeVideo) -> None:
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 5, QTableWidgetItem("正在下载"))
                break
        self.status_label.setText(f"正在下载公开视频封面：已启动 {index}/{total}，当前 {self._short_title(video.title)}")

    def _on_cover_progress(
        self,
        index: int,
        total: int,
        video: YouTubeVideo,
        result: CoverDownloadResult,
    ) -> None:
        if result.path:
            self._cover_paths[video.video_id] = result.path
            status = "已保存"
        else:
            status = "下载失败"
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 5, QTableWidgetItem(status))
                break
        self.status_label.setText(
            f"正在下载公开视频封面：已完成 {index}/{total}，刚完成 {self._short_title(video.title)}"
        )
        self._set_cover_progress(index, total, "封面下载")

    def _on_covers_downloaded(self, results: list[CoverDownloadResult]) -> None:
        success_count = sum(1 for result in results if result.path)
        failure_count = len(results) - success_count
        self._task_control = None
        self.pause_button.setText("暂停")
        self.status_label.setText(
            f"封面下载完成：成功 {success_count} 条，失败 {failure_count} 条。"
        )
        self._set_cover_progress(len(results), len(results), "封面下载完成")
        self._set_busy(False)

    def _on_covers_cancelled(self, results: list[CoverDownloadResult]) -> None:
        success_count = sum(1 for result in results if result.path)
        self._task_control = None
        self.pause_button.setText("暂停")
        self.status_label.setText(f"封面下载已取消，已保留 {success_count} 条成功封面。")
        self._set_cover_progress(len(results), len(self._videos), "封面下载已取消")
        self._set_busy(False)

    def _set_cover_progress(self, completed: int, total: int, label: str) -> None:
        total = max(1, int(total or 1))
        self.cover_progress_bar.setRange(0, total)
        self.cover_progress_bar.setValue(max(0, min(int(completed or 0), total)))
        self.cover_progress_bar.setFormat(f"{label}：%v/%m")

    def _review_covers(self) -> None:
        videos = [video for video in self._selected_videos() if video.video_id in self._cover_paths]
        if not videos:
            QMessageBox.information(self, "没有可检测封面", "请先下载并勾选需要检测的封面。")
            return
        if self._parallel_cover_review_is_running():
            QMessageBox.information(self, "封面检测进行中", "当前已有封面检测任务正在执行。")
            return
        if self._thread_is_running():
            if not isinstance(self._thread, _PrepareThread):
                QMessageBox.information(self, "任务进行中", "封面检测只能与字幕获取同步执行，请等待当前任务完成。")
                return
            self._start_parallel_cover_review(videos)
            return
        self._task_control = TaskControl()
        self._set_busy(True)
        self.status_label.setText(f"正在检测封面：0/{len(videos)}")
        self._thread = _CoverReviewThread(videos, self._cover_paths, self._task_control)
        self._thread.progress.connect(self._on_cover_review_progress)
        self._thread.succeeded.connect(self._on_cover_review_finished)
        self._thread.cancelled.connect(self._on_cover_review_cancelled)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _start_parallel_cover_review(self, videos: list[YouTubeVideo], *, prefer_url: bool = False, concurrency: int | None = None) -> None:
        """Run cover-model calls beside subtitle acquisition without replacing its worker."""
        control = TaskControl()
        thread = _CoverReviewThread(videos, self._cover_paths, control, prefer_url=prefer_url, concurrency=concurrency)
        self._parallel_cover_review_control = control
        self._parallel_cover_review_thread = thread
        thread.progress.connect(self._on_parallel_cover_review_progress)
        thread.succeeded.connect(lambda results: self._on_parallel_cover_review_finished(thread, results, False))
        thread.cancelled.connect(lambda results: self._on_parallel_cover_review_finished(thread, results, True))
        thread.failed.connect(lambda message: self._on_parallel_cover_review_failed(thread, message))
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda: self._finish_parallel_cover_review_thread(thread))
        progress_label = "封面 URL 优先检测" if prefer_url else "封面检测（与字幕同步）"
        self._set_cover_progress(0, len(videos), progress_label)
        self.cover_review_button.setEnabled(False)
        if prefer_url:
            message = f"已启动 URL 优先封面检测：0/{len(videos)}（失败自动回退本地下载，最多 3 路）。"
        else:
            message = f"字幕获取继续进行，已同步启动封面检测：0/{len(videos)}（最多 3 路）。"
        self.status_label.setText(message)
        if hasattr(self, "_cover_page"):
            self._cover_page.set_status(message)
        thread.start()
        # Re-evaluate controls after the QThread has actually started. The
        # earlier busy refresh runs before isRunning() becomes true.
        self._set_busy(self._thread_is_running())
        self._cover_page.pause_button.setEnabled(True)

    def _on_parallel_cover_review_progress(
        self,
        index: int,
        total: int,
        video: YouTubeVideo,
        result: CoverReviewResult,
    ) -> None:
        self._cover_review_results = [
            item for item in self._cover_review_results if item.video_id != result.video_id
        ] + [result]
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 5, QTableWidgetItem("已检测" if not result.error else "检测失败"))
                break
        self._schedule_cover_review_table_refresh()
        if hasattr(self, "_cover_page"):
            self._cover_page.update_review_by_video_id(video.video_id, result)
            self._cover_page.set_status(f"封面检测：{index}/{total}，当前完成 {video.title}")
        self._set_cover_progress(index, total, "封面检测（与字幕同步）")

    def _on_parallel_cover_review_finished(
        self,
        thread: _CoverReviewThread,
        results: list[CoverReviewResult],
        cancelled: bool,
    ) -> None:
        if self._parallel_cover_review_thread is not thread:
            return
        self._refresh_cover_review_table()
        self._set_cover_progress(len(results), len(results) or 1, "封面检测已取消" if cancelled else "封面检测完成")

    def _on_parallel_cover_review_failed(self, thread: _CoverReviewThread, message: str) -> None:
        if self._parallel_cover_review_thread is not thread:
            return
        self._set_cover_progress(0, 1, "封面检测失败")
        QMessageBox.warning(self, "封面检测失败", message)

    def _finish_parallel_cover_review_thread(self, thread: _CoverReviewThread) -> None:
        if self._parallel_cover_review_thread is thread:
            self._parallel_cover_review_thread = None
            self._parallel_cover_review_control = None
            self._set_busy(self._thread_is_running())

    def _on_cover_review_progress(
        self,
        index: int,
        total: int,
        video: YouTubeVideo,
        result: CoverReviewResult,
    ) -> None:
        self._cover_review_results = [
            item for item in self._cover_review_results if item.video_id != result.video_id
        ] + [result]
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 5, QTableWidgetItem("已检测" if not result.error else "检测失败"))
                break
        self._schedule_cover_review_table_refresh()
        self.status_label.setText(
            f"正在检测封面：{index}/{total}，当前 {self._short_title(video.title)}"
        )

    def _refresh_cover_review_table(self) -> None:
        previous_selection = {
            str(self.cover_review_table.item(row, 2).text() or "")
            for row in range(self.cover_review_table.rowCount())
            if self.cover_review_table.item(row, 0) is not None
            and self.cover_review_table.item(row, 0).checkState() == Qt.CheckState.Checked
            and self.cover_review_table.item(row, 2) is not None
        }
        use_default_selection = self.cover_review_table.rowCount() == 0
        self.cover_review_table.setUpdatesEnabled(False)
        self.cover_review_table.setRowCount(len(self._cover_review_results))
        labels = {
            "safe": "安全",
            "review": "待复核",
            "risk": "疑似风险",
            "unknown": "无法判断",
        }
        for row, result in enumerate(self._cover_review_results):
            check = QTableWidgetItem()
            if use_default_selection or result.video_id in previous_selection:
                check.setCheckState(Qt.CheckState.Checked)
            else:
                check.setCheckState(Qt.CheckState.Unchecked)
            self.cover_review_table.setItem(row, 0, check)
            conclusion = "检测失败" if result.error else labels.get(result.overall_risk, "无法判断")
            tags = "、".join(result.risk_tags) if result.risk_tags else "未发现明确标签"
            confidence = "-" if result.error else f"{result.confidence:.0%}"
            evidence = result.error or result.evidence or result.summary or "未返回依据"
            values = (result.title, result.video_id, conclusion, tags, confidence, evidence)
            for column, value in enumerate(values, start=1):
                item = QTableWidgetItem(str(value))
                item.setToolTip(str(value))
                self.cover_review_table.setItem(row, column, item)
        self.task_tabs.setTabText(2, f"封面检测 ({len(self._cover_review_results)})")
        self.cover_review_table.setUpdatesEnabled(True)

    def _schedule_cover_review_table_refresh(self) -> None:
        """Coalesce high-frequency worker updates into bounded UI refreshes."""
        if self._cover_review_refresh_pending:
            return
        self._cover_review_refresh_pending = True

        def refresh() -> None:
            self._cover_review_refresh_pending = False
            self._refresh_cover_review_table()

        QTimer.singleShot(300, refresh)

    def _on_cover_review_finished(self, results: list[CoverReviewResult]) -> None:
        success_count = sum(1 for result in results if not result.error)
        failure_count = len(results) - success_count
        self._task_control = None
        self.pause_button.setText("暂停")
        self._refresh_cover_review_table()
        self.task_tabs.setCurrentIndex(2)
        self.status_label.setText(
            f"封面检测完成：成功 {success_count} 条，失败 {failure_count} 条。"
        )
        self._set_busy(False)

    def _on_cover_review_cancelled(self, results: list[CoverReviewResult]) -> None:
        self._task_control = None
        self.pause_button.setText("暂停")
        self._refresh_cover_review_table()
        self.status_label.setText(f"封面检测已取消，已保留 {len(results)} 条结果。")
        self._set_busy(False)

    def _open_cover_directory(self) -> None:
        COVER_DIR.mkdir(parents=True, exist_ok=True)
        if not QDesktopServices.openUrl(QUrl.fromLocalFile(str(COVER_DIR))):
            QMessageBox.warning(self, "无法打开目录", f"请手动打开封面目录：\n{COVER_DIR}")

    def _open_output_directory(self) -> None:
        output_directory = PROJECT_ROOT / "output"
        output_directory.mkdir(parents=True, exist_ok=True)
        if not QDesktopServices.openUrl(QUrl.fromLocalFile(str(output_directory))):
            QMessageBox.warning(self, "无法打开目录", f"请手动打开输出目录：\n{output_directory}")

    def _choose_export_path(self, title: str, default_path: Path) -> Path | None:
        dialog = ExportDestinationDialog(title, default_path, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None
        return dialog.output_path

    def _prepare(self) -> None:
        self._save_proxy()
        videos = self._selected_videos()
        if not videos:
            QMessageBox.warning(self, "没有选择", "请至少选择一条视频。")
            return
        skip_caption_probe = self.skip_caption_probe_check.isChecked()
        self.status_label.setText(
            "正在直接下载音频并 ASR 转写..."
            if skip_caption_probe
            else "正在获取字幕，不下载完整视频..."
        )
        self._active_video_ids = {video.video_id for video in videos}
        self._completed_video_ids = set()
        self._task_control = TaskControl()
        self._set_busy(True)
        auto_asr = self.auto_asr_fallback_check.isChecked()
        self._task_spec = {
            "kind": "prepare",
            "leading_seconds": int(self.seconds_combo.currentData()),
            "allow_asr_fallback": bool(auto_asr or skip_caption_probe),
            "skip_caption_probe": bool(skip_caption_probe),
            "caption_concurrency": int(self.caption_concurrency_combo.currentData()),
            "download_concurrency": int(self.asr_download_concurrency_combo.currentData()),
            "asr_concurrency": int(self.asr_transcribe_concurrency_combo.currentData()),
        }
        self._thread = _PrepareThread(
            videos,
            int(self.seconds_combo.currentData()),
            self._task_control,
            allow_asr_fallback=auto_asr or skip_caption_probe,
            skip_caption_probe=skip_caption_probe,
            caption_concurrency=int(self.caption_concurrency_combo.currentData()),
            download_concurrency=int(self.asr_download_concurrency_combo.currentData()),
            asr_concurrency=int(self.asr_transcribe_concurrency_combo.currentData()),
        )
        self._thread.succeeded.connect(self._on_prepared)
        self._thread.progress.connect(self._on_prepare_progress)
        self._thread.stage.connect(self._on_asr_stage)
        self._thread.cancelled.connect(self._on_task_cancelled)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._set_busy(True)
        self._thread.start()
        if self.auto_cover_review_check.isChecked():
            cover_videos = [video for video in videos if video.video_id in self._cover_paths]
            if cover_videos:
                self._start_parallel_cover_review(cover_videos)

    def _on_prepared(self, ready: list[dict[str, object]], pending_asr: list[dict[str, object]]) -> None:
        self._merge_preparation_results(ready, pending_asr)
        pending_ids = {str(item["video"]["video_id"]) for item in pending_asr}
        asr_ready_ids = {
            str(item.get("source_video_id") or "")
            for item in ready
            if str(item.get("source_caption_source") or "") == "asr"
        }
        for row, video in enumerate(self._videos):
            if video.video_id not in self._active_video_ids:
                continue
            if video.video_id in pending_ids:
                status = "转写未完成，待重试"
            elif video.video_id in asr_ready_ids:
                status = "ASR 转写完成"
            else:
                status = "字幕已准备"
            self.video_table.setItem(row, 4, QTableWidgetItem(status))
        if pending_asr:
            message = f"字幕准备完成：{len(ready)} 条可匹配，{len(pending_asr)} 条转写未完成，可在待兜底列表重试。"
        else:
            message = f"字幕准备完成：{len(ready)} 条可匹配，直出字幕与自动 ASR 兜底均已完成。"
        self.status_label.setText(message)
        self._set_busy(False)
        if self._resume_auto_asr and self._pending_asr:
            self._resume_auto_asr = False
            self._start_asr_thread(list(self._pending_asr), {})
        else:
            self._resume_auto_asr = False
        self._refresh_table_statuses()

    def _merge_preparation_results(
        self,
        ready: list[dict[str, object]],
        pending_asr: list[dict[str, object]],
    ) -> None:
        active_ids = self._active_video_ids
        self._ready_items = VerificationWorkflow().coalesce_video_match_items([
            item for item in self._ready_items
            if str(item.get("source_video_id") or "") not in active_ids
        ] + list(ready))
        self._pending_asr = [
            item for item in self._pending_asr
            if str((item.get("video") or {}).get("video_id") or "") not in active_ids
        ] + list(pending_asr)
        self._refresh_fallback_table()

    def _refresh_fallback_table(self) -> None:
        previous_selection = {
            str(self.fallback_table.item(row, 2).text() or "")
            for row in range(self.fallback_table.rowCount())
            if self.fallback_table.item(row, 0) is not None
            and self.fallback_table.item(row, 0).checkState() == Qt.CheckState.Checked
            and self.fallback_table.item(row, 2) is not None
        }
        use_default_selection = self.fallback_table.rowCount() == 0
        self.fallback_table.setRowCount(len(self._pending_asr))
        for row, item in enumerate(self._pending_asr):
            video = item.get("video") or {}
            inspection = item.get("inspection") or {}
            video_id = str(video.get("video_id") or "")
            check = QTableWidgetItem()
            if use_default_selection or video_id in previous_selection:
                check.setCheckState(Qt.CheckState.Checked)
            else:
                check.setCheckState(Qt.CheckState.Unchecked)
            reason = str(inspection.get("asr_error") or "")
            if not reason:
                reason = "未获取到 YouTube 直出字幕"
            self.fallback_table.setItem(row, 0, check)
            self.fallback_table.setItem(row, 1, QTableWidgetItem(str(video.get("title") or "")))
            self.fallback_table.setItem(row, 2, QTableWidgetItem(video_id))
            self.fallback_table.setItem(row, 3, QTableWidgetItem(str(video.get("channel") or "未知频道")))
            self.fallback_table.setItem(row, 4, QTableWidgetItem(reason))
        self.task_tabs.setTabText(1, f"待兜底 ({len(self._pending_asr)})")

    def _selected_pending_asr_items(self) -> list[dict[str, object]]:
        selected_ids = {
            str(self.fallback_table.item(row, 2).text() or "")
            for row in range(self.fallback_table.rowCount())
            if self.fallback_table.item(row, 0) is not None
            and self.fallback_table.item(row, 0).checkState() == Qt.CheckState.Checked
            and self.fallback_table.item(row, 2) is not None
        }
        return [
            item for item in self._pending_asr
            if str((item.get("video") or {}).get("video_id") or "") in selected_ids
        ]

    def _refresh_table_statuses(self) -> None:
        ready_ids = {
            str(item.get("source_video_id") or "") for item in self._ready_items
        }
        asr_ready_ids = {
            str(item.get("source_video_id") or "")
            for item in self._ready_items
            if str(item.get("source_caption_source") or "") == "asr"
        }
        pending_by_id = {
            str((item.get("video") or {}).get("video_id") or ""): item
            for item in self._pending_asr
        }
        for row, video in enumerate(self._videos):
            pending = pending_by_id.get(video.video_id)
            if pending is not None:
                inspection = pending.get("inspection") or {}
                asr_status = str(inspection.get("asr_status") or "")
                status = "转写失败，待重试" if asr_status == "failed" else "等待选择下载转写"
            elif video.video_id in ready_ids:
                status = "ASR 转写完成" if video.video_id in asr_ready_ids else "字幕已准备"
            else:
                continue
            self.video_table.setItem(row, 4, QTableWidgetItem(status))

    def _on_prepare_progress(
        self,
        index: int,
        total: int,
        video: YouTubeVideo,
        inspection: dict[str, object],
    ) -> None:
        self._completed_video_ids.add(video.video_id)
        self._upsert_preparation_progress(video, inspection)
        if inspection.get("status") == "asr_required":
            status = "转写未完成，待重试" if self.auto_asr_fallback_check.isChecked() else "等待选择下载转写"
        elif inspection.get("source_kind") == "asr":
            status = "ASR 转写完成"
        else:
            status = "字幕已准备"
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 4, QTableWidgetItem(status))
                break
        phase = "下载并 ASR 转写" if inspection.get("source_kind") == "asr" else "获取直出字幕"
        self.status_label.setText(f"正在{phase}：{index}/{total}，当前 {self._short_title(video.title)}")

    def _upsert_preparation_progress(
        self,
        video: YouTubeVideo,
        inspection: dict[str, object],
    ) -> None:
        """Keep each completed subtitle inspection durable during a batch."""
        video_id = video.video_id
        self._ready_items = [
            item for item in self._ready_items
            if str(item.get("source_video_id") or "") != video_id
        ]
        self._pending_asr = [
            item for item in self._pending_asr
            if str((item.get("video") or {}).get("video_id") or "") != video_id
        ]
        if inspection.get("status") == "asr_required":
            self._pending_asr.append(
                {"video": video.to_dict(), "inspection": dict(inspection)}
            )
        else:
            workflow = VerificationWorkflow()
            self._ready_items.append(workflow._build_video_match_item(video, inspection))  # noqa: SLF001
        self._ready_items = VerificationWorkflow().coalesce_video_match_items(self._ready_items)
        self._matching_page.set_items(self._ready_items)
        self._refresh_fallback_table()
        self._persist_workspace_state()

    def _on_asr_stage(self, message: str) -> None:
        """Keep long ASR batches observable between per-video completion events."""
        self.status_label.setText(message)
        if hasattr(self, "_dashboard_page"):
            self._dashboard_page.refresh(self._workspace_snapshot())
        if hasattr(self, "_task_center_page"):
            self._task_center_page.refresh(self._workspace_snapshot())
        # Download stages happen before per-video ASR results exist. Persist
        # them so a restart does not appear stuck at 0/N.
        if not self._state_persist_timer.isActive():
            self._state_persist_timer.start(300)

    def _offer_asr_fallback(self, pending_asr: list[dict[str, object]]) -> None:
        if not pending_asr:
            return
        self.task_tabs.setCurrentIndex(1)
        answer = QMessageBox.question(
            self,
            "部分视频没有直出字幕",
            f"有 {len(pending_asr)} 条视频未能直接取得字幕。\n是否现在下载音频并执行 ASR 转写？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer == QMessageBox.StandardButton.Yes:
            self._prepare_asr_fallback()

    def _prepare_asr_fallback(self) -> None:
        if not YouTubeAsrService().is_ready():
            message = "ASR 尚未配置或没有可用的 API 密钥，未开始下载。请先打开“ASR 配置”并测试可用性。"
            self.status_label.setText(message)
            if hasattr(self, "_subtitle_page"):
                self._subtitle_page.set_status(message)
            QMessageBox.warning(self, "ASR 未配置", message)
            return
        pending_items = self._selected_pending_asr_items()
        if not pending_items:
            existing_pending = {
                str((item.get("video") or {}).get("video_id") or ""): item
                for item in self._pending_asr
            }
            pending_items = [
                existing_pending.get(video.video_id)
                or {
                    "video": video.to_dict(),
                    "inspection": VerificationWorkflow.asr_required_inspection(
                        video,
                        int(self.seconds_combo.currentData()),
                    ),
                }
                for video in self._selected_direct_asr_videos()
            ]
        if not pending_items:
            QMessageBox.information(self, "无需转写", "当前勾选的视频没有需要下载转写的条目。")
            return
        self._active_video_ids = {
            str((item.get("video") or {}).get("video_id") or "")
            for item in pending_items
        }
        self._set_busy(True)
        self.status_label.setText(f"正在下载音频并转写：0/{len(pending_items)}")
        if self.audio_strategy_combo.currentData() == "browser":
            self.status_label.setText(f"正在浏览器高速采集音频：0/{len(pending_items)}")
            self._browser_capture_dialog = BrowserAudioCaptureDialog(
                pending_items,
                int(self.seconds_combo.currentData()),
                int(self.browser_concurrency_combo.currentData()),
                self,
            )
            self._browser_capture_dialog.completed.connect(
                lambda audio_sources, _failed: self._on_browser_capture_finished(pending_items, audio_sources)
            )
            self._browser_capture_dialog.show()
            self._browser_capture_dialog.raise_()
            self._browser_capture_dialog.activateWindow()
            return
        self._start_asr_thread(pending_items, {})

    def _selected_direct_asr_videos(self) -> list[YouTubeVideo]:
        """Return selected videos that do not already have usable subtitles."""
        ready_ids = {
            str(item.get("source_video_id") or "")
            for item in self._ready_items
            if str(item.get("source_video_id") or "")
        }
        return [
            video for video in self._selected_videos()
            if video.video_id not in ready_ids
        ]

    def _on_browser_capture_finished(
        self,
        pending_items: list[dict[str, object]],
        audio_sources: dict[str, str],
    ) -> None:
        if self._task_control is not None and self._task_control.cancelled:
            self._on_task_cancelled([], [])
            return
        self._start_asr_thread(pending_items, audio_sources)

    def _start_asr_thread(self, pending_items: list[dict[str, object]], audio_sources: dict[str, str]) -> None:
        self._completed_video_ids = set()
        # Recovery enters here directly, without the normal button handler.
        # Keep result merging correct for both entry points.
        self._active_video_ids = {
            str((item.get("video") or {}).get("video_id") or "")
            for item in pending_items
            if isinstance(item, dict)
            and str((item.get("video") or {}).get("video_id") or "")
        }
        self._task_control = TaskControl()
        self._task_spec = {
            "kind": "asr_fallback",
            "leading_seconds": int(self.seconds_combo.currentData()),
            "download_concurrency": int(self.asr_download_concurrency_combo.currentData()),
            "asr_concurrency": int(self.asr_transcribe_concurrency_combo.currentData()),
        }
        self._set_busy(True)
        self._thread = _AsrFallbackThread(
            pending_items,
            int(self.seconds_combo.currentData()),
            self._task_control,
            audio_sources,
            int(self.asr_download_concurrency_combo.currentData()),
            int(self.asr_transcribe_concurrency_combo.currentData()),
        )
        self._thread.succeeded.connect(self._on_asr_fallback_finished)
        self._thread.progress.connect(self._on_asr_progress)
        self._thread.stage.connect(self._on_asr_stage)
        self._thread.cancelled.connect(self._on_task_cancelled)
        self._thread.failed.connect(self._on_worker_failed)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()
        self._persist_workspace_state()

    def _on_asr_progress(
        self,
        index: int,
        total: int,
        video: YouTubeVideo,
        inspection: dict[str, object],
    ) -> None:
        self._completed_video_ids.add(video.video_id)
        self._upsert_preparation_progress(video, inspection)
        status = "转写已完成" if inspection.get("status") != "asr_required" else "转写失败，待重试"
        for row, item in enumerate(self._videos):
            if item.video_id == video.video_id:
                self.video_table.setItem(row, 4, QTableWidgetItem(status))
                break
        for row in range(self.fallback_table.rowCount()):
            item = self.fallback_table.item(row, 2)
            if item is not None and item.text() == video.video_id:
                self.fallback_table.setItem(
                    row,
                    4,
                    QTableWidgetItem(f"正在下载音频并转写：{index}/{total}"),
                )
                break
        self.status_label.setText(
            f"正在下载音频并转写：{index}/{total}，当前 {self._short_title(video.title)}"
        )

    def _on_task_cancelled(self, ready: list[dict[str, object]], pending_asr: list[dict[str, object]]) -> None:
        self._active_video_ids = set(self._completed_video_ids)
        self._merge_preparation_results(ready, pending_asr)
        self._refresh_table_statuses()
        self._task_control = None
        self.pause_button.setText("暂停")
        self.status_label.setText(f"任务已取消，已保留 {len(self._completed_video_ids)} 条已完成结果。")
        self._set_busy(False)

    def _on_asr_fallback_finished(
        self,
        ready: list[dict[str, object]],
        pending_asr: list[dict[str, object]],
    ) -> None:
        self._merge_preparation_results(ready, pending_asr)
        self._refresh_table_statuses()
        not_configured = bool(pending_asr) and all(
            str((item.get("inspection") or {}).get("asr_status") or "") == "not_configured"
            for item in pending_asr
        )
        if not_configured:
            message = f"ASR 未配置，未开始下载；仍有 {len(pending_asr)} 条待处理。"
        else:
            message = f"兜底转写完成：新增 {len(ready)} 条可匹配视频字幕，仍有 {len(pending_asr)} 条未完成。"
        self.status_label.setText(message)
        if hasattr(self, "_subtitle_page"):
            self._subtitle_page.set_status(message)
        if not_configured:
            QMessageBox.warning(self, "未开始兜底转写", "请先配置并测试 ASR 服务，再重新点击“下载并转写兜底”。")
        self._set_busy(False)

    def _export_subtitles(self) -> None:
        selected_video_ids = {video.video_id for video in self._selected_videos()}
        if not selected_video_ids:
            QMessageBox.warning(self, "没有选择", "请至少勾选一条视频。")
            return
        default_path = PROJECT_ROOT / "output" / (
            f"youtube_subtitles_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        output_path = self._choose_export_path("导出字幕 Excel", default_path)
        if output_path is None:
            return
        try:
            exported_count = export_subtitles_to_xlsx(
                output_path,
                self._ready_items,
                selected_video_ids,
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        if not exported_count:
            QMessageBox.warning(self, "没有可导出的字幕", "所选视频尚未成功获取字幕。")
            return
        self.status_label.setText(f"已导出 {exported_count} 条视频字幕：{output_path}")
        QMessageBox.information(self, "导出完成", f"已导出 {exported_count} 条视频字幕。\n{output_path}")

    def _export_matching_results(self) -> None:
        if not self._matching_result_rows:
            QMessageBox.information(self, "暂无结果", "请先完成至少一条视频的字幕匹配，再导出结果。")
            return
        default_path = PROJECT_ROOT / "output" / f"youtube_matching_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output_path = self._choose_export_path("导出匹配结果 Excel", default_path)
        if output_path is None:
            return
        try:
            exported_count = export_matching_results_to_xlsx(output_path, self._matching_result_rows)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.status_label.setText(f"已导出 {exported_count} 条视频匹配结果：{output_path}")
        self._matching_page.set_status(self.status_label.text())
        QMessageBox.information(self, "导出完成", f"已导出 {exported_count} 条视频匹配结果。\n{output_path}")

    def _export_cover_review_results(self) -> None:
        if not self._cover_review_results:
            QMessageBox.information(self, "暂无结果", "请先完成至少一条封面检测，再导出结果。")
            return
        default_path = PROJECT_ROOT / "output" / f"youtube_cover_review_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output_path = self._choose_export_path("导出封面检测结果 Excel", default_path)
        if output_path is None:
            return
        source_urls = {video.video_id: video.source_url for video in self._videos}
        thumbnail_urls = {video.video_id: video.thumbnail_url for video in self._videos}
        if hasattr(self, "_cover_page"):
            for item in self._cover_page.items:
                raw_video = item.get("video") if isinstance(item.get("video"), dict) else {}
                video_id = str(raw_video.get("video_id") or "")
                if video_id:
                    source_urls.setdefault(video_id, str(raw_video.get("source_url") or ""))
                    thumbnail_urls.setdefault(video_id, str(raw_video.get("thumbnail_url") or ""))
        try:
            exported_count = export_cover_review_results_to_xlsx(
                output_path,
                self._cover_review_results,
                source_urls,
                thumbnail_urls,
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.status_label.setText(f"已导出 {exported_count} 条封面检测结果：{output_path}")
        self._cover_page.set_status(self.status_label.text())
        QMessageBox.information(self, "导出完成", f"已导出 {exported_count} 条封面检测结果。\n{output_path}")

    def _match(self) -> None:
        if not self._ready_items:
            return
        selected_video_ids = {video.video_id for video in self._selected_videos()}
        if not selected_video_ids:
            QMessageBox.warning(self, "没有选择", "请至少勾选一条已准备字幕的视频后再提交匹配。")
            return
        workflow = VerificationWorkflow()
        selected_items = workflow.filter_video_match_items(self._ready_items, selected_video_ids)
        selected_items = workflow.coalesce_video_match_items(selected_items)
        if not selected_items:
            QMessageBox.information(
                self,
                "所选视频暂无字幕",
                "所选视频尚未准备可匹配字幕，请先获取字幕或完成 ASR 转写。",
            )
            return
        self._matching_page.set_items(selected_items)
        credentials = self._matching_service_credentials()
        if credentials is None:
            return
        server, username, password = credentials
        if self._thread_is_running() or self._matching_thread_is_running():
            QMessageBox.information(self, "任务进行中", "请等待当前本地任务或匹配任务完成后再提交匹配。")
            return
        self._match_paused = False
        self._set_busy(True, scope="matching")
        submit_message = f"正在提交已勾选的 {len(selected_items)} 条视频字幕到匹配服务..."
        self.status_label.setText(submit_message)
        self._matching_page.set_status(submit_message)
        self._active_matching_items = [dict(item) for item in selected_items]
        thread = _MatchThread(
            selected_items,
            server,
            username,
            password,
        )
        self._matching_thread = thread
        thread.task_created.connect(self._on_match_task_created)
        thread.progress.connect(self._on_match_progress)
        thread.succeeded.connect(self._on_matched)
        thread.failed.connect(self._on_matching_worker_failed)
        thread.finished.connect(lambda: self._finish_matching_thread(thread))
        thread.start()

    def _on_match_task_created(self, task_id: str) -> None:
        message = f"匹配任务已提交：{task_id[:12]}...，正在读取服务端排队状态；可继续获取字幕或下载素材。"
        self.status_label.setText(message)
        if hasattr(self, "_matching_page"):
            self._matching_page.set_status(message)
        self._set_busy(True, scope="matching")

    def _on_match_progress(self, detail: dict[str, object]) -> None:
        task = detail.get("task") if isinstance(detail.get("task"), dict) else {}
        status = str(task.get("status") or "处理中")
        queue = task.get("queue") if isinstance(task.get("queue"), dict) else {}

        def duration_text(value: object) -> str:
            try:
                seconds = max(0, int(float(value)))
            except (TypeError, ValueError):
                return ""
            if seconds < 60:
                return f"{seconds} 秒"
            return f"{seconds // 60} 分钟" + (f" {seconds % 60} 秒" if seconds % 60 else "")

        def count(*keys: str) -> int | None:
            for key in keys:
                value = task.get(key)
                if isinstance(value, (int, float)):
                    return int(value)
            counts = task.get("counts") if isinstance(task.get("counts"), dict) else {}
            for key in keys:
                aliases = {
                    "accepted_input_count": "accepted",
                    "completed_input_count": "completed",
                    "failed_input_count": "failed",
                }
                value = counts.get(aliases.get(key, key))
                if isinstance(value, (int, float)):
                    return int(value)
            return None

        completed = count("completed_input_count", "completed_count", "done_count")
        failed = count("failed_input_count", "failed_count", "failure_count")
        processed = count("processed_count")
        if processed is None and completed is not None and failed is not None:
            processed = completed + failed
        total = count("accepted_input_count", "total_count", "item_count", "total")
        progress = f"{processed}/{total}" if processed is not None and total is not None else "处理中"
        failed_text = f"，失败 {failed} 条" if failed is not None else ""
        if status == "queued":
            ahead = queue.get("tasks_ahead_count")
            eta = duration_text(queue.get("estimated_wait_seconds"))
            message = f"排队中：前方 {ahead if ahead is not None else '若干'} 个任务"
            if eta:
                message += f"，预计等待约 {eta}"
        elif status == "running":
            eta = duration_text(queue.get("estimated_remaining_seconds"))
            message = f"正在检测：已完成 {progress}{failed_text}"
            if eta:
                message += f"，预计剩余约 {eta}"
        elif status in {"pause_requested", "cancel_requested"}:
            message = "正在处理暂停/取消请求，将在当前字幕检测结束后生效。"
        elif status == "paused":
            message = "任务已暂停，不占用队列；点击继续可重新进入队列。"
        else:
            source_video_count = self._matching_source_video_count(self._active_matching_items or self._ready_items)
            message = f"匹配任务：{status}，视频 {source_video_count} 条，字幕任务进度 {progress}{failed_text}"
        self.status_label.setText(message)
        if hasattr(self, "_matching_page"):
            self._matching_page.set_status(message)

    def _on_matched(self, detail: dict[str, object]) -> None:
        result_rows = build_matching_result_rows(detail, self._active_matching_items or self._ready_items)
        self._matching_result_rows = result_rows
        self.result_edit.setPlainText(
            "\n".join(
                [f"匹配任务完成，共 {len(result_rows)} 条来源视频。"]
                + [
                    f"{row.get('source_title') or row.get('source_video_id')} | {row.get('match_status')} | "
                    f"{row.get('matched_book_names') or '无命中剧名'} | {row.get('matched_book_ids') or '无 Book ID'}"
                    for row in result_rows
                ]
            )
        )
        if hasattr(self, "_matching_page"):
            self._matching_page.set_result_rows(result_rows)
        task = detail.get("task") if isinstance(detail.get("task"), dict) else {}
        status = str(task.get("status") or "completed").lower()
        self._match_paused = False
        self.pause_button.setText("暂停")
        if status in {"cancelled", "canceled"}:
            message = "视频级匹配已取消，已保留已完成视频的结果。"
        elif status == "partial_failed":
            message = "视频级匹配完成，但存在部分失败项。"
        else:
            message = "视频级匹配已完成。"
        self.status_label.setText(message)
        if hasattr(self, "_matching_page"):
            self._matching_page.set_status(message)
        self._set_busy(False, scope="matching")

    def _on_matching_worker_failed(self, message: str) -> None:
        self._match_paused = False
        self.pause_button.setText("暂停")
        self.status_label.setText("匹配任务失败")
        if hasattr(self, "_matching_page"):
            self._matching_page.set_status(f"匹配任务失败：{message}")
        self._set_busy(False, scope="matching")
        QMessageBox.critical(self, "匹配失败", message)

    def _on_worker_failed(self, message: str) -> None:
        self._task_control = None
        self.pause_button.setText("暂停")
        self.status_label.setText("任务失败")
        if hasattr(self, "_download_page"):
            self._download_page.set_status(f"任务失败：{message}")
        if hasattr(self, "_subtitle_page"):
            self._subtitle_page.set_status(f"任务失败：{message}")
        self._set_busy(False, scope="local")
        QMessageBox.critical(self, "处理失败", message)

    def _clear(self) -> None:
        if self._parallel_cover_review_control is not None:
            self._parallel_cover_review_control.cancel()
        self._videos = []
        self._ready_items = []
        self._active_matching_items = []
        self._matching_result_rows = []
        self._pending_asr = []
        self._cover_paths = {}
        self._cover_review_results = []
        self._standalone_download_results = []
        self._refresh_downloaded_subtitle_import()
        self.video_table.setRowCount(0)
        self.fallback_table.setRowCount(0)
        self.task_tabs.setTabText(1, "待兜底 (0)")
        self.cover_review_table.setRowCount(0)
        self.task_tabs.setTabText(2, "封面检测 (0)")
        self._refresh_channel_selector()
        self.result_edit.clear()
        self.status_label.setText("请输入频道链接开始采集")
        self._set_busy(False)

    def _open_asr_config(self) -> None:
        if self._asr_config_dialog is None:
            self._asr_config_dialog = AsrConfigDialog(self)
            self._asr_config_dialog.finished.connect(self._on_asr_config_closed)
        self._asr_config_dialog.show()
        self._asr_config_dialog.raise_()
        self._asr_config_dialog.activateWindow()

    def _on_asr_config_closed(self, _result: int) -> None:
        dialog = self._asr_config_dialog
        self._asr_config_dialog = None
        if dialog is not None:
            dialog.deleteLater()

    def _open_llm_config(self) -> None:
        if self._llm_config_dialog is None:
            self._llm_config_dialog = LlmConfigDialog(self)
            self._llm_config_dialog.finished.connect(self._on_llm_config_closed)
        self._llm_config_dialog.show()
        self._llm_config_dialog.raise_()
        self._llm_config_dialog.activateWindow()

    def _on_llm_config_closed(self, _result: int) -> None:
        dialog = self._llm_config_dialog
        self._llm_config_dialog = None
        if dialog is not None:
            dialog.deleteLater()

    def _open_matching_config(self) -> None:
        if self._matching_config_dialog is None:
            self._matching_config_dialog = MatchingConfigDialog(self)
            self._matching_config_dialog.config_saved.connect(self._apply_matching_service_config)
            self._matching_config_dialog.finished.connect(self._on_matching_config_closed)
        self._matching_config_dialog.show()
        self._matching_config_dialog.raise_()
        self._matching_config_dialog.activateWindow()

    def _on_matching_config_closed(self, _result: int) -> None:
        dialog = self._matching_config_dialog
        self._matching_config_dialog = None
        if dialog is not None:
            dialog.deleteLater()

    def _apply_matching_service_config(self, config: dict[str, object]) -> None:
        """Refresh the matching page after its standalone configuration changes."""
        if hasattr(self, "_matching_page"):
            state = "已加载保存的匹配服务配置。" if self._api_config_service.is_matching_service_ready(config) else "请先打开“匹配配置”填写服务账号。"
            self._matching_page.set_status(state)

    def _matching_service_credentials(self) -> tuple[str, str, str] | None:
        config = self._api_config_service.get_matching_service_config()
        if not self._api_config_service.is_matching_service_ready(config):
            QMessageBox.warning(
                self,
                "缺少匹配配置",
                "请先打开“匹配配置”，填写并保存匹配服务地址、账号和密码。",
            )
            return None
        return (
            str(config.get("base_url") or "").strip(),
            str(config.get("username") or "").strip(),
            str(config.get("password") or ""),
        )

    def _open_youtube_login(self) -> None:
        if self._youtube_login_dialog is not None:
            self._youtube_login_dialog.show()
            self._youtube_login_dialog.raise_()
            self._youtube_login_dialog.activateWindow()
            return
        self._youtube_login_dialog = YouTubeLoginDialog(self.channel_input.text().strip(), self)
        self._youtube_login_dialog.finished.connect(self._on_youtube_login_closed)
        self._youtube_login_dialog.show()

    def _on_youtube_login_closed(self, _result: int) -> None:
        dialog = self._youtube_login_dialog
        self._youtube_login_dialog = None
        if dialog is not None:
            dialog.deleteLater()
        self.status_label.setText("正在同步 YouTube 账号登录状态...")
        QTimer.singleShot(350, self._sync_closed_browser_login)

    def _sync_closed_browser_login(self) -> None:
        try:
            result = sync_browser_cookie_file()
        except YouTubeCookieSyncError as exc:
            self.status_label.setText(f"YouTube 登录状态未同步：{exc}")
            return
        if result.authenticated:
            self.status_label.setText(
                f"YouTube 账号登录状态已同步（{result.cookie_count} 个 Cookie），后续下载会自动使用。"
            )
        else:
            self.status_label.setText(
                f"已同步 {result.cookie_count} 个基础 Cookie，未检测到账号会话。"
            )

    def _offer_detected_system_proxy(self) -> None:
        if self.proxy_input.text().strip():
            return
        self._detect_system_proxy(interactive=False)

    def _detect_system_proxy(self, *, interactive: bool) -> None:
        candidates, pac_detected = ProxyDiscoveryService.discover()
        if not candidates:
            if interactive:
                message = "未发现可直接使用的 HTTP 系统代理。"
                if pac_detected:
                    message += "\n检测到 PAC 自动代理脚本，当前无法安全解析为固定地址，请向网络工具确认本地 HTTP 代理端口。"
                QMessageBox.information(self, "未检测到代理", message)
            return
        candidate = candidates[0]
        message = (
            f"检测到 {candidate.source}：\n{candidate.address}\n\n"
            "是否将其用于 YouTube 采集和 Range 音频下载？"
        )
        answer = QMessageBox.question(
            self,
            "检测到本机代理",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if answer != QMessageBox.StandardButton.Yes:
            if interactive:
                self.status_label.setText("未修改 YouTube 下载代理。")
            return
        self.proxy_input.setText(candidate.address)
        self._save_proxy(show_feedback=interactive)

    def _save_proxy(self, *, show_feedback: bool = False) -> None:
        value = self.proxy_input.text().strip()
        if value and not value.lower().startswith(("http://", "https://")):
            QMessageBox.warning(self, "代理格式错误", "代理地址应类似 http://127.0.0.1:7897。")
            return
        YOUTUBE_PROXY_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        if value:
            YOUTUBE_PROXY_CONFIG_PATH.write_text(value, encoding="utf-8")
            self.status_label.setText(f"YouTube 下载代理已保存：{value}")
            if show_feedback:
                QMessageBox.information(
                    self,
                    "代理已保存",
                    f"YouTube 下载将使用以下代理：\n{value}\n\n后续采集和音频下载会自动生效。",
                )
        else:
            YOUTUBE_PROXY_CONFIG_PATH.unlink(missing_ok=True)
            self.status_label.setText("YouTube 下载代理已清除，将使用直连网络。")
            if show_feedback:
                QMessageBox.information(self, "代理已清除", "YouTube 下载将使用直连网络。")


def run() -> int:
    app = QApplication.instance() or QApplication([])
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    # Recovery must be single-owner. Without a process lock, launching the
    # VBS twice lets both windows consume the same session_state.json and
    # multiply yt-dlp/Node workers until Windows starts rejecting cmd/node.
    app_lock = QLockFile(str(RUNTIME_DIR / "application.lock"))
    app_lock.setStaleLockTime(10_000)
    if not app_lock.tryLock(0):
        QMessageBox.warning(
            None,
            "素材分析助手已在运行",
            "已有一个素材分析助手实例正在运行。\n请关闭已有窗口后再启动，避免重复恢复同一批任务。",
        )
        return 1
    apply_app_theme(app)
    window = MainWindow()
    window.show()
    # Wait until the window is visible before opening a modal dialog. Showing
    # it from MainWindow.__init__ can race the first event-loop turn and leave
    # the recovery prompt hidden behind the startup window on some systems.
    QTimer.singleShot(250, window._offer_workspace_recovery)
    try:
        return app.exec()
    finally:
        app_lock.unlock()
