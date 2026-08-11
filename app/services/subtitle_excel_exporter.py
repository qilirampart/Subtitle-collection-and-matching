from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


_HEADERS = (
    "视频标题",
    "视频链接",
    "视频 ID",
    "频道",
    "上传日期",
    "字幕范围（秒）",
    "字幕语言",
    "字幕来源",
    "完整字幕",
)


def export_subtitles_to_xlsx(
    output_path: Path,
    items: list[dict[str, object]],
    selected_video_ids: set[str],
) -> int:
    """Write one complete, original transcript per selected video."""
    rows: dict[str, dict[str, object]] = {}
    for item in items:
        video_id = str(item.get("source_video_id") or "").strip()
        if not video_id or video_id not in selected_video_ids or video_id in rows:
            continue
        rows[video_id] = item

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "字幕"
    sheet.append(_HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in rows.values():
        sheet.append((
            str(item.get("source_display_title") or ""),
            str(item.get("source_description") or ""),
            str(item.get("source_video_id") or ""),
            str(item.get("source_channel") or ""),
            str(item.get("source_upload_date") or ""),
            f"{item.get('source_caption_start', 0)}-{item.get('source_caption_end', 0)}",
            str(item.get("source_caption_language") or ""),
            str(item.get("source_caption_source") or ""),
            str(item.get("source_caption_text") or item.get("source_text_original") or ""),
        ))

    widths = (36, 55, 16, 20, 14, 16, 12, 24, 100)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(rows)


__all__ = ["export_subtitles_to_xlsx"]
