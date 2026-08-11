from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.models import YouTubeVideo
from app.services.text_normalizer import SubtitleTextNormalizer
from app.services.youtube_collector import YouTubeCollector
from app.services.task_state import TaskStateStore
from app.workflow import VerificationWorkflow


CHANNEL_URL = "https://www.youtube.com/@%E9%9B%B2%E4%B8%8A%E5%8A%87%E5%A0%B4_99"
COMPLETED_XLSX = Path("output/youtube_subtitles_completed_20260807_153321.xlsx")
RECOVERY_DIR = Path("runtime/recovery")


def range_end_seconds(value: object) -> int:
    text = str(value or "").strip()
    try:
        return max(1, int(float(text)))
    except ValueError:
        pass
    try:
        return max(1, int(float(text.rsplit("-", 1)[-1])))
    except ValueError:
        return 180


def completed_items() -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(COMPLETED_XLSX, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(min_row=2, values_only=True)
    normalizer = SubtitleTextNormalizer()
    workflow = VerificationWorkflow(normalizer=normalizer)
    completed: dict[str, dict[str, object]] = {}
    ready: list[dict[str, object]] = []
    for row in rows:
        video_id = str(row[2] or "").strip()
        source_url = str(row[1] or "").strip()
        text = str(row[8] or "").strip()
        if not video_id or not source_url or not text:
            continue
        video = YouTubeVideo(
            video_id=video_id,
            source_url=source_url,
            title=str(row[0] or video_id).strip(),
            channel=str(row[3] or "").strip(),
            upload_date=str(row[4] or "").strip(),
        )
        inspection = {
            "status": "ready_for_matching",
            "asr_required": False,
            "asr_status": "completed",
            "language_code": str(row[6] or "").strip(),
            "source_kind": str(row[7] or "recovered").strip(),
            "source_path": "",
            "start_seconds": 0,
            "end_seconds": range_end_seconds(row[5]),
            "text": text,
            "normalized_text": normalizer.normalize(text, language_code=str(row[6] or "").strip()),
            "matching_language_code": normalizer.matching_language_code(str(row[6] or "").strip()),
            "cues": [],
        }
        completed[video_id] = video.to_dict()
        ready.append(workflow._build_video_match_item(video, inspection))  # noqa: SLF001
    return completed, ready


def main() -> int:
    parser = argparse.ArgumentParser(description="Rebuild the interrupted 141-video ASR resume state.")
    parser.add_argument("--write", action="store_true", help="Write the recovery manifest and session_state.json.")
    args = parser.parse_args()

    if not COMPLETED_XLSX.is_file():
        raise FileNotFoundError(f"Completed subtitle export is missing: {COMPLETED_XLSX}")
    completed, ready_items = completed_items()
    videos = YouTubeCollector().collect_channel(CHANNEL_URL, max_items=0)
    current_ids = {video.video_id for video in videos}
    completed_ids = set(completed)
    if len(videos) != 141 or len(completed_ids) != 66 or not completed_ids.issubset(current_ids):
        raise RuntimeError(
            f"Unsafe recovery set: current={len(videos)} completed={len(completed_ids)} overlap={len(current_ids & completed_ids)}"
        )

    pending_asr = []
    for video in videos:
        if video.video_id in completed_ids:
            continue
        pending_asr.append(
            {
                "video": video.to_dict(),
                "inspection": {
                    "status": "asr_required",
                    "asr_required": True,
                    "asr_status": "pending_resume",
                    "source_kind": "none",
                    "source_path": "",
                    "text": "",
                    "normalized_text": "",
                    "matching_language_code": "",
                    "language_code": "",
                    "start_seconds": 0,
                    "end_seconds": 180,
                    "cues": [],
                },
            }
        )

    payload = {
        "version": 1,
        "active": True,
        "task_kind": "ASR 续跑",
        "status": "已恢复原 141 条队列：66 条已完成，75 条等待 ASR 兜底",
        "videos": [video.to_dict() for video in videos],
        "ready_items": ready_items,
        "matching_result_rows": [],
        "pending_asr": pending_asr,
        "cover_paths": {},
        "cover_review_results": [],
        "download_results": [],
        "download_jobs": [],
        "subtitle_jobs": [],
    }
    summary = {
        "channel": CHANNEL_URL,
        "total_videos": len(videos),
        "completed_videos": len(completed_ids),
        "pending_asr_videos": len(pending_asr),
        "completed_ids": sorted(completed_ids),
        "pending_ids": [str(item["video"]["video_id"]) for item in pending_asr],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not args.write:
        return 0

    RECOVERY_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest = RECOVERY_DIR / f"asr_resume_manifest_{stamp}.json"
    manifest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    store = TaskStateStore()
    from app.settings import TASK_STATE_PATH

    if TASK_STATE_PATH.is_file():
        shutil.copy2(TASK_STATE_PATH, RECOVERY_DIR / f"session_state_before_resume_{stamp}.json")
    store.save(payload)
    print(f"manifest={manifest.resolve()}")
    print(f"session_state={TASK_STATE_PATH.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
