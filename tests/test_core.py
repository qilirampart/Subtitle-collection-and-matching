from __future__ import annotations

import unittest
from pathlib import Path
from threading import Event, Thread
from tempfile import TemporaryDirectory
from time import sleep
from types import SimpleNamespace
from unittest.mock import MagicMock, Mock, patch

from openpyxl import load_workbook

from app.audio_transcription_models import AudioTranscriptionResult, PreparedAudio
from app.models import CaptionCue, YouTubeVideo
from app.services.matching_api import (
    DramaSubtitleMatchingClient,
    MatchingServiceConfig,
    aggregate_video_results,
    build_matching_result_rows,
)
from app.services.subtitle_service import YouTubeSubtitleService
from app.services.text_normalizer import SubtitleTextNormalizer
from app.services.youtube_collector import YouTubeCollectionError, YouTubeCollector
from app.services.youtube_cover_service import CoverDownloadResult, YouTubeCoverService
from app.services.youtube_cover_review_service import CoverReviewResult, YouTubeCoverReviewService
from app.services.youtube_cookie_service import _chrome_expiry_to_unix
from app.services.youtube_service import YouTubeDownloadCancelled
from app.services.youtube_transcript_service import YouTubeTranscriptPanelService, YouTubeTranscriptUnavailable
from app.services.youtube_asr import YouTubeAsrService
from app.services.subtitle_excel_exporter import export_subtitles_to_xlsx
from app.services.review_excel_exporter import export_cover_review_results_to_xlsx, export_matching_results_to_xlsx
from app.services.proxy_discovery_service import ProxyDiscoveryService
from app.services.task_state import TaskStateStore
from app.services.youtube_audio_service import YouTubeAudioService
from app.task_control import TaskControl
from app.ui.workspace_pages import MatchingPage
from app.workflow import VerificationWorkflow


class YouTubeCollectorTests(unittest.TestCase):
    def test_normalizes_channel_and_subtab_urls(self) -> None:
        self.assertEqual(
            YouTubeCollector.normalize_channel_url("https://www.youtube.com/@example/shorts"),
            "https://www.youtube.com/@example/videos",
        )
        self.assertEqual(
            YouTubeCollector.normalize_channel_url("https://www.youtube.com/@example"),
            "https://www.youtube.com/@example/videos",
        )

    def test_parses_srt_cues(self) -> None:
        cues = YouTubeCollector.srt_cues(
            "1\n00:00:00,000 --> 00:00:01,500\nHello\n\n2\n00:00:02,000 --> 00:00:03,000\nWorld"
        )
        self.assertEqual(cues, [(0.0, 1.5, "Hello"), (2.0, 3.0, "World")])

    def test_preserves_thumbnail_url_from_channel_entry(self) -> None:
        collector = YouTubeCollector()
        collector._extract = Mock(return_value={  # type: ignore[method-assign]
            "channel": "Demo channel",
            "entries": [{
                "id": "abc123",
                "title": "Demo video",
                "thumbnail": "https://img.example/cover.webp",
            }],
        })

        videos = collector.collect_channel("https://www.youtube.com/@example")

        self.assertEqual(len(videos), 1)
        self.assertEqual(videos[0].thumbnail_url, "https://img.example/cover.webp")


class MatchingSelectionTests(unittest.TestCase):
    def test_filters_matching_items_to_checked_videos_before_coalescing(self) -> None:
        workflow = VerificationWorkflow()
        items = [
            {"source_video_id": "channel-a-1", "source_segment_order": 1},
            {"source_video_id": "channel-a-1", "source_segment_order": 2},
            {"source_video_id": "channel-b-1", "source_segment_order": 1},
        ]

        selected = workflow.filter_video_match_items(items, {"channel-a-1"})

        self.assertEqual(len(selected), 2)
        self.assertTrue(all(item["source_video_id"] == "channel-a-1" for item in selected))


class YouTubeCoverServiceTests(unittest.TestCase):
    def test_downloads_thumbnail_to_simple_video_id_filename(self) -> None:
        response = Mock()
        response.headers = {"content-type": "image/jpeg"}
        response.content = b"\xff\xd8" + (b"x" * 2048)
        response.raise_for_status.return_value = None
        session = Mock()
        session.get.return_value = response
        video = YouTubeVideo(
            "abc123",
            "https://www.youtube.com/watch?v=abc123",
            "Demo video",
            thumbnail_url="https://img.example/cover.jpg",
        )
        with TemporaryDirectory() as directory, patch(
            "app.services.youtube_cover_service.requests.Session", return_value=session
        ), patch.object(YouTubeCoverService, "_proxy_settings", return_value=None):
            result = YouTubeCoverService().download_cover(video, output_dir=Path(directory))

        self.assertFalse(result.error)
        self.assertTrue(result.path.endswith("abc123.jpg"))
        self.assertEqual(session.get.call_count, 1)

    def test_download_batch_reports_start_and_completion_for_each_cover(self) -> None:
        videos = [
            YouTubeVideo("cover-a", "https://www.youtube.com/watch?v=cover-a", "A"),
            YouTubeVideo("cover-b", "https://www.youtube.com/watch?v=cover-b", "B"),
            YouTubeVideo("cover-c", "https://www.youtube.com/watch?v=cover-c", "C"),
        ]
        service = YouTubeCoverService()
        service.download_cover = Mock(side_effect=lambda video: CoverDownloadResult(video.video_id, f"{video.video_id}.jpg"))  # type: ignore[method-assign]
        started: list[tuple[int, str]] = []
        completed: list[tuple[int, str]] = []

        results, cancelled = service.download_batch(
            videos,
            concurrency=2,
            started_callback=lambda index, _total, video: started.append((index, video.video_id)),
            progress_callback=lambda index, _total, video, _result: completed.append((index, video.video_id)),
        )

        self.assertFalse(cancelled)
        self.assertEqual([result.video_id for result in results], ["cover-a", "cover-b", "cover-c"])
        self.assertEqual(sorted(started), [(1, "cover-a"), (2, "cover-b"), (3, "cover-c")])
        self.assertEqual(sorted(index for index, _video_id in completed), [1, 2, 3])


class YouTubeCoverReviewServiceTests(unittest.TestCase):
    def test_parses_structured_multimodal_result(self) -> None:
        response = Mock()
        response.raise_for_status.return_value = None
        response.json.return_value = {
            "choices": [{
                "message": {
                    "content": '{"overall_risk":"review","risk_tags":["student"],"summary":"需要复核","evidence":"可见校园场景","confidence":0.82}'
                }
            }]
        }
        profile = {
            "id": "llm-1",
            "enabled": True,
            "api_base": "https://llm.example",
            "api_key": "test-key",
            "model": "vision-model",
            "temperature": 0,
        }
        config = Mock()
        config.get_llm_profiles.return_value = [profile]
        config.is_llm_profile_ready.return_value = True
        video = YouTubeVideo("abc123", "https://www.youtube.com/watch?v=abc123", "Demo")
        with TemporaryDirectory() as directory, patch(
            "app.services.youtube_cover_review_service.requests.post", return_value=response
        ) as post:
            cover_path = Path(directory) / "abc123.jpg"
            cover_path.write_bytes(b"image-data")
            result = YouTubeCoverReviewService(config).review_cover(video, cover_path)

        self.assertEqual(result.overall_risk, "review")
        self.assertEqual(result.risk_tags, ("学生或校园",))
        self.assertEqual(result.confidence, 0.82)
        self.assertIn('"overall_risk":"review"', result.model_response)
        self.assertTrue(post.call_args.kwargs["json"]["messages"][1]["content"][1]["image_url"]["url"].startswith("data:image/jpeg;base64,"))


class SubtitleExcelExportTests(unittest.TestCase):
    def test_exports_one_complete_transcript_per_video(self) -> None:
        items = [
            {
                "source_video_id": "video-1",
                "source_display_title": "Video one",
                "source_description": "https://youtube.example/watch?v=video-1",
                "source_caption_language": "en",
                "source_caption_source": "youtube_transcript_panel",
                "source_caption_start": 0,
                "source_caption_end": 180,
                "source_caption_text": "The complete transcript.",
                "source_text_original": "Only a matching segment.",
            },
            {
                "source_video_id": "video-1",
                "source_display_title": "Video one",
                "source_text_original": "Another matching segment.",
            },
        ]
        with TemporaryDirectory() as directory:
            output_path = Path(directory) / "subtitles.xlsx"
            self.assertEqual(export_subtitles_to_xlsx(output_path, items, {"video-1"}), 1)
            sheet = load_workbook(output_path).active
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(2, 1).value, "Video one")
            self.assertEqual(sheet.cell(2, 9).value, "The complete transcript.")


class ReviewExcelExportTests(unittest.TestCase):
    def test_exports_complete_matching_review_rows(self) -> None:
        rows = [{
            "source_channel": "Demo channel",
            "source_title": "Demo video",
            "source_url": "https://youtube.example/watch?v=video-1",
            "source_video_id": "video-1",
            "source_subtitle": "Complete subtitle.",
            "match_status": "confirmed_match",
            "user_message": "已确认命中",
            "matched_book_names": "Demo drama",
            "matched_book_ids": "book-1",
            "execution": {"strategy": "fast_screen_only"},
            "evidence_pairs": [{"matched_text": "Library subtitle."}],
        }]
        with TemporaryDirectory() as directory:
            output_path = Path(directory) / "matching.xlsx"
            self.assertEqual(export_matching_results_to_xlsx(output_path, rows), 1)
            sheet = load_workbook(output_path).active
            self.assertEqual(sheet.cell(2, 3).value, "https://youtube.example/watch?v=video-1")
            self.assertEqual(sheet.cell(2, 5).value, "Complete subtitle.")
            self.assertEqual(sheet.cell(2, 9).value, "book-1")
            self.assertIn("fast_screen_only", sheet.cell(2, 18).value)

    def test_exports_cover_review_with_original_model_response(self) -> None:
        review = CoverReviewResult(
            video_id="video-2",
            title="Cover demo",
            cover_path="covers/video-2.jpg",
            overall_risk="review",
            risk_tags=("学生或校园",),
            summary="需要人工复核",
            evidence="可见校服",
            confidence=0.82,
            model_response='{"overall_risk":"review"}',
        )
        with TemporaryDirectory() as directory:
            output_path = Path(directory) / "covers.xlsx"
            self.assertEqual(
                export_cover_review_results_to_xlsx(
                    output_path,
                    [review],
                    {"video-2": "https://youtube.example/watch?v=video-2"},
                ),
                1,
            )
            sheet = load_workbook(output_path).active
            self.assertEqual(sheet.cell(2, 2).value, "https://youtube.example/watch?v=video-2")
            self.assertEqual(sheet.cell(2, 7).value, "需要人工复核")
            self.assertIn('"overall_risk":"review"', sheet.cell(2, 10).value)


class YouTubeCookieExportTests(unittest.TestCase):
    def test_converts_chromium_expiry_to_unix_seconds(self) -> None:
        self.assertEqual(_chrome_expiry_to_unix(0), 0)
        self.assertEqual(_chrome_expiry_to_unix(11_644_473_601_000_000), 1)


class YouTubeTranscriptPanelTests(unittest.TestCase):
    def test_parses_youtube_transcript_timestamps(self) -> None:
        self.assertEqual(YouTubeTranscriptPanelService._timestamp_to_seconds("1:02"), 62)
        self.assertEqual(YouTubeTranscriptPanelService._timestamp_to_seconds("1:02:03"), 3723)
        self.assertIsNone(YouTubeTranscriptPanelService._timestamp_to_seconds("invalid"))

    def test_identifies_common_transcript_scripts_without_translation(self) -> None:
        self.assertEqual(YouTubeTranscriptPanelService._infer_language_code("你好，世界"), "zh")
        self.assertEqual(YouTubeTranscriptPanelService._infer_language_code("こんにちは"), "ja")
        self.assertEqual(YouTubeTranscriptPanelService._infer_language_code("안녕하세요"), "ko")
        self.assertEqual(YouTubeTranscriptPanelService._infer_language_code("Hello world"), "en")


class SubtitleWorkflowTests(unittest.TestCase):
    def test_builds_one_video_match_record_with_complete_cues(self) -> None:
        video = YouTubeVideo("abc", "https://youtu.be/abc", "Video A", channel="Channel A", upload_date="2026-08-01")
        subtitle_service = Mock()
        cue_texts = [f"cue {index} " + ("x" * 90) for index in range(8)]
        caption = SimpleNamespace(
            to_dict=lambda: {
                "video": video.to_dict(),
                "language_code": "en",
                "source_kind": "manual",
                "source_path": "captions.srt",
                "text": "\n".join(cue_texts),
                "start_seconds": 0,
                "end_seconds": 160,
                "asr_required": False,
                "cues": tuple(
                    {"start_seconds": index * 20.0, "end_seconds": index * 20.0 + 18.0, "text": text}
                    for index, text in enumerate(cue_texts)
                ),
            },
            text="\n".join(cue_texts), language_code="en", asr_required=False,
        )
        subtitle_service.acquire_leading_captions.return_value = caption
        workflow = VerificationWorkflow(subtitle_service=subtitle_service)

        ready, pending = workflow.prepare_batch_items([video])

        self.assertFalse(pending)
        self.assertEqual(len(ready), 1)
        self.assertEqual(ready[0]["source_video_id"], "abc")
        self.assertEqual(ready[0]["source_caption_language"], "en")
        self.assertEqual(ready[0]["source_segment_order"], 1)
        self.assertTrue(ready[0]["source_text_original"])
        self.assertEqual(ready[0]["source_time_start"], "0.000")
        self.assertGreater(len(ready[0]["query_text"]), 500)
        self.assertEqual(len(ready[0]["cues"]), len(cue_texts))
        self.assertEqual(ready[0]["matching_mode"], "video")

    def test_prefers_manual_chinese_track_then_automatic(self) -> None:
        self.assertEqual(
            YouTubeSubtitleService._choose_track({"en": [{}], "zh-Hant": [{}]}, {"zh": [{}]}),
            ("zh-Hant", "manual"),
        )
        self.assertEqual(YouTubeSubtitleService._choose_track({}, {"en": [{}]}), ("en", "automatic"))

    def test_marks_video_for_asr_when_no_caption_is_available(self) -> None:
        video = YouTubeVideo("test", "https://www.youtube.com/watch?v=test", "test")
        subtitle_service = Mock()
        subtitle_service.acquire_leading_captions.return_value = type(
            "Caption",
            (),
            {
                "to_dict": lambda self: {"text": "", "language_code": "", "asr_required": True},
                "text": "",
                "language_code": "",
                "asr_required": True,
            },
        )()
        workflow = VerificationWorkflow(subtitle_service=subtitle_service)
        inspected = workflow.inspect_video(video)
        self.assertEqual(inspected["status"], "asr_required")

    def test_marks_video_for_asr_when_caption_track_lookup_fails(self) -> None:
        video = YouTubeVideo("test", "https://www.youtube.com/watch?v=test", "test")
        collector = Mock()
        collector._extract.side_effect = YouTubeCollectionError("format unavailable")
        transcript_panel = Mock()
        transcript_panel.acquire_leading_transcript.side_effect = YouTubeTranscriptUnavailable("no panel")
        service = YouTubeSubtitleService(
            collector=collector,
            transcript_panel_service=transcript_panel,
        )

        caption = service.acquire_leading_captions(video)

        self.assertTrue(caption.asr_required)
        self.assertEqual(caption.source_kind, "none")

    def test_retries_transcript_panel_before_falling_back_to_asr(self) -> None:
        video = YouTubeVideo("test", "https://www.youtube.com/watch?v=test", "test")
        panel = Mock()
        panel.acquire_leading_transcript.side_effect = [
            YouTubeTranscriptUnavailable("temporary response"),
            SimpleNamespace(
                text="Recovered transcript.",
                language_code="en",
                start_seconds=0,
                end_seconds=180,
                cues=(),
            ),
        ]
        service = YouTubeSubtitleService(collector=Mock(), transcript_panel_service=panel)
        with TemporaryDirectory() as directory, patch("app.services.subtitle_service.CAPTION_DIR", Path(directory)), patch(
            "app.services.subtitle_service.time.sleep"
        ):
            caption = service.acquire_leading_captions(video)

        self.assertEqual(caption.source_kind, "youtube_transcript_panel")
        self.assertEqual(panel.acquire_leading_transcript.call_count, 2)

    def test_batch_defers_asr_until_user_explicitly_requests_fallback(self) -> None:
        video = YouTubeVideo("test", "https://www.youtube.com/watch?v=test", "test")
        subtitle_service = Mock()
        subtitle_service.acquire_leading_captions.return_value = type(
            "Caption",
            (),
            {
                "to_dict": lambda self: {"text": "", "language_code": "", "asr_required": True},
                "text": "",
                "language_code": "",
                "asr_required": True,
            },
        )()
        asr_service = Mock()
        workflow = VerificationWorkflow(subtitle_service=subtitle_service, asr_service=asr_service)

        ready, pending = workflow.prepare_batch_items([video])

        self.assertFalse(ready)
        self.assertEqual(len(pending), 1)
        asr_service.is_ready.assert_not_called()

    def test_batch_runs_asr_immediately_when_automatic_fallback_is_enabled(self) -> None:
        video = YouTubeVideo("test", "https://www.youtube.com/watch?v=test", "test")
        subtitle_service = Mock()
        subtitle_service.acquire_leading_captions.return_value = type(
            "Caption",
            (),
            {
                "to_dict": lambda self: {"text": "", "language_code": "", "asr_required": True},
                "text": "",
                "language_code": "",
                "asr_required": True,
            },
        )()
        asr_service = Mock()
        asr_service.is_ready.return_value = True
        asr_service.transcribe_video.return_value = SimpleNamespace(
            source_path="audio.m4a",
            audio_path="audio.m4a",
            text="ASR transcript text.",
        )
        workflow = VerificationWorkflow(subtitle_service=subtitle_service, asr_service=asr_service)

        ready, pending = workflow.prepare_batch_items([video], allow_asr_fallback=True)

        self.assertFalse(pending)
        self.assertEqual(len(ready), 1)
        self.assertEqual(ready[0]["source_caption_source"], "asr")
        asr_service.transcribe_video.assert_called_once()

    def test_maps_youtube_script_labels_without_changing_text(self) -> None:
        normalizer = SubtitleTextNormalizer()
        self.assertEqual(normalizer.matching_language_code("zh-Hant"), "zh")
        self.assertEqual(normalizer.matching_language_code("zh-Hans"), "zh")
        self.assertEqual(normalizer.normalize("繁體 台詞", language_code="zh-Hant"), "繁體 台詞")

    def test_coalesces_legacy_segment_records_into_one_video_request(self) -> None:
        workflow = VerificationWorkflow()
        legacy_items = [
            {
                "source_video_id": "legacy-video",
                "source_ref": "https://www.youtube.com/watch?v=legacy-video#segment-1",
                "source_caption_language": "en",
                "source_caption_text": "Complete subtitle text for a legacy video.",
                "source_text_original": "Complete subtitle text for",
                "source_segment_order": 1,
                "source_time_start": "0",
                "source_time_end": "12",
            },
            {
                "source_video_id": "legacy-video",
                "source_ref": "https://www.youtube.com/watch?v=legacy-video#segment-2",
                "source_caption_language": "en",
                "source_caption_text": "Complete subtitle text for a legacy video.",
                "source_text_original": "a legacy video.",
                "source_segment_order": 2,
                "source_time_start": "10",
                "source_time_end": "20",
            },
        ]

        items = workflow.coalesce_video_match_items(legacy_items)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["matching_mode"], "video")
        self.assertEqual(items[0]["source_ref"], "https://www.youtube.com/watch?v=legacy-video")
        self.assertEqual(items[0]["query_text"], "Complete subtitle text for a legacy video.")
        self.assertEqual(len(items[0]["cues"]), 2)


class MatchingClientTests(unittest.TestCase):
    def test_derives_visible_common_phrase_when_service_has_no_highlight_terms(self) -> None:
        terms = MatchingPage._common_phrase_terms(
            "女主终于拿到了离婚协议书，准备离开这座城市。",
            "她拿到了离婚协议书，随后准备离开这座城市重新开始。",
            [],
        )

        self.assertTrue(any("离婚协议书" in term for term in terms))
        self.assertTrue(any("准备离开这座城市" in term for term in terms))

    def test_aggregates_segment_results_by_video(self) -> None:
        summary = aggregate_video_results({"items": [
            {"source_video_id": "v1", "source_display_title": "Video", "source_time_start": "1", "source_time_end": "2", "matched_book_id": "b1", "matched_book_name": "Book", "lexical_score": -2.0},
            {"source_video_id": "v1", "source_display_title": "Video", "source_time_start": "3", "source_time_end": "4", "matched_book_id": "b1", "matched_book_name": "Book", "lexical_score": -1.0},
            {"source_video_id": "v1", "source_display_title": "Video", "matched_book_id": "b2", "matched_book_name": "Other", "lexical_score": -0.5},
        ]})
        self.assertEqual(len(summary), 1)
        self.assertEqual(summary[0]["segment_count"], 3)
        self.assertEqual(summary[0]["matched_candidates"][0]["book_id"], "b1")
        self.assertEqual(summary[0]["matched_candidates"][0]["segment_hits"], 2)

    def test_aggregates_final_decision_and_keeps_review_segments(self) -> None:
        detail = {
            "items": [
                {
                    "source_video_id": "v2",
                    "source_display_title": "Video 2",
                    "source_segment_order": 1,
                    "source_time_start": "0",
                    "source_time_end": "20",
                    "result_payload_json": '{"decision":{"matched":false,"status":"review_required","reason":"semantic_candidate_requires_manual_review","book_id":"b9","book_name":"Candidate","matched_episode_order":3}}',
                },
                {
                    "source_video_id": "v2",
                    "source_display_title": "Video 2",
                    "source_segment_order": 2,
                    "matched_book_id": "b1",
                    "matched_book_name": "Confirmed",
                    "matched_episode_order": 1,
                    "result_payload_json": '{"decision":{"matched":true,"status":"matched","reason":"strong_lexical_evidence","book_id":"b1","book_name":"Confirmed","matched_episode_order":1}}',
                },
            ]
        }
        summary = aggregate_video_results(detail)
        self.assertEqual(summary[0]["matched_segment_count"], 1)
        self.assertEqual(summary[0]["review_segment_count"], 1)
        self.assertEqual(summary[0]["not_matched_segment_count"], 0)
        self.assertEqual(summary[0]["review_items"][0]["reason"], "semantic_candidate_requires_manual_review")
        self.assertEqual(summary[0]["matched_candidates"][0]["matched_episode_orders"], [1])

    def test_builds_final_result_rows_with_source_subtitle_and_book_id(self) -> None:
        detail = {
            "items": [
                {
                    "source_video_id": "video-1",
                    "source_display_title": "Source drama",
                    "source_channel": "Source channel",
                    "source_time_start": "0",
                    "source_time_end": "20",
                    "result_payload_json": '{"decision":{"matched":true,"status":"matched","reason":"strong_lexical_evidence","book_id":"book-9","book_name":"Matched drama","matched_episode_order":2}}',
                }
            ]
        }
        rows = build_matching_result_rows(
            detail,
            [{
                "source_video_id": "video-1",
                "source_ref": "https://www.youtube.com/watch?v=video-1",
                "source_channel": "Source channel",
                "source_display_title": "Source drama",
                "source_caption_text": "Complete source subtitle.",
            }],
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["source_channel"], "Source channel")
        self.assertEqual(rows[0]["source_title"], "Source drama")
        self.assertEqual(rows[0]["source_subtitle"], "Complete source subtitle.")
        self.assertEqual(rows[0]["source_url"], "https://www.youtube.com/watch?v=video-1")
        self.assertEqual(rows[0]["matched_book_names"], "Matched drama")
        self.assertEqual(rows[0]["matched_book_ids"], "book-9")

    def test_keeps_candidate_evidence_for_side_by_side_review(self) -> None:
        detail = {
            "items": [
                {
                    "source_video_id": "video-2",
                    "source_segment_order": 1,
                    "result_payload_json": {
                        "decision": {
                            "matched": True,
                            "status": "matched",
                            "candidate_rank": 1,
                            "book_id": "book-2",
                            "book_name": "Matched drama",
                            "matched_episode_order": 3,
                            "reason": "strong_lexical_evidence",
                        },
                        "candidates": [{
                            "rank": 1,
                            "book_id": "book-2",
                            "book_name": "Matched drama",
                            "episode_order": 3,
                            "match_metrics": {"shared_trigrams": ["Source subtitle"]},
                            "evidence": {
                                "window_uid": "book-2:3:1-4",
                                "window_text": "Matched library subtitle.",
                            },
                        }],
                    },
                }
            ]
        }
        rows = build_matching_result_rows(
            detail,
            [{
                "source_video_id": "video-2",
                "source_segment_order": 1,
                "source_text_original": "Source subtitle segment.",
            }],
        )

        pair = rows[0]["evidence_pairs"][0]
        self.assertEqual(pair["source_text"], "Source subtitle segment.")
        self.assertEqual(pair["matched_text"], "Matched library subtitle.")
        self.assertEqual(pair["window_uid"], "book-2:3:1-4")
        self.assertEqual(pair["highlight_terms"], ["Source subtitle"])

    def test_keeps_all_strong_candidates_for_ambiguous_content_match(self) -> None:
        detail = {
            "items": [{
                "source_video_id": "video-ambiguous",
                "source_segment_order": 1,
                "result_payload_json": {
                    "decision": {
                        "matched": False,
                        "status": "review_required",
                        "outcome": "content_matched_ambiguous",
                        "content_match_status": "matched",
                        "title_resolution": "ambiguous",
                        "user_message": "内容已命中：多个剧集存在连续台词复用。",
                        "strong_match_candidates": [
                            {
                                "review_priority": 1,
                                "book_id": "book-a",
                                "book_name": "Drama A",
                                "matched_episode_order": 1,
                                "aggregate_text_coverage_rate": 0.7788,
                                "text_coverage_rate": 0.3641,
                                "matched_window_count": 5,
                                "semantic_score": 0.8815,
                            },
                            {
                                "review_priority": 2,
                                "book_id": "book-b",
                                "book_name": "Drama B",
                                "matched_episode_order": 1,
                                "aggregate_text_coverage_rate": 0.7000,
                                "text_coverage_rate": 0.3500,
                                "matched_window_count": 4,
                                "semantic_score": 0.8700,
                            },
                        ],
                    },
                    "candidates": [
                        {"book_id": "book-a", "evidence_window_uid": "book-a:1:1-5", "window_text": "Shared subtitle A"},
                        {"book_id": "book-b", "evidence_window_uid": "book-b:1:1-4", "window_text": "Shared subtitle B"},
                    ],
                },
            }],
        }

        rows = build_matching_result_rows(
            detail,
            [{
                "source_video_id": "video-ambiguous",
                "source_segment_order": 1,
                "source_text_original": "Source subtitle with shared content.",
            }],
        )

        self.assertEqual(rows[0]["match_status"], "content_matched_ambiguous")
        self.assertEqual(rows[0]["matched_book_names"], "Drama A | Drama B")
        self.assertEqual(len(rows[0]["strong_candidates"]), 2)
        self.assertEqual(rows[0]["evidence_pairs"][0]["window_uid"], "book-a:1:1-5")
        self.assertEqual(rows[0]["evidence_pairs"][0]["coverage_text"], "77.88%")
        self.assertIn("内容已命中", rows[0]["user_message"])

    def test_keeps_translation_assisted_match_as_review_not_confirmed(self) -> None:
        detail = {
            "items": [{
                "source_video_id": "video-translation",
                "source_segment_order": 1,
                "result_payload_json": {
                    "decision": {
                        "matched": False,
                        "status": "review_required",
                        "outcome": "translation_assisted_match",
                        "user_message": "Translation-assisted candidate requires review.",
                    },
                    "translation_fallback": {
                        "status": "matched",
                        "source_language_code": "en",
                        "matched_target_language_code": "zh",
                        "attempts": [{"target_language_code": "zh", "cache_hit": False}],
                    },
                    "translated_query_text": "Translated subtitle",
                    "translation_candidates": [{
                        "book_id": "book-translation",
                        "book_name": "Translated Drama",
                        "episode_order": 2,
                        "window_text": "Candidate evidence",
                    }],
                },
            }],
        }
        rows = build_matching_result_rows(
            detail,
            [{"source_video_id": "video-translation", "source_segment_order": 1, "source_text_original": "English source"}],
        )

        self.assertEqual(rows[0]["match_status"], "translation_assisted_match")
        self.assertEqual(rows[0]["matched_segment_count"], 0)
        self.assertEqual(rows[0]["review_segment_count"], 1)
        self.assertEqual(rows[0]["translated_query_text"], "Translated subtitle")
        self.assertEqual(rows[0]["translation_fallback"]["matched_target_language_code"], "zh")

    def test_login_and_compare_preserve_session_on_one_client(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig("https://matching.example"))
        response = Mock(ok=True, status_code=200)
        response.json.side_effect = [{"user": {"username": "operator"}}, {"payload": {"candidates": []}}]
        client._session.request = Mock(return_value=response)  # noqa: SLF001

        client.login("operator", "password")
        result = client.compare("subtitle text", language_code="en")

        self.assertEqual(result, {"payload": {"candidates": []}})
        self.assertEqual(client._session.request.call_count, 2)  # noqa: SLF001
        self.assertEqual(client._session.request.call_args.kwargs["json"]["language_code"], "en")  # noqa: SLF001

    def test_video_compare_posts_complete_subtitle_and_cues(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig("https://matching.example"))
        response = Mock(ok=True, status_code=200)
        response.json.return_value = {"payload": {"video_decision": {"outcome": "no_match"}}}
        client._session.request = Mock(return_value=response)  # noqa: SLF001

        result = client.video_compare(
            "Complete subtitle text.",
            cues=[{"start_seconds": 0, "end_seconds": 2.5, "text": "First cue."}],
            language_code="en",
            top_k=50,
            semantic_enabled=False,
        )

        self.assertEqual(result["payload"]["video_decision"]["outcome"], "no_match")
        request = client._session.request.call_args  # noqa: SLF001
        self.assertTrue(any("/api/v1/drama-subtitles/video-compare" in str(value) for value in request.args))
        body = request.kwargs["json"]
        self.assertEqual(body["query_text"], "Complete subtitle text.")
        self.assertEqual(body["cues"][0]["text"], "First cue.")
        self.assertEqual(body["top_k"], 20)
        self.assertFalse(body["semantic_enabled"])
        self.assertTrue(body["translation_fallback"])

    def test_submit_batch_builds_parser_compatible_csv(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig("https://matching.example"))
        response = Mock(ok=True, status_code=200)
        response.json.return_value = {"task_id": "task-1", "status": "queued"}
        client._session.request = Mock(return_value=response)  # noqa: SLF001

        result = client.submit_batch(
            [
                {
                    "source_ref": "https://www.youtube.com/watch?v=test",
                    "query_text": "台词内容",
                    "source_display_title": "测试视频",
                }
            ]
        )

        self.assertEqual(result["task_id"], "task-1")
        files = client._session.request.call_args.kwargs["files"]  # noqa: SLF001
        csv_text = files["file"][1].decode("utf-8-sig")
        self.assertIn("source_ref,query_text", csv_text)
        self.assertIn("https://www.youtube.com/watch?v=test", csv_text)
        form_data = client._session.request.call_args.kwargs["data"]  # noqa: SLF001
        self.assertEqual(form_data["semantic_enabled"], "true")
        self.assertEqual(form_data["semantic_window_limit"], 100)

    def test_partial_failed_is_a_terminal_task_status(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig("https://matching.example"))
        client.task_detail = Mock(return_value={"task": {"status": "partial_failed"}})  # type: ignore[method-assign]

        detail = client.wait_for_task("task-1", poll_seconds=0.2)

        self.assertEqual(detail["task"]["status"], "partial_failed")
        client.task_detail.assert_called_once_with("task-1")

    def test_requests_evidence_context_with_bounded_parameters(self) -> None:
        client = DramaSubtitleMatchingClient(MatchingServiceConfig("https://matching.example"))
        response = Mock(ok=True, status_code=200)
        response.json.return_value = {"context": {"context": {"text": "Evidence text."}}}
        client._session.request = Mock(return_value=response)  # noqa: SLF001

        result = client.evidence_context("book-1:2:3-6", context_chars=9000, before_lines=99)

        self.assertEqual(result["context"]["context"]["text"], "Evidence text.")
        request = client._session.request.call_args  # noqa: SLF001
        self.assertTrue(any("/api/v1/drama-subtitles/evidence-context/book-1:2:3-6" in str(value) for value in request.args))
        self.assertEqual(request.kwargs["params"], {"context_chars": 5000, "before_lines": 30})


class YouTubeAsrAdapterTests(unittest.TestCase):
    def test_uses_migrated_audio_download_and_transcription_services(self) -> None:
        service = YouTubeAsrService()
        prepared = PreparedAudio(
            source_path="audio.m4a",
            audio_path="audio.mp3",
            duration_ms=1000,
            size_bytes=128,
            chunk_paths=["audio.mp3"],
            chunk_offsets_ms=[0],
        )
        transcription = AudioTranscriptionResult(
            source_path="audio.m4a",
            audio_path="audio.mp3",
            text="识别台词",
            srt_text="",
        )
        with (
            patch.object(service, "is_ready", return_value=True),
            patch.object(
                service._audio_service,
                "download_audio",
                return_value=SimpleNamespace(local_path="audio.m4a"),
            ) as download_mock,
            patch.object(
                service._transcription_service,
                "transcribe_source",
                return_value=(prepared, transcription),
            ) as transcribe_mock,
        ):
            result = service.transcribe_video("https://www.youtube.com/watch?v=test", leading_seconds=180)

        self.assertEqual(result.text, "识别台词")
        self.assertEqual(download_mock.call_args.kwargs["max_duration_seconds"], 180)
        self.assertEqual(download_mock.call_args.kwargs["concurrency"], 6)
        transcribe_mock.assert_called_once_with("audio.m4a", should_cancel=None)

    def test_workflow_uses_bounded_download_and_asr_stages(self) -> None:
        videos = [
            YouTubeVideo("video-1", "https://www.youtube.com/watch?v=video-1", "Video 1"),
            YouTubeVideo("video-2", "https://www.youtube.com/watch?v=video-2", "Video 2"),
        ]
        pending = [
            {"video": video.to_dict(), "inspection": {"status": "asr_required", "asr_required": True}}
            for video in videos
        ]
        download_segments: list[int] = []

        class WorkerService:
            def download_video_audio(self, source_url, *, segment_concurrency, **_kwargs):
                download_segments.append(segment_concurrency)
                return f"{source_url}.m4a"

            def transcribe_audio_source(self, source_path, **_kwargs):
                return SimpleNamespace(source_path=source_path, audio_path=source_path, text="识别台词")

        workflow = VerificationWorkflow(asr_service=MagicMock(is_ready=Mock(return_value=True)))
        with patch("app.workflow.YouTubeAsrService", side_effect=WorkerService):
            ready, still_pending = workflow.prepare_asr_fallback_items(
                pending,
                download_concurrency=2,
                asr_concurrency=2,
            )

        self.assertEqual(len(ready), 2)
        self.assertEqual(still_pending, [])
        self.assertEqual(download_segments, [3, 3])


class ProxyAndRangeRouteTests(unittest.TestCase):
    def test_normalizes_windows_proxy_forms(self) -> None:
        self.assertEqual(ProxyDiscoveryService.normalize_proxy("127.0.0.1:7897"), "http://127.0.0.1:7897")
        self.assertEqual(
            ProxyDiscoveryService.normalize_proxy("http=127.0.0.1:7890;https=127.0.0.1:7897"),
            "http://127.0.0.1:7897",
        )
        self.assertEqual(ProxyDiscoveryService.normalize_proxy("DIRECT"), "")
        self.assertEqual(ProxyDiscoveryService.normalize_proxy("invalid-address"), "")

    def test_failed_range_routes_enter_cooldown_and_success_resets_them(self) -> None:
        service = YouTubeAudioService()
        self.assertTrue(service._is_direct_route_available("direct"))  # noqa: SLF001

        service._mark_direct_route_failure("direct")  # noqa: SLF001
        self.assertFalse(service._is_direct_route_available("direct"))  # noqa: SLF001

        service._mark_direct_route_success("direct")  # noqa: SLF001
        self.assertTrue(service._is_direct_route_available("direct"))  # noqa: SLF001

        service._mark_direct_route_failure("configured-proxy")  # noqa: SLF001
        self.assertTrue(service._is_direct_route_available("configured-proxy"))  # noqa: SLF001
        service._mark_direct_route_failure("configured-proxy")  # noqa: SLF001
        self.assertFalse(service._is_direct_route_available("configured-proxy"))  # noqa: SLF001

    def test_proxy_403_refreshes_metadata_once_before_fallback(self) -> None:
        metadata = {
            "url": "https://media.example/audio",
            "filesize": 2_000_000,
            "duration": 180,
            "ext": "m4a",
            "http_headers": {},
        }
        response_403 = MagicMock()
        response_403.__enter__.return_value = response_403
        response_403.status_code = 403
        response_ok = MagicMock()
        response_ok.__enter__.return_value = response_ok
        response_ok.status_code = 206
        response_ok.iter_content.return_value = [b"x" * 256_000]
        session = MagicMock()
        session.get.side_effect = [response_403, response_ok]
        service = YouTubeAudioService()
        with TemporaryDirectory() as directory, patch(
            "app.services.youtube_audio_service.EXTRACTED_AUDIO_DIR", Path(directory)
        ), patch.object(
            service._youtube, "_ffmpeg_proxy_from_environment", return_value="http://127.0.0.1:7897"
        ), patch.object(
            service, "_resolve_direct_range_metadata", side_effect=[metadata, metadata]
        ) as resolve_mock, patch(
            "app.services.youtube_audio_service.requests.Session", return_value=session
        ), patch.object(service, "_read_last_audio_timestamp", return_value=180.0):
            result = service._download_audio_with_direct_range(
                "https://www.youtube.com/watch?v=test",
                duration=180,
                requested_duration=180,
                progress_callback=None,
                should_cancel=None,
            )

        self.assertTrue(Path(result.local_path).name.endswith(".m4a"))
        self.assertEqual(resolve_mock.call_count, 2)
        self.assertEqual(session.get.call_count, 2)

    def test_compat_process_stops_immediately_when_cancelled(self) -> None:
        process = MagicMock()
        process.poll.return_value = None
        process.communicate.return_value = ("", "")
        service = YouTubeAudioService()
        with patch("app.services.youtube_audio_service.subprocess.Popen", return_value=process), patch.object(
            service._youtube, "_terminate_process_tree"
        ) as terminate_mock:
            with self.assertRaises(YouTubeDownloadCancelled):
                service._run_compat_process(  # noqa: SLF001
                    ["yt-dlp"],
                    env={},
                    should_cancel=lambda: True,
                    timeout_seconds=90,
                    description="test segment",
                )

        terminate_mock.assert_called_once_with(process)

    def test_compat_timeout_diagnostics_identifies_missing_first_byte(self) -> None:
        diagnostics = YouTubeAudioService._compat_timeout_diagnostics(  # noqa: SLF001
            started_at=0.0,
            observed_bytes=0,
            first_output_at=None,
            last_growth_at=None,
        )

        self.assertIn("state=no_first_byte", diagnostics)
        self.assertIn("bytes=0", diagnostics)

    def test_compat_segment_bytes_includes_partial_download(self) -> None:
        with TemporaryDirectory() as directory:
            output_dir = Path(directory)
            (output_dir / "segment_002.m4a.part").write_bytes(b"partial")
            (output_dir / "segment_002.m4a").write_bytes(b"complete")
            (output_dir / "segment_003.m4a").write_bytes(b"other")

            observed = YouTubeAudioService._compat_segment_bytes(output_dir, 2)  # noqa: SLF001

        self.assertEqual(observed, len(b"partialcomplete"))


class TaskControlTests(unittest.TestCase):
    def test_pause_resume_and_cancel_checkpoints(self) -> None:
        control = TaskControl()
        entered = Event()
        completed = Event()

        def wait_at_checkpoint() -> None:
            entered.set()
            if control.checkpoint():
                completed.set()

        control.pause()
        worker = Thread(target=wait_at_checkpoint)
        worker.start()
        self.assertTrue(entered.wait(0.5))
        sleep(0.05)
        self.assertFalse(completed.is_set())
        control.resume()
        worker.join(0.5)
        self.assertTrue(completed.is_set())

        control.pause()
        control.cancel()
        self.assertTrue(control.cancelled)
        self.assertFalse(control.paused)
        self.assertFalse(control.checkpoint())


class TaskStateStoreTests(unittest.TestCase):
    def test_round_trips_atomic_snapshot(self) -> None:
        with TemporaryDirectory() as directory:
            state_path = Path(directory) / "runtime" / "session_state.json"
            with patch("app.services.task_state.TASK_STATE_PATH", state_path):
                store = TaskStateStore()
                store.save({"active": True, "items": [{"video_id": "video-1", "status": "pending"}]})
                loaded = store.load()

        self.assertIsNotNone(loaded)
        self.assertTrue(loaded["active"])
        self.assertEqual(loaded["items"][0]["video_id"], "video-1")
        self.assertTrue(loaded["updated_at"])


if __name__ == "__main__":
    unittest.main()
