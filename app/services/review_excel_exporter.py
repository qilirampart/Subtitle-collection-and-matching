from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _write_sheet(sheet, headers: tuple[str, ...], rows: Iterable[tuple[object, ...]], widths: tuple[int, ...]) -> int:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count = 0
    for row in rows:
        sheet.append(tuple(str(value) if value is not None else "" for value in row))
        count += 1
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return count


def _json_text(value: object) -> str:
    if not value:
        return ""
    try:
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    except (TypeError, ValueError):
        return str(value)


def export_matching_results_to_xlsx(output_path: Path, rows: list[dict[str, object]]) -> int:
    """Export one complete review record per source video."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "字幕匹配结果"
    headers = (
        "来源频道", "来源剧名", "原视频链接", "视频 ID", "完整字幕", "匹配状态", "用户结论",
        "命中剧名", "Book ID", "命中集数", "命中时间范围", "匹配原因", "确认命中数",
        "待复核数", "未命中数", "翻译回退", "译文", "服务端执行", "强证据候选", "命中证据",
    )
    exported = _write_sheet(
        sheet,
        headers,
        (
            (
                row.get("source_channel"), row.get("source_title"), row.get("source_url"),
                row.get("source_video_id"), row.get("source_subtitle"), row.get("match_status"),
                row.get("user_message"), row.get("matched_book_names"), row.get("matched_book_ids"),
                row.get("matched_episode_orders"), row.get("matched_time_ranges"), row.get("match_reasons"),
                row.get("matched_segment_count"), row.get("review_segment_count"), row.get("not_matched_segment_count"),
                _json_text(row.get("translation_fallback")), row.get("translated_query_text"),
                _json_text(row.get("execution")), _json_text(row.get("strong_candidates")),
                _json_text(row.get("evidence_pairs")),
            )
            for row in rows
            if isinstance(row, dict)
        ),
        (20, 38, 60, 18, 100, 20, 42, 32, 18, 16, 22, 42, 14, 14, 14, 54, 80, 48, 70, 100),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return exported


def export_cover_review_results_to_xlsx(
    output_path: Path,
    reviews: Iterable[object],
    source_urls: dict[str, str],
    thumbnail_urls: dict[str, str] | None = None,
    channel_ids: dict[str, str] | None = None,
    channel_names: dict[str, str] | None = None,
) -> int:
    """Export complete cover-review responses, including the original model reply."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "封面检测结果"
    headers = (
        "频道 ID", "频道名", "视频标题", "原视频链接", "视频 ID", "封面 CDN 地址", "封面文件", "检测结论", "风险标签", "摘要",
        "可见证据", "置信度", "模型原始回复", "错误信息",
    )
    exported = _write_sheet(
        sheet,
        headers,
        (
            (
                (channel_ids or {}).get(str(getattr(review, "video_id", "")), ""),
                (channel_names or {}).get(str(getattr(review, "video_id", "")), ""), getattr(review, "title", ""),
                source_urls.get(str(getattr(review, "video_id", "")), ""), getattr(review, "video_id", ""),
                getattr(review, "thumbnail_url", "") or (thumbnail_urls or {}).get(str(getattr(review, "video_id", "")), ""),
                getattr(review, "cover_path", ""),
                getattr(review, "overall_risk", ""), "、".join(getattr(review, "risk_tags", ()) or ()),
                getattr(review, "summary", ""), getattr(review, "evidence", ""),
                getattr(review, "confidence", ""), getattr(review, "model_response", ""),
                getattr(review, "error", ""),
            )
            for review in reviews
        ),
        (28, 26, 38, 60, 18, 70, 55, 16, 24, 42, 60, 12, 100),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return exported


__all__ = ["export_cover_review_results_to_xlsx", "export_matching_results_to_xlsx"]
