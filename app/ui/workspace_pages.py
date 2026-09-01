from __future__ import annotations

import csv
from collections.abc import Callable
from difflib import SequenceMatcher
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QPixmap, QTextCharFormat, QTextCursor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QFrame,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QComboBox,
    QLineEdit,
    QPlainTextEdit,
    QPushButton,
    QMessageBox,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


Navigate = Callable[[str], None]


def _channel_urls_from_file(path: Path) -> list[str]:
    """Read channel URL candidates from text, CSV, or Excel files."""
    suffix = path.suffix.lower()
    values: list[str] = []
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("导入 Excel 需要安装 openpyxl。") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for row in workbook.active.iter_rows(values_only=True):
                values.extend(str(value).strip() for value in row if value is not None)
        finally:
            workbook.close()
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                values.extend(str(value).strip() for value in row)
    else:
        values = path.read_text(encoding="utf-8-sig").splitlines()
    urls: list[str] = []
    seen: set[str] = set()
    for value in values:
        candidate = str(value or "").strip()
        if not candidate or "youtube.com" not in candidate.lower() and "youtu.be" not in candidate.lower():
            continue
        if candidate not in seen:
            seen.add(candidate)
            urls.append(candidate)
    return urls


class ChannelListDialog(QDialog):
    """Editable channel list that always keeps one trailing blank row."""

    def __init__(self, urls: list[str] | None = None, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("频道清单")
        self.resize(820, 560)
        root = QVBoxLayout(self)
        hint = QLabel("每行填写一个 YouTube 频道链接。填写最后一行后会自动新增空行，也可批量导入文件。")
        hint.setWordWrap(True)
        root.addWidget(hint)

        actions = QHBoxLayout()
        import_button = QPushButton("批量导入 TXT / CSV / Excel")
        import_button.setProperty("secondary", True)
        import_button.clicked.connect(self._import_file)
        remove_button = QPushButton("删除选中行")
        remove_button.setProperty("secondary", True)
        remove_button.clicked.connect(self._remove_selected_rows)
        clear_button = QPushButton("清空")
        clear_button.setProperty("danger", True)
        clear_button.clicked.connect(self._clear)
        actions.addWidget(import_button)
        actions.addWidget(remove_button)
        actions.addWidget(clear_button)
        actions.addStretch(1)
        root.addLayout(actions)

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["序号", "频道链接"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.itemChanged.connect(self._on_item_changed)
        root.addWidget(self.table, 1)

        self.summary_label = QLabel()
        root.addWidget(self.summary_label)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Ok)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("确认使用")
        buttons.accepted.connect(self._accept)
        buttons.rejected.connect(self.reject)
        root.addWidget(buttons)
        self._replace_urls(urls or [])

    def urls(self) -> list[str]:
        result: list[str] = []
        seen: set[str] = set()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 1)
            value = item.text().strip() if item is not None else ""
            if value and value not in seen:
                seen.add(value)
                result.append(value)
        return result

    def _replace_urls(self, urls: list[str]) -> None:
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        for url in urls:
            self._append_row(str(url or "").strip())
        self._append_row("")
        self.table.blockSignals(False)
        self._renumber()
        self._update_summary()

    def _append_row(self, value: str) -> None:
        row = self.table.rowCount()
        self.table.insertRow(row)
        number = QTableWidgetItem(str(row + 1))
        number.setFlags(number.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 0, number)
        self.table.setItem(row, 1, QTableWidgetItem(value))

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() != 1:
            return
        if item.row() == self.table.rowCount() - 1 and item.text().strip():
            self.table.blockSignals(True)
            self._append_row("")
            self.table.blockSignals(False)
        self._renumber()
        self._update_summary()

    def _renumber(self) -> None:
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item is not None:
                item.setText(str(row + 1))

    def _update_summary(self) -> None:
        values = []
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 1)
            value = item.text().strip() if item is not None else ""
            if value:
                values.append(value)
        self.summary_label.setText(f"已填写 {len(values)} 条，去重后 {len(dict.fromkeys(values))} 条")

    def _import_file(self) -> None:
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "导入频道清单",
            "",
            "频道清单 (*.txt *.csv *.xlsx);;文本文件 (*.txt);;CSV 文件 (*.csv);;Excel 文件 (*.xlsx)",
        )
        if not file_name:
            return
        try:
            imported = _channel_urls_from_file(Path(file_name))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "导入失败", str(exc))
            return
        if not imported:
            QMessageBox.information(self, "没有频道链接", "文件中没有识别到 YouTube 频道链接。")
            return
        existing = self.urls()
        self._replace_urls(list(dict.fromkeys(existing + imported)))
        self.summary_label.setText(f"已导入 {len(imported)} 条；当前去重后共 {len(self.urls())} 条")

    def _remove_selected_rows(self) -> None:
        rows = sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True)
        for row in rows:
            self.table.removeRow(row)
        values = self.urls()
        self._replace_urls(values)

    def _clear(self) -> None:
        self._replace_urls([])

    def _accept(self) -> None:
        if not self.urls():
            QMessageBox.information(self, "频道清单为空", "请至少填写或导入一个频道链接。")
            return
        self.accept()


def _section_title(title: str, hint: str = "") -> QWidget:
    widget = QWidget()
    layout = QVBoxLayout(widget)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(2)
    heading = QLabel(title)
    heading.setObjectName("PageTitle")
    layout.addWidget(heading)
    if hint:
        description = QLabel(hint)
        description.setObjectName("PageHint")
        description.setWordWrap(True)
        layout.addWidget(description)
    return widget


class DashboardPage(QWidget):
    """Lightweight landing page for the modular workspace."""

    def __init__(self, navigate: Navigate, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._navigate = navigate
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.addWidget(_section_title("工作台", "从单个功能开始，或按既有全流程连续处理。"), 1)
        self._status = QLabel("当前没有运行任务")
        self._status.setObjectName("WorkspaceStatus")
        self._status.setWordWrap(True)
        header_row.addWidget(self._status)
        root.addLayout(header_row)

        metrics = QGridLayout()
        metrics.setHorizontalSpacing(12)
        metrics.setVerticalSpacing(12)
        self._metric_values: dict[str, QLabel] = {}
        for index, (key, label, hint) in enumerate((
            ("videos", "视频队列", "已采集的视频"),
            ("ready", "可匹配字幕", "已取得或转写完成"),
            ("fallback", "待 ASR 兜底", "未直接取得字幕"),
            ("reviews", "封面检测", "已有检测结果"),
        )):
            card = QFrame()
            card.setObjectName("DashboardMetric")
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(16, 13, 16, 13)
            title = QLabel(label)
            title.setObjectName("MetricLabel")
            value = QLabel("0")
            value.setObjectName("MetricValue")
            subtext = QLabel(hint)
            subtext.setObjectName("MetricHint")
            card_layout.addWidget(title)
            card_layout.addWidget(value)
            card_layout.addWidget(subtext)
            metrics.addWidget(card, 0, index)
            self._metric_values[key] = value
        root.addLayout(metrics)

        entry_row = QHBoxLayout()
        entry_row.setSpacing(12)
        full_flow = self._action_card(
            "一键全流程",
            "频道采集、字幕获取、ASR 兜底与匹配核验仍保持现有稳定链路。",
            "进入全流程",
            lambda: self._navigate("full_flow"),
            primary=True,
        )
        direct = self._action_card(
            "按功能处理",
            "下载、字幕转写、匹配核验将逐页独立；当前快捷入口会定位到对应的稳定功能。",
            "选择处理方式",
            lambda: self._navigate("full_flow"),
        )
        entry_row.addWidget(full_flow, 1)
        entry_row.addWidget(direct, 1)
        root.addLayout(entry_row)

        quick_card = QFrame()
        quick_card.setObjectName("WorkspaceCard")
        quick_layout = QVBoxLayout(quick_card)
        quick_layout.setContentsMargins(16, 14, 16, 14)
        quick_layout.addWidget(QLabel("快速开始"))
        quick_buttons = QGridLayout()
        actions = (
            ("采集频道或 Shorts", "collect"),
            ("下载音频并转写", "download"),
            ("获取字幕", "subtitles"),
            ("提交匹配核验", "matching"),
            ("检测视频封面", "cover"),
            ("查看任务状态", "tasks"),
        )
        for index, (label, target) in enumerate(actions):
            button = QPushButton(label)
            button.setProperty("secondary", True)
            button.clicked.connect(lambda _checked=False, page=target: self._navigate(page))
            quick_buttons.addWidget(button, index // 3, index % 3)
        quick_layout.addLayout(quick_buttons)
        root.addWidget(quick_card)

        note = QLabel(
            "本轮改造先保留全流程页面作为兼容入口。独立下载、字幕转写、匹配核验页面会在不改变既有服务逻辑的前提下逐步迁移。"
        )
        note.setObjectName("PageHint")
        note.setWordWrap(True)
        root.addWidget(note)
        root.addStretch(1)

    @staticmethod
    def _action_card(title: str, hint: str, action: str, callback: Callable[[], None], *, primary: bool = False) -> QFrame:
        card = QFrame()
        card.setObjectName("QuickActionCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(8)
        heading = QLabel(title)
        heading.setObjectName("QuickActionTitle")
        body = QLabel(hint)
        body.setObjectName("PageHint")
        body.setWordWrap(True)
        button = QPushButton(action)
        if primary:
            button.setProperty("primary", True)
        else:
            button.setProperty("secondary", True)
        button.clicked.connect(callback)
        layout.addWidget(heading)
        layout.addWidget(body, 1)
        layout.addWidget(button, 0, Qt.AlignmentFlag.AlignLeft)
        return card

    def refresh(self, snapshot: dict[str, object]) -> None:
        for key, value in (
            ("videos", snapshot.get("videos", 0)),
            ("ready", snapshot.get("ready", 0)),
            ("fallback", snapshot.get("fallback", 0)),
            ("reviews", snapshot.get("reviews", 0)),
        ):
            self._metric_values[key].setText(str(value))
        self._status.setText(str(snapshot.get("status") or "当前没有运行任务"))


class TaskCenterPage(QWidget):
    """Session task overview. Persistent recovery remains owned by TaskStateStore."""

    def __init__(self, open_full_flow: Callable[[], None], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._open_full_flow = open_full_flow
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("任务中心", "查看当前会话中正在处理的采集、字幕、ASR、封面检测和匹配任务。"))

        self._summary = QLabel("当前没有运行任务。")
        self._summary.setObjectName("WorkspaceStatus")
        self._summary.setWordWrap(True)
        root.addWidget(self._summary)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["任务", "当前状态", "队列", "已完成", "待处理"])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setMinimumHeight(240)
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()
        open_button = QPushButton("打开全流程任务页")
        open_button.setProperty("primary", True)
        open_button.clicked.connect(self._open_full_flow)
        actions.addWidget(open_button)
        actions.addStretch(1)
        root.addLayout(actions)

    def refresh(self, snapshot: dict[str, object]) -> None:
        active = bool(snapshot.get("active"))
        status = str(snapshot.get("status") or "当前没有运行任务")
        self._summary.setText(status)
        rows = (
            ("视频队列", "已就绪" if snapshot.get("videos") else "等待采集", snapshot.get("videos", 0), snapshot.get("videos", 0), 0),
            (
                "独立下载",
                "处理中" if active and snapshot.get("task_kind") == "视频下载" else "按需执行",
                snapshot.get("downloads", 0),
                snapshot.get("download_completed", 0),
                max(0, int(snapshot.get("downloads", 0)) - int(snapshot.get("download_completed", 0))),
            ),
            ("字幕获取", "处理中" if active and snapshot.get("task_kind") == "字幕获取" else "等待处理", snapshot.get("videos", 0), snapshot.get("ready", 0), snapshot.get("fallback", 0)),
            ("ASR 兜底", "处理中" if active and snapshot.get("task_kind") == "ASR 兜底" else "按需执行", snapshot.get("fallback", 0), snapshot.get("ready", 0), snapshot.get("fallback", 0)),
            ("封面检测", "处理中" if active and snapshot.get("task_kind") == "封面检测" else "按需执行", snapshot.get("videos", 0), snapshot.get("reviews", 0), 0),
            ("匹配核验", "处理中" if active and snapshot.get("task_kind") == "匹配核验" else "等待提交", snapshot.get("ready", 0), 0, snapshot.get("ready", 0)),
        )
        self.table.setRowCount(len(rows))
        for row, values in enumerate(rows):
            for column, value in enumerate(values):
                self.table.setItem(row, column, QTableWidgetItem(str(value)))


class DownloadPage(QWidget):
    """Standalone download surface backed by the existing YouTube services."""

    def __init__(
        self,
        *,
        start_download: Callable[[list[str], str, int, int], None],
        pause_download: Callable[[], None],
        cancel_download: Callable[[], None],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._start_download = start_download
        self._pause_download = pause_download
        self._cancel_download = cancel_download
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("视频下载", "粘贴单条或多条 YouTube 视频链接。下载完成后素材会保留在本地输出目录。"))

        input_card = QFrame()
        input_card.setObjectName("WorkspaceCard")
        input_layout = QGridLayout(input_card)
        input_layout.setContentsMargins(16, 14, 16, 14)
        input_layout.setHorizontalSpacing(12)
        input_layout.setVerticalSpacing(9)
        input_layout.addWidget(QLabel("视频链接"), 0, 0, Qt.AlignmentFlag.AlignTop)
        self.url_input = QPlainTextEdit()
        self.url_input.setPlaceholderText("每行一条 YouTube 视频链接，可直接粘贴 watch、shorts 或 youtu.be 链接")
        self.url_input.setFixedHeight(76)
        input_layout.addWidget(self.url_input, 0, 1, 1, 4)
        self.add_button = QPushButton("加入下载列表")
        self.add_button.setProperty("secondary", True)
        self.add_button.clicked.connect(self._add_urls)
        input_layout.addWidget(self.add_button, 0, 5, Qt.AlignmentFlag.AlignTop)

        self.content_combo = QComboBox()
        self.content_combo.addItem("下载视频", "video")
        self.content_combo.addItem("仅下载音频", "audio")
        self.content_combo.addItem("视频和音频都下载", "both")
        self.duration_combo = QComboBox()
        self.duration_combo.addItem("完整内容", 0)
        self.duration_combo.addItem("前 1 分钟", 60)
        self.duration_combo.addItem("前 3 分钟", 180)
        self.duration_combo.addItem("前 5 分钟", 300)
        self.duration_combo.setCurrentIndex(2)
        self.concurrency_spin = QSpinBox()
        self.concurrency_spin.setRange(1, 6)
        self.concurrency_spin.setValue(3)
        for widget in (self.content_combo, self.duration_combo, self.concurrency_spin):
            widget.setFixedHeight(32)
        input_layout.addWidget(QLabel("下载内容"), 1, 0)
        input_layout.addWidget(self.content_combo, 1, 1)
        input_layout.addWidget(QLabel("下载范围"), 1, 2)
        input_layout.addWidget(self.duration_combo, 1, 3)
        input_layout.addWidget(QLabel("音频并发"), 1, 4)
        input_layout.addWidget(self.concurrency_spin, 1, 5)
        for column in (1, 3):
            input_layout.setColumnStretch(column, 1)
        root.addWidget(input_card)

        queue_card = QFrame()
        queue_card.setObjectName("WorkspaceCard")
        queue_layout = QVBoxLayout(queue_card)
        queue_layout.setContentsMargins(16, 14, 16, 14)
        queue_header = QHBoxLayout()
        self.queue_title = QLabel("下载列表（0）")
        self.queue_title.setObjectName("QuickActionTitle")
        clear_button = QPushButton("清空列表")
        clear_button.clicked.connect(self._clear_jobs)
        queue_header.addWidget(self.queue_title)
        queue_header.addStretch(1)
        queue_header.addWidget(clear_button)
        queue_layout.addLayout(queue_header)
        self.queue_table = QTableWidget(0, 5)
        self.queue_table.setHorizontalHeaderLabels(["链接", "下载内容", "范围", "状态", "输出文件"])
        self.queue_table.verticalHeader().setVisible(False)
        self.queue_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.queue_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        header = self.queue_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.queue_table.setMinimumHeight(250)
        queue_layout.addWidget(self.queue_table, 1)
        root.addWidget(queue_card, 1)

        footer = QHBoxLayout()
        self.status_label = QLabel("请粘贴视频链接并加入下载列表。")
        self.status_label.setObjectName("WorkspaceStatus")
        self.status_label.setWordWrap(True)
        self.start_button = QPushButton("开始下载")
        self.start_button.setProperty("primary", True)
        self.start_button.clicked.connect(self._request_start)
        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(self._pause_download)
        self.cancel_button = QPushButton("取消")
        self.cancel_button.setProperty("danger", True)
        self.cancel_button.clicked.connect(self._cancel_download)
        footer.addWidget(self.status_label, 1)
        footer.addWidget(self.pause_button)
        footer.addWidget(self.cancel_button)
        footer.addWidget(self.start_button)
        root.addLayout(footer)
        self._jobs: list[str] = []
        self.set_busy(False)

    def _add_urls(self) -> None:
        values = [line.strip() for line in self.url_input.toPlainText().splitlines() if line.strip()]
        if not values:
            self.status_label.setText("没有读取到可加入的链接。")
            return
        self._jobs.extend(values)
        self.url_input.clear()
        self._refresh_table()
        self.status_label.setText(f"已加入 {len(values)} 条链接，当前列表共 {len(self._jobs)} 条。")

    def _clear_jobs(self) -> None:
        self._jobs = []
        self._refresh_table()
        self.status_label.setText("下载列表已清空。")

    def _request_start(self) -> None:
        if not self._jobs:
            self.status_label.setText("请先加入至少一条视频链接。")
            return
        self._start_download(
            list(self._jobs),
            str(self.content_combo.currentData() or "video"),
            int(self.duration_combo.currentData() or 0),
            self.concurrency_spin.value(),
        )

    def _refresh_table(self) -> None:
        self.queue_table.setRowCount(len(self._jobs))
        mode = self.content_combo.currentText()
        duration = self.duration_combo.currentText()
        for row, url in enumerate(self._jobs):
            self.queue_table.setItem(row, 0, QTableWidgetItem(url))
            self.queue_table.setItem(row, 1, QTableWidgetItem(mode))
            self.queue_table.setItem(row, 2, QTableWidgetItem(duration))
            self.queue_table.setItem(row, 3, QTableWidgetItem("等待下载"))
            self.queue_table.setItem(row, 4, QTableWidgetItem(""))
        self.queue_title.setText(f"下载列表（{len(self._jobs)}）")

    def update_job(self, row: int, status: str, output: str = "") -> None:
        if row < 0 or row >= self.queue_table.rowCount():
            return
        self.queue_table.setItem(row, 3, QTableWidgetItem(status))
        if output:
            self.queue_table.setItem(row, 4, QTableWidgetItem(output))

    def set_busy(self, busy: bool) -> None:
        for widget in (
            self.url_input,
            self.add_button,
            self.content_combo,
            self.duration_combo,
            self.concurrency_spin,
            self.start_button,
        ):
            widget.setEnabled(not busy)
        self.pause_button.setEnabled(busy)
        self.cancel_button.setEnabled(busy)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)


class SubtitlePage(QWidget):
    """Independent caption acquisition and ASR surface."""

    def __init__(
        self,
        *,
        start_transcription: Callable[[list[str], int, bool, bool], None],
        import_downloads: Callable[[], list[str]],
        pause_transcription: Callable[[], None],
        cancel_transcription: Callable[[], None],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._start_transcription = start_transcription
        self._import_downloads = import_downloads
        self._sources: list[str] = []
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("字幕与转写", "优先获取 YouTube 直出字幕；没有字幕时，再按你的选择调用 ASR。"))

        input_card = QFrame()
        input_card.setObjectName("WorkspaceCard")
        input_layout = QGridLayout(input_card)
        input_layout.setContentsMargins(16, 14, 16, 14)
        input_layout.setHorizontalSpacing(10)
        input_layout.setVerticalSpacing(8)
        input_layout.addWidget(QLabel("输入素材"), 0, 0, Qt.AlignmentFlag.AlignTop)
        self.source_input = QPlainTextEdit()
        self.source_input.setPlaceholderText("每行一条 YouTube 视频链接或本地视频/音频路径")
        self.source_input.setFixedHeight(70)
        input_layout.addWidget(self.source_input, 0, 1, 1, 3)
        add_button = QPushButton("加入列表")
        add_button.setProperty("secondary", True)
        add_button.clicked.connect(self._add_sources)
        input_layout.addWidget(add_button, 0, 4, Qt.AlignmentFlag.AlignTop)
        local_button = QPushButton("选择本地文件")
        local_button.clicked.connect(self._choose_local_files)
        input_layout.addWidget(local_button, 1, 1)
        self.downloaded_button = QPushButton("导入“视频下载”完成素材 (0)")
        self.downloaded_button.setToolTip("导入“视频下载”页面成功完成的文件；每条下载任务只加入一个素材。")
        self.downloaded_button.clicked.connect(self._add_downloaded)
        self.downloaded_button.setEnabled(False)
        input_layout.addWidget(self.downloaded_button, 1, 2)
        self.duration_combo = QComboBox()
        self.duration_combo.addItem("前 1 分钟", 60)
        self.duration_combo.addItem("前 3 分钟", 180)
        self.duration_combo.addItem("前 5 分钟", 300)
        self.duration_combo.setCurrentIndex(1)
        self.asr_fallback_check = QCheckBox("无直出字幕时执行 ASR 兜底")
        self.asr_fallback_check.setChecked(True)
        self.skip_caption_probe_check = QCheckBox("已知无字幕时跳过探测，直接下载转写")
        self.skip_caption_probe_check.setToolTip(
            "适用于已确认没有字幕接口的视频；开启后不请求字幕接口，直接下载音频并进行 ASR。"
        )
        input_layout.addWidget(QLabel("识别范围"), 1, 3)
        input_layout.addWidget(self.duration_combo, 1, 4)
        input_layout.addWidget(self.asr_fallback_check, 2, 1, 1, 4)
        input_layout.addWidget(self.skip_caption_probe_check, 3, 1, 1, 4)
        input_layout.setColumnStretch(1, 1)
        root.addWidget(input_card)

        content_splitter = QSplitter(Qt.Orientation.Vertical)
        content_splitter.setChildrenCollapsible(False)
        list_card = QFrame()
        list_card.setObjectName("WorkspaceCard")
        list_card.setMinimumHeight(280)
        list_layout = QVBoxLayout(list_card)
        list_layout.setContentsMargins(16, 14, 16, 14)
        self.queue_title = QLabel("字幕任务（0）")
        self.queue_title.setObjectName("QuickActionTitle")
        list_layout.addWidget(self.queue_title)
        self.queue_table = QTableWidget(0, 5)
        self.queue_table.setHorizontalHeaderLabels(["素材", "类型", "字幕来源", "状态", "字数"])
        self.queue_table.verticalHeader().setVisible(False)
        self.queue_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.queue_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.queue_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.queue_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.queue_table.itemSelectionChanged.connect(self._show_selected_preview)
        list_layout.addWidget(self.queue_table, 1)
        content_splitter.addWidget(list_card)

        preview_card = QFrame()
        preview_card.setObjectName("WorkspaceCard")
        preview_layout = QVBoxLayout(preview_card)
        preview_layout.setContentsMargins(16, 14, 16, 14)
        preview_layout.addWidget(QLabel("字幕结果预览"))
        self.preview_edit = QPlainTextEdit()
        self.preview_edit.setReadOnly(True)
        self.preview_edit.setPlaceholderText("选择一条素材后查看字幕文本")
        preview_layout.addWidget(self.preview_edit, 1)
        content_splitter.addWidget(preview_card)
        content_splitter.setStretchFactor(0, 3)
        content_splitter.setStretchFactor(1, 2)
        root.addWidget(content_splitter, 1)

        footer = QHBoxLayout()
        self.status_label = QLabel("请加入视频、音频或 YouTube 链接。")
        self.status_label.setObjectName("WorkspaceStatus")
        self.start_button = QPushButton("开始识别")
        self.start_button.setProperty("primary", True)
        self.start_button.clicked.connect(self._request_start)
        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(pause_transcription)
        self.cancel_button = QPushButton("取消")
        self.cancel_button.setProperty("danger", True)
        self.cancel_button.clicked.connect(cancel_transcription)
        footer.addWidget(self.status_label, 1)
        footer.addWidget(self.pause_button)
        footer.addWidget(self.cancel_button)
        footer.addWidget(self.start_button)
        root.addLayout(footer)
        self._preview_texts: dict[int, str] = {}
        self.set_busy(False)

    def _add_sources(self) -> None:
        values = [line.strip() for line in self.source_input.toPlainText().splitlines() if line.strip()]
        if not values:
            self.status_label.setText("没有读取到素材路径或链接。")
            return
        self._sources.extend(values)
        self.source_input.clear()
        self._refresh_table()
        self.status_label.setText(f"已加入 {len(values)} 条素材。")

    def _choose_local_files(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "选择视频或音频",
            "",
            "视频和音频 (*.mp4 *.mov *.mkv *.avi *.mp3 *.wav *.m4a *.webm);;所有文件 (*.*)",
        )
        if files:
            self._sources.extend(files)
            self._refresh_table()
            self.status_label.setText(f"已加入 {len(files)} 个本地文件。")

    def _add_downloaded(self) -> None:
        files = [str(value) for value in self._import_downloads() if str(value).strip()]
        if not files:
            self.status_label.setText("当前没有可导入的“视频下载”完成素材。")
            return
        existing = {source.casefold() for source in self._sources}
        additions = [path for path in files if path.casefold() not in existing]
        if not additions:
            self.status_label.setText("已下载素材均已在字幕任务列表中。")
            return
        self._sources.extend(additions)
        self._refresh_table()
        self.status_label.setText(f"已从“视频下载”导入 {len(additions)} 个素材。")

    def set_downloaded_material_count(self, count: int) -> None:
        available = max(0, int(count or 0))
        self.downloaded_button.setText(f"导入“视频下载”完成素材 ({available})")
        self.downloaded_button.setEnabled(available > 0)

    def _request_start(self) -> None:
        if not self._sources:
            self.status_label.setText("请先加入至少一条素材。")
            return
        self._start_transcription(
            list(self._sources),
            int(self.duration_combo.currentData() or 180),
            self.asr_fallback_check.isChecked() or self.skip_caption_probe_check.isChecked(),
            self.skip_caption_probe_check.isChecked(),
        )

    def _refresh_table(self) -> None:
        self.queue_table.setRowCount(len(self._sources))
        for row, source in enumerate(self._sources):
            source_type = "YouTube" if source.lower().startswith(("http://", "https://")) else "本地文件"
            self.queue_table.setItem(row, 0, QTableWidgetItem(source))
            self.queue_table.setItem(row, 1, QTableWidgetItem(source_type))
            self.queue_table.setItem(row, 2, QTableWidgetItem("待获取"))
            self.queue_table.setItem(row, 3, QTableWidgetItem("等待识别"))
            self.queue_table.setItem(row, 4, QTableWidgetItem("-"))
        self.queue_title.setText(f"字幕任务（{len(self._sources)}）")

    def update_job(self, row: int, status: str, source_kind: str = "", text: str = "") -> None:
        if row < 0 or row >= self.queue_table.rowCount():
            return
        if source_kind:
            self.queue_table.setItem(row, 2, QTableWidgetItem(source_kind))
        self.queue_table.setItem(row, 3, QTableWidgetItem(status))
        self.queue_table.setItem(row, 4, QTableWidgetItem(str(len(text)) if text else "-"))
        self._preview_texts[row] = text

    def _show_selected_preview(self) -> None:
        rows = self.queue_table.selectionModel().selectedRows()
        self.preview_edit.setPlainText(self._preview_texts.get(rows[0].row(), "") if rows else "")

    def set_busy(self, busy: bool) -> None:
        for widget in (
            self.source_input,
            self.start_button,
            self.duration_combo,
            self.asr_fallback_check,
            self.skip_caption_probe_check,
            self.downloaded_button,
        ):
            widget.setEnabled(not busy)
        if not busy:
            self.downloaded_button.setEnabled("(0)" not in self.downloaded_button.text())
        self.pause_button.setEnabled(busy)
        self.cancel_button.setEnabled(busy)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)


class MatchingPage(QWidget):
    """Independent matching surface for subtitle assets and service results."""

    def __init__(
        self,
        *,
        import_ready_items: Callable[[], list[dict[str, object]]],
        import_files: Callable[[], list[dict[str, object]]],
        start_matching: Callable[[list[dict[str, object]], int], None],
        pause_matching: Callable[[], None],
        cancel_matching: Callable[[], None],
        load_evidence_context: Callable[[str], None],
        export_results: Callable[[], None],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._import_ready_items = import_ready_items
        self._import_files = import_files
        self._start_matching = start_matching
        self._load_evidence_context = load_evidence_context
        self._export_results = export_results
        self._items: list[dict[str, object]] = []
        self._evidence_context_cache: dict[str, str] = {}
        self._selected_evidence_uid = ""
        self._selected_evidence_pairs: list[dict[str, object]] = []
        self._selected_result_row: dict[str, object] = {}
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("匹配核验", "把字幕结果单独提交到中央匹配服务，不需要重新下载或转写视频。"))

        source_card = QFrame()
        source_card.setObjectName("WorkspaceCard")
        source_layout = QHBoxLayout(source_card)
        source_layout.setContentsMargins(16, 14, 16, 14)
        for label, callback in (
            ("导入字幕结果", self._load_ready),
            ("导入 TXT / Excel", self._load_files),
        ):
            button = QPushButton(label)
            button.setProperty("secondary", True)
            button.clicked.connect(callback)
            source_layout.addWidget(button)
        source_layout.addStretch(1)
        self.item_count_label = QLabel("当前 0 条")
        source_layout.addWidget(self.item_count_label)
        root.addWidget(source_card)

        self.top_k_spin = QSpinBox()
        self.top_k_spin.setRange(1, 20)
        self.top_k_spin.setValue(10)
        source_layout.addWidget(QLabel("候选数"))
        source_layout.addWidget(self.top_k_spin)

        content_splitter = QSplitter(Qt.Orientation.Vertical)
        list_card = QFrame()
        list_card.setObjectName("WorkspaceCard")
        list_layout = QVBoxLayout(list_card)
        list_layout.setContentsMargins(16, 14, 16, 14)
        list_layout.addWidget(QLabel("待匹配字幕"))
        self.item_table = QTableWidget(0, 4)
        self.item_table.setHorizontalHeaderLabels(["视频/来源", "语言", "字数", "状态"])
        self.item_table.verticalHeader().setVisible(False)
        self.item_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.item_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.item_table.setMinimumHeight(220)
        list_layout.addWidget(self.item_table, 1)
        content_splitter.addWidget(list_card)
        result_card = QFrame()
        result_card.setObjectName("WorkspaceCard")
        result_layout = QVBoxLayout(result_card)
        result_layout.setContentsMargins(16, 14, 16, 14)
        result_layout.addWidget(QLabel("匹配结果"))
        self.result_edit = QPlainTextEdit()
        self.result_edit.setReadOnly(True)
        self.result_edit.setPlaceholderText("提交后显示确认命中、待人工复核和未命中结果")
        result_layout.addWidget(self.result_edit, 1)
        self.result_edit.hide()
        self.result_table = QTableWidget(0, 7)
        self.result_table.setHorizontalHeaderLabels([
            "来源频道", "来源剧名", "原视频链接", "字幕预览", "匹配状态", "命中剧名", "Book ID",
        ])
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        result_header = self.result_table.horizontalHeader()
        for column in range(7):
            result_header.setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)
        result_header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        result_header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.result_table.itemSelectionChanged.connect(self._show_selected_result_detail)
        result_layout.addWidget(self.result_table, 3)
        result_layout.addWidget(QLabel("完整字幕与匹配详情"))
        self.result_detail = QPlainTextEdit()
        self.result_detail.setReadOnly(True)
        self.result_detail.setPlaceholderText("选择一条记录后，显示完整字幕、命中片段和匹配原因。")
        self.result_detail.setMinimumHeight(145)
        result_layout.addWidget(self.result_detail, 2)
        self.result_detail.hide()
        self.comparison_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.comparison_splitter.setObjectName("EvidenceComparison")
        # Keep a quiet visual gutter between the two evidence panels. The
        # comparison is read side-by-side, not resized as a primary action.
        self.comparison_splitter.setHandleWidth(14)
        source_card = QFrame()
        source_layout = QVBoxLayout(source_card)
        source_layout.setContentsMargins(0, 0, 0, 0)
        source_layout.addWidget(QLabel("原字幕（参与匹配的片段）"))
        self.source_context_hint = QLabel("原片段由当前匹配结果自动定位")
        self.source_context_hint.setObjectName("SectionHint")
        self.source_context_hint.setMinimumHeight(20)
        source_layout.addWidget(self.source_context_hint)
        self.source_evidence_edit = QPlainTextEdit()
        self.source_evidence_edit.setReadOnly(True)
        self.source_evidence_edit.setPlaceholderText("选择已命中或待复核的结果后显示原字幕片段。")
        source_layout.addWidget(self.source_evidence_edit, 1)
        matched_card = QFrame()
        matched_layout = QVBoxLayout(matched_card)
        matched_layout.setContentsMargins(0, 0, 0, 0)
        matched_layout.addWidget(QLabel("匹配库字幕证据"))
        self.evidence_selector = QComboBox()
        self.evidence_selector.setMinimumWidth(180)
        self.evidence_selector.setMaximumWidth(360)
        self.evidence_selector.setToolTip("歧义命中时选择不同强证据候选")
        self.evidence_selector.currentIndexChanged.connect(self._show_selected_evidence_pair)
        self.load_context_button = QPushButton("加载扩展上下文")
        self.load_context_button.setProperty("secondary", True)
        self.load_context_button.setEnabled(False)
        self.load_context_button.clicked.connect(self._request_selected_evidence_context)
        matched_layout.addWidget(self.load_context_button)
        self.matched_evidence_edit = QPlainTextEdit()
        self.matched_evidence_edit.setReadOnly(True)
        self.matched_evidence_edit.setPlaceholderText("选择已命中或待复核的结果后显示库内字幕证据。")
        matched_layout.addWidget(self.matched_evidence_edit, 1)

        # Rebuild both headers as one aligned row. The context action stays
        # compact instead of taking an entire line above the right text panel.
        source_title = source_layout.itemAt(0).widget()
        source_layout.removeWidget(source_title)
        source_layout.removeWidget(self.source_context_hint)
        self.source_context_hint.hide()
        source_header = QWidget()
        source_header.setFixedHeight(30)
        source_header_layout = QHBoxLayout(source_header)
        source_header_layout.setContentsMargins(0, 0, 0, 0)
        source_header_layout.addWidget(source_title)
        source_header_layout.addStretch(1)
        self.highlight_hint = QLabel("")
        self.highlight_hint.setObjectName("SectionHint")
        self.highlight_hint.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        source_header_layout.addWidget(self.highlight_hint)
        source_layout.insertWidget(0, source_header)

        matched_title = matched_layout.itemAt(0).widget()
        matched_layout.removeWidget(matched_title)
        matched_layout.removeWidget(self.load_context_button)
        matched_header = QWidget()
        matched_header.setFixedHeight(30)
        matched_header_layout = QHBoxLayout(matched_header)
        matched_header_layout.setContentsMargins(0, 0, 0, 0)
        matched_header_layout.addWidget(matched_title)
        matched_header_layout.addWidget(self.evidence_selector, 1)
        matched_header_layout.addStretch(1)
        self.load_context_button.setFixedSize(126, 24)
        matched_header_layout.addWidget(self.load_context_button)
        matched_layout.insertWidget(0, matched_header)

        self.comparison_splitter.addWidget(source_card)
        self.comparison_splitter.addWidget(matched_card)
        self.comparison_splitter.setStretchFactor(0, 1)
        self.comparison_splitter.setStretchFactor(1, 1)
        result_layout.addWidget(self.comparison_splitter, 2)
        content_splitter.addWidget(result_card)
        content_splitter.setStretchFactor(0, 2)
        content_splitter.setStretchFactor(1, 5)
        content_splitter.setSizes([310, 720])
        root.addWidget(content_splitter, 1)

        footer = QHBoxLayout()
        self.status_label = QLabel("请导入字幕结果后提交匹配。")
        self.status_label.setObjectName("WorkspaceStatus")
        self.submit_button = QPushButton("提交匹配")
        self.submit_button.setProperty("primary", True)
        self.submit_button.clicked.connect(self._request_start)
        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(pause_matching)
        self.cancel_button = QPushButton("取消")
        self.cancel_button.setProperty("danger", True)
        self.cancel_button.clicked.connect(cancel_matching)
        self.export_button = QPushButton("导出匹配 Excel")
        self.export_button.setProperty("secondary", True)
        self.export_button.clicked.connect(self._export_results)
        footer.addWidget(self.status_label, 1)
        footer.addWidget(self.pause_button)
        footer.addWidget(self.cancel_button)
        footer.addWidget(self.export_button)
        footer.addWidget(self.submit_button)
        root.addLayout(footer)
        self.set_busy(False)

    def _load_ready(self) -> None:
        self.set_items(self._import_ready_items())

    def _load_files(self) -> None:
        self.set_items(self._import_files())

    def _request_start(self) -> None:
        if not self._items:
            self.set_status("请先导入至少一条字幕。")
            return
        self._start_matching(
            list(self._items),
            self.top_k_spin.value(),
        )

    def set_items(self, items: list[dict[str, object]]) -> None:
        self._items = [item for item in items if str(item.get("query_text") or "").strip()]
        display_items: dict[str, dict[str, object]] = {}
        for index, item in enumerate(self._items, start=1):
            source_key = str(item.get("source_video_id") or item.get("source_ref") or f"item-{index}").strip()
            source_key = source_key.split("#segment-", 1)[0] or f"item-{index}"
            current = display_items.setdefault(
                source_key,
                {
                    "title": str(item.get("source_display_title") or item.get("source_ref") or ""),
                    "language": str(item.get("source_caption_language") or "自动"),
                    "char_count": 0,
                },
            )
            current["char_count"] = int(current["char_count"]) + len(str(item.get("query_text") or ""))

        self.item_count_label.setText(f"当前 {len(display_items)} 条视频")
        self.item_table.setRowCount(len(display_items))
        for row, item in enumerate(display_items.values()):
            self.item_table.setItem(row, 0, QTableWidgetItem(str(item["title"])))
            self.item_table.setItem(row, 1, QTableWidgetItem(str(item["language"])))
            self.item_table.setItem(row, 2, QTableWidgetItem(str(item["char_count"])))
            self.item_table.setItem(row, 3, QTableWidgetItem("待匹配"))

    def update_status(self, text: str) -> None:
        self.set_status(text)

    def set_result_rows(self, rows: list[dict[str, object]]) -> None:
        self._result_rows = [dict(row) for row in rows]
        self.result_table.setRowCount(len(self._result_rows))
        columns = (
            "source_channel",
            "source_title",
            "source_url",
            "source_subtitle",
            "match_status",
            "matched_book_names",
            "matched_book_ids",
        )
        status_labels = {
            "confirmed_match": "确认命中",
            "content_matched_ambiguous": "内容已命中，剧名待确认",
            "potential_match": "疑似命中，待复核",
            "translation_assisted_match": "翻译辅助待复核",
            "no_match": "未命中",
            # Historical rows can still be loaded from an older local state.
            "matched": "确认命中",
            "review_required": "待人工复核",
            "not_matched": "未命中",
        }
        for row_index, row in enumerate(self._result_rows):
            for column_index, field in enumerate(columns):
                value = str(row.get(field) or "")
                display = value
                if field == "source_url" and len(value) > 70:
                    display = f"{value[:67]}..."
                elif field == "source_subtitle" and len(value) > 110:
                    display = f"{value[:110]}..."
                if field == "match_status":
                    display = status_labels.get(value, value)
                item = QTableWidgetItem(display)
                item.setToolTip(value)
                self.result_table.setItem(row_index, column_index, item)
        if self._result_rows:
            self.result_table.selectRow(0)
        else:
            self.result_detail.clear()
            self.source_evidence_edit.clear()
            self.matched_evidence_edit.clear()
            self.highlight_hint.clear()
            self.evidence_selector.clear()
            self.evidence_selector.hide()

    @staticmethod
    def _common_phrase_terms(source_text: str, matched_text: str, provided_terms: list[str]) -> list[str]:
        """Keep server terms, then derive visible common phrases as a local fallback."""
        terms: list[str] = []
        seen: set[str] = set()

        def append_term(value: object) -> None:
            term = str(value or "").strip()
            normalized = term.casefold()
            if len(term) >= 3 and normalized not in seen:
                seen.add(normalized)
                terms.append(term)

        for term in provided_terms:
            append_term(term)

        # The service may omit shared_trigrams for a valid semantic match. Exact
        # common blocks still give the reviewer a visible, auditable comparison.
        matcher = SequenceMatcher(None, source_text, matched_text, autojunk=False)
        for block in matcher.get_matching_blocks():
            if block.size >= 4:
                append_term(source_text[block.a : block.a + min(block.size, 80)])
        return terms[:120]

    @staticmethod
    def _apply_match_highlight(editor: QPlainTextEdit, terms: list[str]) -> int:
        text = editor.toPlainText()
        selections: list[QTextEdit.ExtraSelection] = []
        seen_terms: set[str] = set()
        for raw_term in terms:
            term = str(raw_term or "").strip()
            normalized = term.casefold()
            if len(term) < 3 or normalized in seen_terms:
                continue
            seen_terms.add(normalized)
            start = 0
            while len(selections) < 120:
                index = text.casefold().find(normalized, start)
                if index < 0:
                    break
                cursor = QTextCursor(editor.document())
                cursor.setPosition(index)
                cursor.setPosition(index + len(term), QTextCursor.MoveMode.KeepAnchor)
                selection = QTextEdit.ExtraSelection()
                selection.cursor = cursor
                selection.format.setBackground(QColor("#FFE28A"))
                selection.format.setForeground(QColor("#4B3312"))
                selection.format.setFontWeight(700)
                selection.format.setUnderlineStyle(QTextCharFormat.UnderlineStyle.SingleUnderline)
                selection.format.setUnderlineColor(QColor("#D69200"))
                selections.append(selection)
                start = index + len(term)
        editor.setExtraSelections(selections)
        return len(selections)

    def _show_selected_result_detail(self) -> None:
        row_index = self.result_table.currentRow()
        if row_index < 0 or row_index >= len(getattr(self, "_result_rows", [])):
            return
        row = self._result_rows[row_index]
        status = {
            "confirmed_match": "确认命中",
            "content_matched_ambiguous": "内容已命中，剧名待确认",
            "potential_match": "疑似命中，待复核",
            "translation_assisted_match": "翻译辅助待复核",
            "no_match": "未命中",
            "matched": "确认命中",
            "review_required": "待人工复核",
            "not_matched": "未命中",
        }.get(
            str(row.get("match_status") or ""),
            str(row.get("match_status") or ""),
        )
        evidence_pairs = row.get("evidence_pairs")
        self._selected_result_row = row
        self._selected_evidence_pairs = [
            value for value in evidence_pairs
            if isinstance(value, dict)
        ] if isinstance(evidence_pairs, list) else []
        self.evidence_selector.blockSignals(True)
        self.evidence_selector.clear()
        for index, pair in enumerate(self._selected_evidence_pairs, start=1):
            priority = pair.get("review_priority") or index
            name = str(pair.get("book_name") or "未命名候选")
            episode = str(pair.get("episode_order") or "未知集数")
            coverage = str(pair.get("coverage_text") or "历史任务未计算")
            self.evidence_selector.addItem(f"优先级 {priority} | {name} | 第{episode}集 | 覆盖 {coverage}", index - 1)
        self.evidence_selector.blockSignals(False)
        self.evidence_selector.setVisible(len(self._selected_evidence_pairs) > 1)
        self.evidence_selector.setCurrentIndex(0)
        if not self._selected_evidence_pairs:
            self._selected_evidence_uid = ""
            self.load_context_button.setEnabled(False)
            self.load_context_button.setText("无可扩展证据")
            self._render_evidence_pair({}, row, status)
            return
        self._show_selected_evidence_pair()

    def _show_selected_evidence_pair(self) -> None:
        if not self._selected_result_row:
            return
        index = self.evidence_selector.currentData()
        if not isinstance(index, int):
            index = self.evidence_selector.currentIndex()
        pair = self._selected_evidence_pairs[index] if 0 <= index < len(self._selected_evidence_pairs) else {}
        status = {
            "confirmed_match": "确认命中",
            "content_matched_ambiguous": "内容已命中，剧名待确认",
            "potential_match": "疑似命中，待复核",
            "translation_assisted_match": "翻译辅助待复核",
            "no_match": "未命中",
            "matched": "确认命中",
            "review_required": "待人工复核",
            "not_matched": "未命中",
        }.get(str(self._selected_result_row.get("match_status") or ""), str(self._selected_result_row.get("match_status") or ""))
        self._render_evidence_pair(pair, self._selected_result_row, status)

    def _render_evidence_pair(self, pair: dict[str, object], row: dict[str, object], status: str) -> None:
        source_text = str(pair.get("source_text") or row.get("source_subtitle") or "暂无可用字幕").strip()
        execution = pair.get("execution") if isinstance(pair.get("execution"), dict) else row.get("execution")
        window_uid = str(pair.get("window_uid") or "").strip()
        self._selected_evidence_uid = window_uid
        self.load_context_button.setEnabled(bool(window_uid))
        evidence_context = str(self._evidence_context_cache.get(window_uid) or "").strip()
        matched_text = evidence_context or str(pair.get("matched_text") or "暂无匹配库字幕证据").strip()
        if window_uid:
            self.load_context_button.setText("已加载扩展上下文" if evidence_context else "加载扩展上下文")
        else:
            self.load_context_button.setText("无可扩展证据")
        self.source_evidence_edit.setPlainText(
            "\n".join(
                (
                    f"来源频道：{row.get('source_channel') or '未提供'}",
                    f"来源剧名：{row.get('source_title') or '未提供'}",
                    f"原视频链接：{row.get('source_url') or '未提供'}",
                    "来源字幕：完整字幕或服务端定位的命中内容",
                    "",
                    source_text,
                )
            )
        )
        translation = row.get("translation_fallback") if isinstance(row.get("translation_fallback"), dict) else {}
        translated_text = str(row.get("translated_query_text") or "").strip()
        translation_lines = []
        if translation:
            translation_lines.append(f"翻译回退：{translation.get('status') or '未提供'}")
            translation_lines.append(f"检索目标语言：{translation.get('matched_target_language_code') or '未提供'}")
        if translated_text:
            translation_lines.extend(("翻译字幕（仅供复核，不替代原字幕）：", translated_text))
        self.matched_evidence_edit.setPlainText(
            "\n".join(
                (
                    f"匹配状态：{status or '未提供'}",
                    f"命中剧名：{pair.get('book_name') or row.get('matched_book_names') or '无'}",
                    f"Book ID：{pair.get('book_id') or row.get('matched_book_ids') or '无'}",
                    f"命中集数：{pair.get('episode_order') or row.get('matched_episode_orders') or '无'}",
                    f"用户结论：{row.get('user_message') or pair.get('reason') or row.get('match_reasons') or '无'}",
                    f"累计覆盖率：{pair.get('coverage_text') or '历史任务未计算'}",
                    f"单窗口覆盖率：{pair.get('single_window_coverage_text') or '历史任务未计算'}",
                    f"命中窗口数：{pair.get('matched_window_count') or '无'}",
                    f"语义相似度：{pair.get('semantic_score') if pair.get('semantic_score') is not None else '无'}",
                    f"服务端执行：{self._execution_summary(execution)}",
                    *translation_lines,
                    "",
                    matched_text,
                )
            )
        )
        provided_terms = pair.get("highlight_terms") if isinstance(pair.get("highlight_terms"), list) else []
        highlight_terms = self._common_phrase_terms(source_text, matched_text, provided_terms)
        source_hits = self._apply_match_highlight(self.source_evidence_edit, highlight_terms)
        matched_hits = self._apply_match_highlight(self.matched_evidence_edit, highlight_terms)
        if source_hits and matched_hits:
            self.highlight_hint.setText(f"黄色标注 = 重合字幕（{min(source_hits, matched_hits)} 处）")
        elif window_uid:
            self.highlight_hint.setText("加载扩展上下文后显示重合标注")
        else:
            self.highlight_hint.setText("未提供可定位的匹配证据")

        strong_candidates = row.get("strong_candidates")
        if isinstance(strong_candidates, list) and len(strong_candidates) > 1:
            candidate_lines = ["", "强证据候选（按复核优先级）："]
            for candidate in strong_candidates:
                if not isinstance(candidate, dict):
                    continue
                candidate_lines.append(
                    "优先级 {priority}：{name} | Book ID {book_id} | 第{episode}集 | 累计覆盖 {coverage} | 单窗口 {single} | 窗口 {windows} | 语义 {semantic}".format(
                        priority=candidate.get("review_priority") or "-",
                        name=candidate.get("book_name") or "未命名",
                        book_id=candidate.get("book_id") or "-",
                        episode=candidate.get("episode_order") or "-",
                        coverage=self._format_rate(candidate.get("aggregate_text_coverage_rate")),
                        single=self._format_rate(candidate.get("text_coverage_rate")),
                        windows=candidate.get("matched_window_count") or "-",
                        semantic=candidate.get("semantic_score") if candidate.get("semantic_score") is not None else "-",
                    )
                )
            current = self.matched_evidence_edit.toPlainText()
            self.matched_evidence_edit.setPlainText(current + "\n" + "\n".join(candidate_lines))
            self._apply_match_highlight(self.matched_evidence_edit, highlight_terms)

    @staticmethod
    def _format_rate(value: object) -> str:
        return f"{float(value) * 100:.2f}%" if isinstance(value, (int, float)) else "历史任务未计算"

    @staticmethod
    def _execution_summary(value: object) -> str:
        execution = value if isinstance(value, dict) else {}
        if not execution:
            return "历史任务未提供"
        strategy = str(execution.get("strategy") or "")
        if strategy == "fast_screen_only":
            return "完整字幕初筛直接得出结论，未执行补检"
        if strategy == "fast_screen_then_segment_fallback":
            processed = execution.get("processed_segment_count") or 0
            total = execution.get("segment_count") or 0
            early = "，命中后提前停止" if execution.get("stopped_early") else ""
            return f"完整初筛后补检 {processed}/{total} 段{early}"
        return "服务端已执行视频级匹配"

    def _request_selected_evidence_context(self) -> None:
        if self._selected_evidence_uid:
            self.load_context_button.setEnabled(False)
            self.load_context_button.setText("加载中...")
            self.set_status("正在加载匹配库扩展字幕，完成后会替换右侧的短证据片段。")
            self._load_evidence_context(self._selected_evidence_uid)

    def set_evidence_context(self, window_uid: str, text: str) -> None:
        uid = str(window_uid or "").strip()
        if not uid:
            return
        self._evidence_context_cache[uid] = str(text or "").strip()
        self.load_context_button.setText("已加载扩展上下文")
        self.load_context_button.setEnabled(bool(self._selected_evidence_uid))
        if uid == self._selected_evidence_uid:
            self._show_selected_evidence_pair()

    def set_evidence_context_error(self) -> None:
        self.load_context_button.setText("重试加载上下文")
        self.load_context_button.setEnabled(bool(self._selected_evidence_uid))

    def set_busy(self, busy: bool, *, submission_blocked: bool = False) -> None:
        self.top_k_spin.setEnabled(not busy and not submission_blocked)
        self.submit_button.setEnabled(not busy and not submission_blocked)
        self.export_button.setEnabled(not busy)
        self.pause_button.setEnabled(busy)
        self.cancel_button.setEnabled(busy)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)


class CoverPage(QWidget):
    """Independent cover-material page: acquire thumbnails, then optionally review them."""

    def __init__(
        self,
        *,
        import_videos: Callable[[], list[dict[str, object]]],
        collect_channels: Callable[[list[str], int], None],
        start_review: Callable[[list[dict[str, object]]], None],
        start_url_review: Callable[[list[dict[str, object]]], None],
        remove_items: Callable[[set[str]], None],
        start_cover: Callable[[list[dict[str, object]], bool, bool], None],
        pause_cover: Callable[[], None],
        cancel_cover: Callable[[], None],
        export_reviews: Callable[[], None],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._import_videos = import_videos
        self._collect_channels = collect_channels
        self._start_review = start_review
        self._start_url_review = start_url_review
        self._remove_items_callback = remove_items
        self._start_cover = start_cover
        self._export_reviews = export_reviews
        self._items: list[dict[str, object]] = []
        self._review_details: dict[int, dict[str, object]] = {}
        self._cover_preview_pixmap: QPixmap | None = None
        self._selection_buttons: list[QPushButton] = []
        self._channel_urls: list[str] = []
        self._visible_indices: list[int] = []
        self._queue_tab_index = 0
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("封面素材", "封面检测包含两个步骤：先获取封面文件，再调用模型检测；也可以只下载封面。"))

        source_card = QFrame()
        source_card.setObjectName("WorkspaceCard")
        source_layout = QVBoxLayout(source_card)
        source_layout.setContentsMargins(16, 14, 16, 14)
        input_row = QHBoxLayout()
        self.manage_channels_button = QPushButton("管理频道清单")
        self.manage_channels_button.setProperty("secondary", True)
        self.manage_channels_button.clicked.connect(self._open_channel_dialog)
        input_row.addWidget(self.manage_channels_button)
        self.channel_count_label = QLabel("尚未添加频道")
        self.channel_count_label.setObjectName("PageHint")
        input_row.addWidget(self.channel_count_label, 1)
        input_row.addWidget(QLabel("每频道最多"))
        self.channel_limit_spin = QSpinBox()
        self.channel_limit_spin.setRange(0, 10000)
        self.channel_limit_spin.setValue(0)
        self.channel_limit_spin.setSpecialValueText("全部")
        self.channel_limit_spin.setToolTip("每个频道最多采集多少条视频；0 表示全部")
        input_row.addWidget(self.channel_limit_spin)
        self.collect_button = QPushButton("采集频道视频")
        self.collect_button.setProperty("secondary", True)
        self.collect_button.clicked.connect(self._collect_channel_videos)
        input_row.addWidget(self.collect_button)
        source_layout.addLayout(input_row)
        action_row = QHBoxLayout()
        import_button = QPushButton("从采集结果导入")
        import_button.setProperty("secondary", True)
        import_button.clicked.connect(self._load_videos)
        action_row.addWidget(import_button)
        self.download_check = QCheckBox("自动获取封面")
        self.download_check.setChecked(True)
        self.detect_check = QCheckBox("获取后检测封面")
        self.detect_check.setChecked(True)
        action_row.addStretch(1)
        self.item_count_label = QLabel("当前 0 条")
        action_row.addWidget(self.item_count_label)
        source_layout.addLayout(action_row)
        root.addWidget(source_card)

        queue_card = QFrame()
        queue_card.setObjectName("WorkspaceCard")
        queue_layout = QVBoxLayout(queue_card)
        queue_layout.setContentsMargins(16, 14, 16, 14)
        queue_layout.addWidget(QLabel("封面处理列表"))
        self.queue_tabs = QTabWidget()
        self.queue_tabs.addTab(QWidget(), "采集结果 (0)")
        self.queue_tabs.addTab(QWidget(), "已下载封面 (0)")
        self.queue_tabs.currentChanged.connect(self._on_queue_tab_changed)
        queue_layout.addWidget(self.queue_tabs)
        selection_bar = QHBoxLayout()
        for label, callback in (("全选", self._select_all), ("反选", self._invert_selection), ("清除选择", self._clear_selection), ("移除选中", self._remove_selected)):
            button = QPushButton(label)
            button.setProperty("secondary", True)
            button.clicked.connect(callback)
            selection_bar.addWidget(button)
            self._selection_buttons.append(button)
        selection_bar.addWidget(QLabel("前 N 条"))
        self.select_count_spin = QSpinBox()
        self.select_count_spin.setRange(1, 10000)
        self.select_count_spin.setValue(10)
        selection_bar.addWidget(self.select_count_spin)
        select_first_button = QPushButton("选择前 N 条")
        select_first_button.setProperty("secondary", True)
        select_first_button.clicked.connect(self._select_first_n)
        selection_bar.addWidget(select_first_button)
        selection_bar.addWidget(QLabel("按频道"))
        self.channel_selection_combo = QComboBox()
        self.channel_selection_combo.addItem("选择频道", "")
        self.channel_selection_combo.setMinimumWidth(180)
        selection_bar.addWidget(self.channel_selection_combo, 1)
        select_channel_button = QPushButton("选择该频道")
        select_channel_button.setProperty("secondary", True)
        select_channel_button.clicked.connect(self._select_channel)
        selection_bar.addWidget(select_channel_button)
        queue_layout.addLayout(selection_bar)
        self.selected_count_label = QLabel("已选 0 条")
        queue_layout.addWidget(self.selected_count_label)
        self.item_table = QTableWidget(0, 9)
        self.item_table.setHorizontalHeaderLabels(["选择", "标题", "视频 ID", "频道", "采集状态", "封面状态", "检测状态", "封面文件", "检测结论"])
        self.item_table.verticalHeader().setVisible(False)
        self.item_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.item_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.item_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.item_table.itemSelectionChanged.connect(self._show_selected_review_detail)
        self.item_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.item_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        self.item_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self.item_table.itemChanged.connect(self._on_item_changed)
        review_detail = QFrame()
        review_detail.setObjectName("WorkspaceInset")
        review_detail_layout = QVBoxLayout(review_detail)
        review_detail_layout.setContentsMargins(12, 10, 12, 10)
        review_detail_layout.setSpacing(6)
        review_detail_layout.addWidget(QLabel("检测详情（选中列表中的一条封面查看）"))
        detail_content = QHBoxLayout()
        detail_content.setSpacing(12)
        self.cover_preview_label = QLabel("选择一条封面后显示预览")
        self.cover_preview_label.setObjectName("CoverPreview")
        self.cover_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.cover_preview_label.setMinimumSize(240, 150)
        self.cover_preview_label.setMaximumWidth(360)
        detail_content.addWidget(self.cover_preview_label, 1)
        self.review_detail_edit = QPlainTextEdit()
        self.review_detail_edit.setReadOnly(True)
        self.review_detail_edit.setMinimumHeight(130)
        self.review_detail_edit.setPlaceholderText("模型完成检测后，会显示结论、可见证据与原始模型回复。")
        detail_content.addWidget(self.review_detail_edit, 2)
        review_detail_layout.addLayout(detail_content)
        content_splitter = QSplitter(Qt.Orientation.Vertical)
        content_splitter.addWidget(self.item_table)
        content_splitter.addWidget(review_detail)
        content_splitter.setStretchFactor(0, 3)
        content_splitter.setStretchFactor(1, 2)
        content_splitter.setSizes([360, 190])
        queue_layout.addWidget(content_splitter, 1)
        root.addWidget(queue_card, 1)

        footer = QHBoxLayout()
        self.status_label = QLabel("请从采集结果导入视频。")
        self.status_label.setObjectName("WorkspaceStatus")
        self.download_button = QPushButton("下载选中封面")
        self.download_button.setProperty("secondary", True)
        self.download_button.clicked.connect(self._request_download)
        self.review_button = QPushButton("检测选中封面")
        self.review_button.setProperty("secondary", True)
        self.review_button.setToolTip("只检测当前勾选且已经下载的封面")
        self.review_button.clicked.connect(self._request_review)
        self.url_review_button = QPushButton("URL优先检测选中封面")
        self.url_review_button.setProperty("secondary", True)
        self.url_review_button.setToolTip("先把封面 CDN 地址交给模型，失败后自动下载本地图片再检测")
        self.url_review_button.clicked.connect(self._request_url_review)
        # Keep the experimental remote-URL path available in code, but hide it
        # from the production UI until its latency is acceptable.
        self.url_review_button.setVisible(False)
        self.start_button = QPushButton("下载封面并检测")
        self.start_button.setProperty("primary", True)
        self.start_button.clicked.connect(self._request_start)
        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(pause_cover)
        self.cancel_button = QPushButton("取消")
        self.cancel_button.setProperty("danger", True)
        self.cancel_button.clicked.connect(cancel_cover)
        self.export_button = QPushButton("导出封面检测 Excel")
        self.export_button.setProperty("secondary", True)
        self.export_button.clicked.connect(self._export_reviews)
        footer.addWidget(self.status_label, 1)
        footer.addWidget(self.pause_button)
        footer.addWidget(self.cancel_button)
        footer.addWidget(self.export_button)
        footer.addWidget(self.download_button)
        footer.addWidget(self.review_button)
        footer.addWidget(self.url_review_button)
        footer.addWidget(self.start_button)
        root.addLayout(footer)
        self.set_busy(False)

    def _load_videos(self) -> None:
        self.set_items(self._import_videos())

    @property
    def items(self) -> list[dict[str, object]]:
        return [dict(item) for item in self._items]

    def _collect_channel_videos(self) -> None:
        if not self._channel_urls:
            self.set_status("请先点击“管理频道清单”填写或导入频道链接。")
            return
        self._collect_channels(list(self._channel_urls), self.channel_limit_spin.value())

    def _open_channel_dialog(self) -> None:
        dialog = ChannelListDialog(self._channel_urls, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        self._channel_urls = dialog.urls()
        self.channel_count_label.setText(f"已添加 {len(self._channel_urls)} 个频道")
        self.set_status(f"频道清单已更新：{len(self._channel_urls)} 个频道，点击“采集频道视频”开始。")

    def _request_start(self) -> None:
        if not self._items:
            self.set_status("请先从采集结果导入视频。")
            return
        selected = self._selected_items()
        if not selected:
            self.set_status("请先勾选要处理的视频，可使用全选、前 N 条或按频道选择。")
            return
        self._start_cover(selected, True, True)

    def _request_download(self) -> None:
        selected = self._selected_items()
        if not selected:
            self.set_status("请先勾选要下载封面的视频。")
            return
        self.set_status(f"已选择 {len(selected)} 条，准备下载封面。")
        self._start_cover(selected, True, False)

    def _request_review(self) -> None:
        selected = self._selected_items()
        selected = [item for item in selected if str(item.get("cover_path") or "").strip()]
        if not selected:
            self.set_status("请先勾选要检测的封面；未下载封面的项目不会进入检测。")
            return
        self.set_status(f"已选择 {len(selected)} 条，准备检测选中封面。")
        self._start_review(selected)

    def _request_url_review(self) -> None:
        selected = [item for item in self._selected_items() if isinstance(item.get("video"), dict) and str((item.get("video") or {}).get("thumbnail_url") or "").strip()]
        if not selected:
            self.set_status("请先勾选带封面 CDN 地址的视频。")
            return
        self.set_status(f"已选择 {len(selected)} 条，准备 URL 优先检测（失败自动回退本地下载）。")
        self._start_url_review(selected)

    def set_items(self, items: list[dict[str, object]]) -> None:
        self._items = []
        for raw in items:
            item = dict(raw)
            item.setdefault("collection_status", "success")
            item.setdefault("cover_status", "downloaded" if str(item.get("cover_path") or "") else "pending")
            item.setdefault("review_status", "completed" if item.get("review") else "pending")
            self._items.append(item)
        self._review_details.clear()
        self.review_detail_edit.clear()
        self._set_cover_preview(-1)
        self.item_count_label.setText(f"当前 {len(self._items)} 条")
        self._render_queue_table()
        self._refresh_queue_tabs()
        self._refresh_channel_combo()
        self._update_selected_count()
        if self._items:
            self.set_status(f"已载入 {len(self._items)} 条视频，可勾选后执行封面操作。")

    def _render_queue_table(self) -> None:
        self._visible_indices = [
            index for index, item in enumerate(self._items)
            if (self._queue_tab_index == 0 and not str(item.get("cover_path") or "").strip())
            or (self._queue_tab_index == 1 and str(item.get("cover_path") or "").strip())
        ]
        self.item_table.blockSignals(True)
        self.item_table.setRowCount(len(self._visible_indices))
        for row, source_index in enumerate(self._visible_indices):
            item = self._items[source_index]
            video = item.get("video") if isinstance(item.get("video"), dict) else item
            video = video if isinstance(video, dict) else {}
            cover_path = str(item.get("cover_path") or "")
            review = item.get("review") if isinstance(item.get("review"), dict) else {}
            checkbox = QTableWidgetItem()
            checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            checkbox.setCheckState(Qt.CheckState.Unchecked)
            self.item_table.setItem(row, 0, checkbox)
            self.item_table.setItem(row, 1, QTableWidgetItem(str(video.get("title") or "")))
            self.item_table.setItem(row, 2, QTableWidgetItem(str(video.get("video_id") or "")))
            self.item_table.setItem(row, 3, QTableWidgetItem(str(video.get("channel") or "")))
            self.item_table.setItem(row, 4, self._status_item("collection", item.get("collection_status") or "success"))
            self.item_table.setItem(row, 5, self._status_item("cover", item.get("cover_status") or "pending"))
            self.item_table.setItem(row, 6, self._status_item("review", item.get("review_status") or "pending"))
            self.item_table.setItem(row, 7, QTableWidgetItem(cover_path or "待获取"))
            self.item_table.setItem(row, 8, QTableWidgetItem(""))
            if review:
                self.update_job(row, "检测失败" if review.get("error") else "处理完成", cover_path, review)
            elif cover_path:
                self.item_table.setItem(row, 5, self._status_item("cover", "downloaded"))
        self.item_table.blockSignals(False)

    def _refresh_queue_tabs(self) -> None:
        downloaded = sum(1 for item in self._items if str(item.get("cover_path") or "").strip())
        self.queue_tabs.setTabText(0, f"采集结果 ({len(self._items) - downloaded})")
        self.queue_tabs.setTabText(1, f"已下载封面 ({downloaded})")

    def _on_queue_tab_changed(self, index: int) -> None:
        self._queue_tab_index = max(0, min(index, 1))
        self._render_queue_table()
        self._refresh_channel_combo()
        self._update_selected_count()

    def append_items(self, items: list[dict[str, object]]) -> None:
        """Append newly collected rows while preserving the same queue schema."""
        if not items:
            return
        self.set_items(self._items + items)

    def refresh_stage_views(self) -> None:
        """Refresh tab counts and the filtered table after a worker update."""
        self._refresh_queue_tabs()
        if self._queue_tab_index == 1:
            self._render_queue_table()
            self._refresh_channel_combo()
            self._update_selected_count()

    def update_review_by_video_id(self, video_id: str, result: object) -> None:
        target_id = str(video_id or "")
        for source_index, item in enumerate(self._items):
            video = item.get("video") if isinstance(item.get("video"), dict) else item
            if not isinstance(video, dict) or str(video.get("video_id") or "") != target_id:
                continue
            payload = result.to_dict() if hasattr(result, "to_dict") else dict(result) if isinstance(result, dict) else {}
            item["review"] = payload
            item["review_status"] = "检测失败" if payload.get("error") else "检测完成"
            if source_index in self._visible_indices:
                self.update_job(self._visible_indices.index(source_index), "检测失败" if payload.get("error") else "处理完成", str(item.get("cover_path") or ""), payload)
            return

    def update_job(self, row: int, status: str, cover_path: str = "", result: object = "") -> None:
        if row < 0 or row >= len(self._visible_indices):
            return
        source_index = self._visible_indices[row]
        if cover_path:
            self.item_table.setItem(row, 7, QTableWidgetItem(cover_path))
            self._items[source_index]["cover_path"] = cover_path
            self._items[source_index]["cover_status"] = "downloaded"
        status_column = 6 if "检测" in status or "处理" in status else 5
        stage = "review" if status_column == 6 else "cover"
        self.item_table.setItem(row, status_column, self._status_item(stage, status))
        self._items[source_index]["review_status" if status_column == 6 else "cover_status"] = status
        if isinstance(result, dict):
            self._review_details[source_index] = dict(result)
            self._items[source_index]["review"] = dict(result)
            detail = self._review_details[source_index]
            if detail.get("error"):
                display = f"检测失败：{detail.get('error')}"
            else:
                labels = {"safe": "安全", "review": "待复核", "risk": "疑似风险", "unknown": "无法判断"}
                summary = str(detail.get("summary") or detail.get("evidence") or "模型未提供文字说明").strip()
                display = f"{labels.get(str(detail.get('overall_risk') or ''), '无法判断')}｜{summary}"
            item = QTableWidgetItem(display)
            item.setToolTip(display)
            self.item_table.setItem(row, 8, item)
            if self.item_table.currentRow() == row:
                self._show_selected_review_detail()
        elif result:
            self.item_table.setItem(row, 4, QTableWidgetItem(str(result)))

    def _show_selected_review_detail(self) -> None:
        row = self.item_table.currentRow()
        self._set_cover_preview(row)
        source_index = self._visible_indices[row] if 0 <= row < len(self._visible_indices) else -1
        detail = self._review_details.get(source_index)
        if not detail:
            self.review_detail_edit.setPlainText("该封面尚未完成模型检测。")
            return
        if detail.get("error"):
            self.review_detail_edit.setPlainText(f"检测失败：{detail.get('error')}")
            return
        labels = {"safe": "安全", "review": "待复核", "risk": "疑似风险", "unknown": "无法判断"}
        tags = detail.get("risk_tags") if isinstance(detail.get("risk_tags"), (list, tuple)) else []
        tag_text = "、".join(str(tag) for tag in tags if str(tag).strip()) or "未发现明确风险标签"
        try:
            confidence = f"{float(detail.get('confidence') or 0):.0%}"
        except (TypeError, ValueError):
            confidence = "未提供"
        model_response = str(detail.get("model_response") or "").strip() or "未保留原始回复（旧任务结果）"
        self.review_detail_edit.setPlainText(
            "\n".join(
                (
                    f"结论：{labels.get(str(detail.get('overall_risk') or ''), '无法判断')}",
                    f"风险标签：{tag_text}",
                    f"置信度：{confidence}",
                    f"摘要：{str(detail.get('summary') or '模型未提供摘要').strip()}",
                    f"可见证据：{str(detail.get('evidence') or '模型未提供证据').strip()}",
                    "",
                    "模型原始回复：",
                    model_response,
                )
            )
        )

    @staticmethod
    def _status_item(stage: str, raw_status: object) -> QTableWidgetItem:
        value = str(raw_status or "pending").strip()
        labels = {
            "success": "采集完成",
            "collection_failed": "采集失败",
            "pending": "待处理",
            "downloading": "下载中",
            "downloaded": "封面已获取",
            "cover_failed": "下载失败",
            "reviewing": "检测中",
            "completed": "检测完成",
            "review_failed": "检测失败",
            "获取封面": "下载中",
            "封面失败": "下载失败",
            "封面已获取": "封面已获取",
            "检测封面": "检测中",
            "处理完成": "检测完成",
            "检测失败": "检测失败",
            "处理失败": "处理失败",
            "采集完成": "采集完成",
            "待获取": "待处理",
        }
        text = labels.get(value, value)
        item = QTableWidgetItem(text)
        palette = {
            "pending": ("#fff4d6", "#8a5a00"),
            "downloading": ("#e4f0ff", "#175a9e"),
            "downloaded": ("#e6f7ed", "#176b3a"),
            "completed": ("#dff5ec", "#176b3a"),
            "success": ("#e6f7ed", "#176b3a"),
            "collection_failed": ("#ffe8e6", "#a12622"),
            "cover_failed": ("#ffe8e6", "#a12622"),
            "review_failed": ("#ffe8e6", "#a12622"),
            "采集完成": ("#e6f7ed", "#176b3a"),
            "检测完成": ("#dff5ec", "#176b3a"),
            "检测失败": ("#ffe8e6", "#a12622"),
            "处理失败": ("#ffe8e6", "#a12622"),
        }
        background, foreground = palette.get(value, ("#f2f4f7", "#4b5563"))
        item.setBackground(QColor(background))
        item.setForeground(QColor(foreground))
        item.setToolTip(f"{stage}阶段：{text}")
        return item

    def _set_cover_preview(self, row: int) -> None:
        if row < 0 or row >= self.item_table.rowCount():
            self._cover_preview_pixmap = None
            self.cover_preview_label.setPixmap(QPixmap())
            self.cover_preview_label.setText("选择一条封面后显示预览")
            return
        path_item = self.item_table.item(row, 7)
        cover_path = str(path_item.text() or "") if path_item is not None else ""
        pixmap = QPixmap(cover_path) if cover_path and cover_path != "待获取" else QPixmap()
        self._cover_preview_pixmap = pixmap if not pixmap.isNull() else None
        if self._cover_preview_pixmap is None:
            self.cover_preview_label.setPixmap(QPixmap())
            self.cover_preview_label.setText("封面图片加载失败或尚未获取")
            return
        self.cover_preview_label.setText("")
        self._render_cover_preview()

    def _render_cover_preview(self) -> None:
        if self._cover_preview_pixmap is None:
            return
        target_size = self.cover_preview_label.size()
        self.cover_preview_label.setPixmap(
            self._cover_preview_pixmap.scaled(
                target_size,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        )

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        self._render_cover_preview()

    def set_busy(self, busy: bool, *, allow_review: bool = False) -> None:
        for widget in (self.download_button, self.review_button, self.url_review_button, self.start_button, self.download_check, self.detect_check, self.export_button, self.collect_button, self.manage_channels_button, self.channel_limit_spin):
            widget.setEnabled(not busy)
        self.pause_button.setEnabled(busy)
        self.cancel_button.setEnabled(busy)
        if allow_review:
            self.review_button.setEnabled(True)

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() == 0:
            self._update_selected_count()

    def _selected_rows(self) -> list[int]:
        return [
            row for row in range(self.item_table.rowCount())
            if (item := self.item_table.item(row, 0)) is not None
            and item.checkState() == Qt.CheckState.Checked
        ]

    def _selected_items(self) -> list[dict[str, object]]:
        return [self._items[self._visible_indices[row]] for row in self._selected_rows() if row < len(self._visible_indices)]

    def _set_rows_checked(self, rows: set[int], checked: bool = True) -> None:
        self.item_table.blockSignals(True)
        for row in range(self.item_table.rowCount()):
            item = self.item_table.item(row, 0)
            if item is not None:
                item.setCheckState(Qt.CheckState.Checked if (checked and row in rows) else Qt.CheckState.Unchecked)
        self.item_table.blockSignals(False)
        self._update_selected_count()

    def _select_all(self) -> None:
        self._set_rows_checked(set(range(self.item_table.rowCount())))
        self.set_status(f"已全选当前页 {self.item_table.rowCount()} 条。")

    def _invert_selection(self) -> None:
        selected = set(self._selected_rows())
        self.item_table.blockSignals(True)
        for row in range(self.item_table.rowCount()):
            item = self.item_table.item(row, 0)
            if item is not None:
                item.setCheckState(Qt.CheckState.Unchecked if row in selected else Qt.CheckState.Checked)
        self.item_table.blockSignals(False)
        self._update_selected_count()
        self.set_status(f"已反选当前页，当前勾选 {len(self._selected_rows())} 条。")

    def _clear_selection(self) -> None:
        self._set_rows_checked(set(), checked=False)
        self.set_status("已清除当前页选择。")

    def _remove_selected(self) -> None:
        selected = self._selected_items()
        if not selected:
            self.set_status("请先勾选要移除的视频。")
            return
        ids = {
            str((item.get("video") or {}).get("video_id") or "")
            for item in selected
            if isinstance(item.get("video"), dict) and str((item.get("video") or {}).get("video_id") or "")
        }
        if not ids:
            self.set_status("选中项目缺少视频 ID，无法移除。")
            return
        self._items = [
            item for item in self._items
            if str((item.get("video") or {}).get("video_id") or "") not in ids
        ]
        self._render_queue_table()
        self._refresh_queue_tabs()
        self._refresh_channel_combo()
        self._update_selected_count()
        self._remove_items_callback(ids)
        self.set_status(f"已移除 {len(ids)} 条封面队列记录，磁盘封面文件未删除。")

    def _select_first_n(self) -> None:
        count = min(self.select_count_spin.value(), self.item_table.rowCount())
        self._set_rows_checked(set(range(count)))
        self.set_status(f"已选择当前页前 {count} 条。")

    def _select_channel(self) -> None:
        channel = str(self.channel_selection_combo.currentData() or "")
        if not channel:
            return
        rows = {
            row for row, source_index in enumerate(self._visible_indices)
            for item in [self._items[source_index]]
            if isinstance(item.get("video"), dict) and str(item["video"].get("channel") or "") == channel
        }
        self._set_rows_checked(rows)
        if rows:
            first_row = min(rows)
            self.item_table.setCurrentCell(first_row, 0)
            target = self.item_table.item(first_row, 0)
            if target is not None:
                self.item_table.scrollToItem(target, QAbstractItemView.ScrollHint.PositionAtTop)
            self.set_status(f"已选择频道“{channel}”的 {len(rows)} 条，并定位到第一条。")
        else:
            self.set_status(f"当前页没有频道“{channel}”的视频。")

    def _refresh_channel_combo(self) -> None:
        current = str(self.channel_selection_combo.currentData() or "")
        channels = sorted({
            str((item.get("video") or {}).get("channel") or "").strip()
            for item in self._items
            if isinstance(item.get("video"), dict) and str((item.get("video") or {}).get("channel") or "").strip()
        })
        self.channel_selection_combo.blockSignals(True)
        self.channel_selection_combo.clear()
        self.channel_selection_combo.addItem("选择频道", "")
        for channel in channels:
            self.channel_selection_combo.addItem(channel, channel)
        index = self.channel_selection_combo.findData(current)
        if index >= 0:
            self.channel_selection_combo.setCurrentIndex(index)
        self.channel_selection_combo.blockSignals(False)

    def _update_selected_count(self) -> None:
        self.selected_count_label.setText(f"已选 {len(self._selected_rows())} 条")

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)


class SettingsHubPage(QWidget):
    """Non-modal entry point for existing configuration dialogs and connection settings."""

    def __init__(
        self,
        *,
        open_youtube_login: Callable[[], None],
        open_asr: Callable[[], None],
        open_llm: Callable[[], None],
        open_connection_settings: Callable[[], None],
        open_matching_config: Callable[[], None],
        check_updates: Callable[[], None],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        root = QVBoxLayout(self)
        root.setContentsMargins(26, 24, 26, 26)
        root.setSpacing(14)
        root.addWidget(_section_title("设置", "将登录、代理、ASR、语言模型和匹配服务配置集中管理，避免干扰处理流程。"))

        cards = QGridLayout()
        cards.setHorizontalSpacing(12)
        cards.setVerticalSpacing(12)
        items = (
            ("YouTube 登录状态", "打开内置浏览器登录并同步 Cookie，用于需要授权的字幕与下载。", "管理登录", open_youtube_login),
            ("下载代理", "设置 YouTube 采集和音频下载使用的代理地址。", "设置代理", open_connection_settings),
            ("ASR 配置", "配置多个语音识别服务，并按既有熔断策略自动兜底。", "打开 ASR 配置", open_asr),
            ("语言模型", "配置文本纠偏、封面检测等能力复用的语言模型。", "打开语言模型", open_llm),
            ("匹配服务", "设置匹配服务地址、账号和密码，并在提交匹配前连接。", "设置匹配服务", open_matching_config),
            ("软件更新", "检查 GitHub 发布的新版本；下载、校验、替换和重启将自动完成。", "检查更新", check_updates),
        )
        for index, (title, hint, action, callback) in enumerate(items):
            card = QFrame()
            card.setObjectName("SettingsCard")
            layout = QVBoxLayout(card)
            layout.setContentsMargins(16, 14, 16, 14)
            heading = QLabel(title)
            heading.setObjectName("QuickActionTitle")
            body = QLabel(hint)
            body.setObjectName("PageHint")
            body.setWordWrap(True)
            button = QPushButton(action)
            button.setProperty("secondary", True)
            button.clicked.connect(callback)
            layout.addWidget(heading)
            layout.addWidget(body, 1)
            layout.addWidget(button, 0, Qt.AlignmentFlag.AlignLeft)
            cards.addWidget(card, index // 2, index % 2)
        root.addLayout(cards)
        root.addStretch(1)
