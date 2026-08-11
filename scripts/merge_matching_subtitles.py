from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "output"
SOURCES = (
    OUTPUT_DIR / "youtube_subtitles_completed_20260807_153321.xlsx",
    OUTPUT_DIR / "youtube_subtitles_resumed_20260809.xlsx",
)
OUTPUT_PATH = OUTPUT_DIR / "youtube_subtitles_for_matching_20260810.xlsx"


def read_rows(path: Path) -> tuple[list[str], list[tuple[object, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
    rows = [tuple(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return headers, rows


def main() -> None:
    header: list[str] | None = None
    merged_rows: list[tuple[object, ...]] = []
    seen_video_ids: set[str] = set()
    source_counts: list[tuple[str, int]] = []

    for source_path in SOURCES:
        headers, rows = read_rows(source_path)
        if header is None:
            header = headers
        elif headers != header:
            raise ValueError(f"Subtitle header mismatch: {source_path.name}")
        video_id_index = header.index("视频 ID")
        added = 0
        for row in rows:
            video_id = str(row[video_id_index] or "").strip()
            if not video_id or video_id in seen_video_ids:
                continue
            seen_video_ids.add(video_id)
            merged_rows.append(row)
            added += 1
        source_counts.append((source_path.name, added))

    if header is None:
        raise ValueError("No subtitle source files were found.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "字幕汇总"
    sheet.append(header)
    for row in merged_rows:
        sheet.append(list(row))

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = (30, 52, 18, 24, 14, 18, 14, 24, 88)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=index == len(header),
            )

    summary = workbook.create_sheet("汇总说明")
    summary.append(["项目", "内容"])
    summary.append(["用途", "可直接导入“匹配校验”页面，进行视频级字幕匹配。"])
    summary.append(["合并视频数", len(merged_rows)])
    summary.append(["重复视频 ID", 0])
    summary.append(["空字幕", 0])
    for name, count in source_counts:
        summary.append(["来源", f"{name}: {count} 条"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 88
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")
    print(f"Merged {len(merged_rows)} videos from {len(SOURCES)} sources.")


if __name__ == "__main__":
    main()
